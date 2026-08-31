from fastapi import FastAPI, HTTPException, Body, Depends, UploadFile, File, Form, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from fastapi.concurrency import run_in_threadpool
from pydantic import BaseModel
from typing import Optional, List, Any, TypedDict, Literal
import httpx
import os
import json
import psycopg2
import psycopg2.extras
from contextlib import contextmanager
from dotenv import load_dotenv

# ── LangGraph — built-in graph engine (no external package needed) ────────────
# Implements the same StateGraph API as the langgraph package.
# If the real package is installed it takes priority; otherwise this runs.
try:
    from langgraph.graph import StateGraph as _LGStateGraph, END as _LGEND
    StateGraph = _LGStateGraph
    END = _LGEND
    LANGGRAPH_AVAILABLE = True
    print("✅ Using installed langgraph package")
except ImportError:
    # ── Pure-Python StateGraph implementation ─────────────────────────────────
    # Supports: add_node, add_edge, add_conditional_edges, set_entry_point,
    #           compile() → graph with invoke(state) method.
    # Retry-safe: invoke catches node exceptions and continues with current state.
    _GRAPH_END = "__END__"
    END = _GRAPH_END

    class _CompiledGraph:
        def __init__(self, nodes, edges, cond_edges, entry):
            self._nodes = nodes          # {name: fn}
            self._edges = edges          # {from: to}
            self._cond = cond_edges      # {from: (condition_fn, {result: target})}
            self._entry = entry

        def invoke(self, state: dict, config=None) -> dict:
            current = self._entry
            visited = []
            max_steps = 30  # safety limit
            while current and current != _GRAPH_END and len(visited) < max_steps:
                visited.append(current)
                fn = self._nodes.get(current)
                if fn:
                    try:
                        result = fn(state)
                        if isinstance(result, dict):
                            state = {**state, **result}
                    except Exception as e:
                        state = {**state, "_node_error": str(e)}
                # Determine next node
                if current in self._cond:
                    cond_fn, mapping = self._cond[current]
                    try:
                        key = cond_fn(state)
                    except Exception:
                        key = list(mapping.values())[-1]  # fallback to last
                    current = mapping.get(key, _GRAPH_END)
                elif current in self._edges:
                    current = self._edges[current]
                else:
                    current = _GRAPH_END
            return state

    class StateGraph:
        """Minimal StateGraph compatible with the langgraph API."""
        def __init__(self, schema=None):
            self._nodes = {}
            self._edges = {}
            self._cond_edges = {}
            self._entry = None

        def add_node(self, name: str, fn):
            self._nodes[name] = fn
            return self

        def add_edge(self, from_node: str, to_node: str):
            self._edges[from_node] = to_node
            return self

        def add_conditional_edges(self, from_node: str, condition_fn, mapping: dict):
            self._cond_edges[from_node] = (condition_fn, mapping)
            return self

        def set_entry_point(self, name: str):
            self._entry = name
            return self

        def compile(self):
            return _CompiledGraph(
                self._nodes, self._edges, self._cond_edges, self._entry)

    LANGGRAPH_AVAILABLE = True  # built-in engine counts as available
    print("✅ Using built-in StateGraph engine (no langgraph package needed)")

from pathlib import Path
load_dotenv(Path(__file__).parent / ".env")

# Adobe IMS authentication (OAuth Authorization Code flow, secret kept server-side).
# Imported AFTER load_dotenv so the module reads IMS_* env vars at import time.
from ims_auth import router as ims_router, get_current_user, require_persona, classify_persona

app = FastAPI(title="Nexus Backend", version="1.0.0")


# ══════════════════════════════════════════════════════════════════════════════
# LANGGRAPH AGENTIC SYSTEM
# All 8 agents implemented as StateGraph workflows with:
#   - Multi-step reasoning nodes
#   - Conditional retry edges (max 2 retries)
#   - Quality judge nodes per agent
#   - Shared context retrieval
# ══════════════════════════════════════════════════════════════════════════════

# ── Agent package imports ─────────────────────────────────────────────────────
import sys, os as _os
_agents_path = _os.path.join(_os.path.dirname(__file__), "agents")
if _agents_path not in sys.path:
    sys.path.insert(0, _os.path.dirname(__file__))

try:
    from agents import (
        GRAPHS as _AGENT_GRAPHS,
        call_socratic,
        call_session_evaluator,
        call_study_notes,
        call_flashcards,
    )
    from agents.rag        import run_rag
    from agents.curriculum import run_curriculum
    from agents.capstone   import run_capstone
    from agents.practice   import run_practice, run_validate_understanding
    from agents.reasoning  import run_reasoning, stream_reasoning
    from agents.crossskill import run_crossskill
    from agents.study_aid  import run_study_aid
    _AGENTS_LOADED = True
    print("✅ Loaded agents/ package (separate files)")
except Exception as _agents_err:
    _AGENTS_LOADED = False
    print(f"⚠ agents/ package not loaded ({_agents_err}), using inline graphs")

# ── Shared LLM call utilities ─────────────────────────────────────────────────
# ── LLM helpers — imported from agents/config.py (single source of truth) ────
try:
    from agents.config import (
        OPENAI_URL, OPENAI_MODEL,
        GROQ_URL, GROQ_MODEL, ANTHROPIC_URL, ANTHROPIC_MODEL,
        groq_call as _groq_call,
        anthropic_call as _anthropic_call,
        openai_call as _openai_call,
        llm_call as _llm_call,
        get_db_url, PRODUCT_DISTINCTIONS,
    )
except ImportError:
    # Fallback if agents/ package not available
    OPENAI_URL      = os.getenv("OPENAI_URL",      "https://api.openai.com/v1/chat/completions")
    OPENAI_MODEL    = os.getenv("OPENAI_MODEL",    "gpt-4o-mini")
    GROQ_URL        = os.getenv("GROQ_URL",        "https://api.groq.com/openai/v1/chat/completions")
    GROQ_MODEL      = os.getenv("GROQ_MODEL",      "openai/gpt-oss-20b")
    ANTHROPIC_URL   = os.getenv("ANTHROPIC_URL",   "https://api.anthropic.com/v1/messages")
    ANTHROPIC_MODEL = os.getenv("ANTHROPIC_MODEL", "claude-sonnet-4-6")
    PRODUCT_DISTINCTIONS = ""

    def _openai_call(messages, system, max_tokens=600):
        key = os.getenv("OPENAI_API_KEY","")
        if not key: return "[OPENAI_API_KEY not set]"
        import requests as _req
        r = _req.post(OPENAI_URL,
            headers={"Authorization":f"Bearer {key}","Content-Type":"application/json"},
            json={"model":OPENAI_MODEL,"max_tokens":max_tokens,
                  "messages":[{"role":"system","content":system}]+messages},
            timeout=30)
        r.raise_for_status()
        return r.json()["choices"][0]["message"]["content"]

    def _groq_call(messages, system, max_tokens=600):
        key = os.getenv("GROQ_API_KEY","")
        if not key: return "[GROQ_API_KEY not set]"
        import requests as _req
        r = _req.post(GROQ_URL,
            headers={"Authorization":f"Bearer {key}","Content-Type":"application/json"},
            json={"model":GROQ_MODEL,"max_tokens":max_tokens,
                  "include_reasoning":False,
                  "messages":[{"role":"system","content":system}]+messages},
            timeout=30)
        r.raise_for_status()
        return r.json()["choices"][0]["message"]["content"]

    def _anthropic_call(messages, system, max_tokens=400):
        key = os.getenv("ANTHROPIC_API_KEY","")
        if not key: return "[ANTHROPIC_API_KEY not set]"
        import requests as _req
        r = _req.post(ANTHROPIC_URL,
            headers={"x-api-key":key,"anthropic-version":"2023-06-01","Content-Type":"application/json"},
            json={"model":ANTHROPIC_MODEL,"max_tokens":max_tokens,"system":system,
                  "messages":[m for m in messages if m.get("role") in ("user","assistant")]},
            timeout=30)
        r.raise_for_status()
        return r.json()["content"][0]["text"]

    def _fallback_input_guardrail(messages):
        """Standalone reimplementation of agents.config.check_input_guardrail
        for this fallback path — imports guardrails.policy directly rather
        than through agents.config (which is exactly the import that failed
        to get us into this branch), so this guardrail check can never be
        silently skipped just because the agents/ package failed to load.
        Returns a refusal string if blocked, else None. Never raises."""
        try:
            from guardrails import policy as _gr_policy
            user_text = " ".join(m.get("content","") for m in messages if m.get("role")=="user")
            blocked = _gr_policy.is_injection(user_text) or _gr_policy.is_unsafe(user_text)
            if blocked:
                return ("I can't help with that request. Let's keep this focused on your "
                        "Adobe Experience Platform learning — what topic would you like to explore?")
        except Exception:
            pass  # guardrail must never break generation
        return None

    def _llm_call(messages, system, max_tokens=600, timeout=30, prefer="openai"):
        """OpenAI→Groq→Anthropic failover (mirrors agents.config.llm_call,
        including the input guardrail check — see _fallback_input_guardrail)."""
        blocked = _fallback_input_guardrail(messages)
        if blocked:
            return blocked
        base = ["openai","groq","anthropic"]
        order = ([prefer] + [p for p in base if p != prefer]) if prefer in base else base
        impls = {"openai": _openai_call, "groq": _groq_call, "anthropic": _anthropic_call}
        errors = []
        for provider in order:
            try:
                return impls[provider](messages, system, max_tokens=max_tokens)
            except Exception as e:
                errors.append(f"{provider}: {e}")
        raise RuntimeError("All LLM providers failed → " + " | ".join(errors))

    def get_db_url():
        return os.getenv("DATABASE_URL","")

# ── Compile graphs ────────────────────────────────────────────────────────────
# Prefer agents/ package graphs; fall back to inline graphs if not loaded.
if _AGENTS_LOADED:
    _GRAPHS = _AGENT_GRAPHS
    print(f"✅ Nexus agents: {len(_GRAPHS)}/6 graphs loaded from agents/ package")
else:
    _GRAPHS = {}
    if LANGGRAPH_AVAILABLE:
        try:
            print("ℹ agents/ package not available — agent endpoints will return errors")
        except Exception as e:
            print(f"⚠ Graph compilation error: {e}")
            _GRAPHS = {}

# ── Learner context helper for agent routes ───────────────────────────────────
def _build_learner_context_dict(profile: dict) -> str:
    """Convert profile dict to a learner context string for agent system prompts."""
    parts = []
    if profile.get("name"):       parts.append(f"Name: {profile['name']}")
    if profile.get("role"):       parts.append(f"Role: {profile['role']}")
    if profile.get("team"):       parts.append(f"Team: {profile['team']}")
    if profile.get("track"):      parts.append(f"Current track: {profile['track'].upper()}")
    if profile.get("module"):     parts.append(f"Current module: {profile['module']}")
    if profile.get("confidence"): parts.append(f"Confidence: {profile['confidence']:.0%}")
    return "\n".join(parts)

# ── Pydantic request model for all agent endpoints ────────────────────────────
class AgentRequest(BaseModel):
    messages: List[dict] = []   # optional: flashcard/notes callers send no messages
    profile: dict = {}
    track: str = "rtcdp"
    module: str = ""
    extra: dict = {}
    system: str = ""          # system prompt for generic /api/agent proxy
    max_tokens: int = 1000
    temperature: float = 0.7
    agent_name: str = "Agent"
    prefer_groq: bool = True

def _run_graph(agent_name: str, body: AgentRequest) -> str:
    """Run a LangGraph agent or fall back to single-call mode."""
    graph = _GRAPHS.get(agent_name)
    if not graph:
        return f"[{agent_name} agent not available — LangGraph not installed or graph failed to compile]"
    initial_state: AgentState = {
        "messages": body.messages,
        "profile": body.profile,
        "track": body.track,
        "module": body.module,
        "docs": [],
        "response": "",
        "quality_ok": False,
        "intent": "",
        "retries": 0,
        "extra": body.extra,
    }
    result = graph.invoke(initial_state)
    return result.get("response", "[No response generated]")

# ── Agent API endpoints ───────────────────────────────────────────────────────
# Routes delegate to agents/ package functions when loaded,
# otherwise fall back to inline _run_graph().

@app.post("/api/agents/socratic")
def agent_socratic(body: AgentRequest):
    """Socratic agent — Anthropic primary, Groq fallback."""
    if _AGENTS_LOADED:
        learner_ctx = _build_learner_context_dict(body.profile)
        result = call_socratic(body.messages, learner_context=learner_ctx)
        return {**result, "graph_used": False}
    return {"response": _run_graph("socratic", body),
            "graph_used": "socratic" in _GRAPHS,
            "meta": {"type":"llm","name":"socratic","steps_executed":1}}

# ── DB-backed rate limiter (per authenticated user) ────────────────────────────
# Sliding window: max REASONING_RATE_LIMIT requests per REASONING_RATE_WINDOW s.
# Backed by Postgres (not an in-memory deque) so the limit is enforced correctly
# no matter how many worker processes or instances are running — an in-memory
# counter is invisible across processes, so a learner hitting different workers
# behind a load balancer could exceed the "global" limit by however many workers
# exist. This reuses the same DATABASE_URL every other table in this file already
# depends on, rather than introducing a new piece of infrastructure (e.g. Redis)
# for a single counter.
REASONING_RATE_LIMIT  = int(os.getenv("REASONING_RATE_LIMIT",  "30"))
REASONING_RATE_WINDOW = int(os.getenv("REASONING_RATE_WINDOW", "60"))
# Raw messages fetched here are what reasoning.py's rolling window can see at
# all — anything cut off before this point is gone with no trace. reasoning.py
# itself only sends the last REASONING_HISTORY_WINDOW (default 12) of these to
# the model verbatim and summarizes the rest, so this cap can safely be larger
# than that window without inflating per-turn token cost.
MAX_HISTORY_MESSAGES  = int(os.getenv("REASONING_MAX_HISTORY", "60"))

_rate_limit_table_ready = False

def _ensure_rate_limit_table():
    try:
        conn = psycopg2.connect(DATABASE_URL)
        conn.autocommit = True
        with conn.cursor() as c:
            c.execute("""
                CREATE TABLE IF NOT EXISTS reasoning_rate_hits (
                    id SERIAL PRIMARY KEY,
                    rate_key VARCHAR(255) NOT NULL,
                    hit_at TIMESTAMP NOT NULL DEFAULT NOW()
                )
            """)
            c.execute("CREATE INDEX IF NOT EXISTS idx_rate_hits_key_time ON reasoning_rate_hits (rate_key, hit_at)")
        conn.close()
    except Exception:
        pass

def _rate_limit(key: str, limit: int = None, window: int = None):
    """Raise 429 if `key` exceeded the window; fails OPEN (allows the request) if
    the DB is unreachable, matching this codebase's rule that a DB blip must never
    block a legitimate learner turn — the trade-off is rate limiting is briefly
    unenforced during an outage rather than blocking all traffic.

    `limit`/`window` default to the reasoning-agent constants (unchanged behavior
    for existing callers); other endpoints (e.g. Study Aid) pass their own so each
    gets an independently-sized budget while sharing the same DB-backed mechanism.
    Callers should prefix `key` (e.g. "flashcard:<email>") so different endpoints'
    counters never collide in the shared reasoning_rate_hits table.

    Pruning and the count+insert decision are two separate transactions: pruning
    always commits regardless of the outcome below, so a chronically-blocked key
    (a caller consistently over the limit) still has its old hit rows cleaned up
    every call instead of accumulating forever."""
    limit = REASONING_RATE_LIMIT if limit is None else limit
    window = REASONING_RATE_WINDOW if window is None else window
    global _rate_limit_table_ready
    if not _rate_limit_table_ready:
        _ensure_rate_limit_table()
        _rate_limit_table_ready = True
    try:
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "DELETE FROM reasoning_rate_hits WHERE rate_key=%s AND hit_at < NOW() - make_interval(secs => %s)",
                    (key, window))
    except Exception:
        return  # DB unreachable — fail open, skip enforcement for this call
    try:
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT COUNT(*) FROM reasoning_rate_hits WHERE rate_key=%s", (key,))
                count = cur.fetchone()[0]
                if count >= limit:
                    raise HTTPException(status_code=429,
                        detail=f"Rate limit exceeded ({limit}/{window}s). Please slow down.")
                cur.execute("INSERT INTO reasoning_rate_hits (rate_key) VALUES (%s)", (key,))
    except HTTPException:
        raise
    except Exception:
        return  # DB unreachable — fail open


# ── Study Aid endpoints: rate-limited but NOT auth-gated ───────────────────────
# Unlike Reasoning (which requires a real Adobe IMS session), Flash Cards/Study
# Notes are also used by demo personas that carry no session cookie at all —
# hard-requiring auth here would silently break that flow. Instead these
# endpoints stay open but get their own DB-backed rate-limit budget, keyed by
# session email when logged in, else by client IP, so an anonymous/demo caller
# still can't spam the Groq quota unbounded.
STUDY_AID_RATE_LIMIT  = int(os.getenv("STUDY_AID_RATE_LIMIT",  "20"))
STUDY_AID_RATE_WINDOW = int(os.getenv("STUDY_AID_RATE_WINDOW", "60"))


def _optional_user(request: Request) -> dict:
    """Like get_current_user but never raises 401 — returns {} with no session.
    Used where a feature is intentionally open to unauthenticated/demo callers
    but still needs a stable per-caller identity for rate limiting."""
    from ims_auth import SESSION_COOKIE, _verify as _verify_session
    payload = _verify_session(request.cookies.get(SESSION_COOKIE, ""))
    return payload or {}


def _rate_limit_key(request: Request, user: dict, prefix: str) -> str:
    email = (user.get("email") or "").lower()
    ident = email if email else f"ip:{request.client.host if request.client else 'unknown'}"
    return f"{prefix}:{ident}"

@app.post("/api/agents/reasoning")
def agent_reasoning(body: AgentRequest, user: dict = Depends(get_current_user)):
    """Reasoning agent — tool-calling agentic pipeline: load context → classify
    intent → retrieve → generate (6 tools) → quality judge (retry ≤2).

    Requires a valid IMS session. The learner identity used for all DB reads and
    writes is taken from the SERVER-SIDE session, never from the request body —
    this prevents a caller from impersonating another learner."""
    email = (user.get("email") or "").lower()
    _rate_limit(email or "anon")

    # Server-authoritative identity: overwrite client-supplied name/email with the
    # authenticated session's, so tools that key off profile.name can't be spoofed.
    session_profile = user.get("profile", {}) or {}
    safe_profile = dict(body.profile or {})
    safe_profile["name"]  = session_profile.get("name") or user.get("name") or ""
    safe_profile["email"] = email

    # Bound conversation history so cost/latency/context stay stable on long chats.
    msgs = (body.messages or [])[-MAX_HISTORY_MESSAGES:]

    if _AGENTS_LOADED:
        ctx = {
            "messages":        msgs,
            "profile":         safe_profile,
            "learner_context": _build_learner_context_dict(safe_profile),
            "track":           body.track,
            "module":          body.module,
            "extra":           body.extra or {},
        }
        question = msgs[-1]["content"] if msgs else ""
        result = run_reasoning(question, ctx, graph=_GRAPHS.get("reasoning"))
        tool_calls = result.get("tool_calls", []) or []
        return {
            "response":      result.get("response", "[No response generated]"),
            "graph_used":    True,
            "quality_ok":    result.get("quality_ok", False),
            "quality_score": result.get("quality_score", 0),
            "quality_issue": result.get("quality_issue"),
            "intent":        result.get("intent", ""),
            "retries":       result.get("retries", 0),
            "degraded":      result.get("degraded", False),
            "grounded":      result.get("grounded"),
            "request_id":    result.get("request_id"),
            "tool_calls":    [{"tool": tc.get("tool", ""), "blocked": bool(tc.get("blocked"))} for tc in tool_calls],
            "tool_count":    len(tool_calls),
            "meta":          result.get("meta", {}),
        }
    return {"response": _run_graph("reasoning", body), "graph_used": "reasoning" in _GRAPHS}


@app.post("/api/agents/reasoning/stream")
def agent_reasoning_stream(body: AgentRequest, user: dict = Depends(get_current_user)):
    """Streaming (SSE) variant of the reasoning agent.

    Same authentication, server-authoritative identity, rate limiting, and history
    bounding as /api/agents/reasoning. The full vetted pipeline runs first (so the
    quality judge, retry, and guardrails all still apply); the APPROVED answer is
    then streamed word-by-word as `text/event-stream`. Each SSE frame is a JSON
    object: {"type":"status"|"token"|"done"|"error", ...}."""
    email = (user.get("email") or "").lower()
    _rate_limit(email or "anon")

    session_profile = user.get("profile", {}) or {}
    safe_profile = dict(body.profile or {})
    safe_profile["name"]  = session_profile.get("name") or user.get("name") or ""
    safe_profile["email"] = email

    msgs = (body.messages or [])[-MAX_HISTORY_MESSAGES:]

    if not _AGENTS_LOADED:
        raise HTTPException(status_code=503, detail="Reasoning agent not available.")

    ctx = {
        "messages":        msgs,
        "profile":         safe_profile,
        "learner_context": _build_learner_context_dict(safe_profile),
        "track":           body.track,
        "module":          body.module,
        "extra":           body.extra or {},
    }
    question = msgs[-1]["content"] if msgs else ""

    def event_stream():
        try:
            for ev in stream_reasoning(question, ctx, graph=_GRAPHS.get("reasoning")):
                yield f"data: {json.dumps(ev)}\n\n"
        except Exception as e:
            yield f"data: {json.dumps({'type': 'error', 'detail': str(e)})}\n\n"

    return StreamingResponse(
        event_stream(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no",   # disable proxy buffering so tokens flush live
        },
    )

@app.post("/api/agents/curriculum")
def agent_curriculum(body: AgentRequest):
    """Curriculum agent — sequencing + NBA + quiz + test-out."""
    if _AGENTS_LOADED:
        request_type = body.extra.get("request_type", "quiz")
        context = {
            "learner_name":       body.profile.get("name",""),
            "track":              body.track,
            "topic":              body.module or body.extra.get("topic",""),
            "module_title":       body.extra.get("module_title",""),
            "overall_confidence": body.profile.get("confidence", 0.5),
            "modules":            body.extra.get("modules", []),
            "done_modules":       body.extra.get("done_modules", []),
            "conf_scores":        body.extra.get("conf_scores", {}),
        }
        result = run_curriculum(request_type, context, graph=_GRAPHS.get("curriculum"))
        return {"result": result["result"], "meta": result["meta"],
                "quality_ok": result.get("quality_ok", True)}
    return {"response": _run_graph("curriculum", body), "graph_used": "curriculum" in _GRAPHS}

def _resolve_real_role_manager_tenure(email: str):
    """Look up a learner's authoritative role/manager/tenure from the HR
    directory (employee_directory), keyed by their session email — this can't
    be spoofed by the client, unlike trusting body.profile.role directly.
    Returns (role, manager_name, tenure_months) — any of which may be None."""
    if not email:
        return None, None, None
    try:
        with get_db() as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute(
                    "SELECT role, manager_name, doj FROM employee_directory WHERE email=%s AND is_active=TRUE",
                    (email.lower(),))
                row = cur.fetchone()
        if not row:
            return None, None, None
        tenure_months = None
        if row.get("doj"):
            from datetime import date
            doj = row["doj"]
            today = date.today()
            tenure_months = (today.year - doj.year) * 12 + (today.month - doj.month)
        return row.get("role"), row.get("manager_name"), tenure_months
    except Exception as e:
        print(f"[advisor] directory lookup error: {e}")
        return None, None, None


@app.post("/api/agents/advisor")
def agent_advisor(body: AgentRequest, request: Request):
    """AI Advisor — cross-skilling guidance (uses crossskill agent).

    Not hard-auth-gated (demo personas use this too, same as flashcards), but
    when a real IMS session IS present, role/manager/tenure are resolved
    server-side from the HR directory rather than trusted from the client —
    those are the fields that ground the recommendation in real org data, so
    they shouldn't be spoofable. Falls back to whatever the client sent
    (demo/no-directory-match case) so the feature keeps working either way."""
    if _AGENTS_LOADED:
        user = _optional_user(request)
        email = (user.get("email") or "").lower()
        real_role, real_manager, tenure_months = _resolve_real_role_manager_tenure(email)
        context = {
            "learner_name":      body.profile.get("name",""),
            "learner_role":      real_role or body.profile.get("role",""),
            "manager_name":      real_manager or body.extra.get("manager_name",""),
            "tenure_months":     tenure_months,
            "completed_tracks":  body.extra.get("completed_tracks",[]),
            "enrolled_tracks":   body.extra.get("enrolled_tracks",[]),
            "skills":            body.extra.get("skills",{}),
            "certifications":    body.extra.get("certifications",[]),
            "bandwidth_pct":     body.profile.get("bw", 80),
            "team_name":         body.profile.get("team",""),
            # The learner's own active track (e.g. 'da', 'rtcdp') is a reliable
            # grounding fallback for the skill-map journey when they aren't in the
            # HR directory and their job title doesn't match a journey row — the
            # track code maps 1:1 to a role_learning_journey role (da→AEP-DA, …).
            "active_track":      body.track or body.profile.get("active_track","") or body.profile.get("track",""),
            # Tracks already shown to this learner in this session (from the
            # "Show me something else" flow) — excluded so retrying doesn't
            # just re-serve the same deterministic top pick.
            "exclude_tracks":    body.extra.get("exclude_tracks", []),
        }
        # Two modes on one endpoint: an empty `messages` list is the initial
        # recommendation-card fetch (run_crossskill, deterministic scoring +
        # one structured-JSON generation); a non-empty list is a follow-up chat
        # turn, which uses tool-calling (run_crossskill_chat) so the model can
        # pull the SAME real ranking/track/learning-path data on demand instead
        # of a static context block getting confused by unrelated content.
        if body.messages:
            from agents.crossskill import run_crossskill_chat
            chat_result = run_crossskill_chat(
                [{"role": m.get("role", "user"), "content": m.get("content", "")} for m in body.messages],
                context,
            )
            return {"result": chat_result, "response": chat_result.get("response", "")}
        result = run_crossskill(context, graph=_GRAPHS.get("crossskill"))
        return {"result": result["guidance"], "meta": result["meta"],
                "role_journey_matched": result.get("role_journey_matched", False),
                "manager_track_focus": result.get("manager_track_focus")}
    return {"response": _run_graph("advisor", body), "graph_used": "advisor" in _GRAPHS}

@app.post("/api/agents/capstone")
def agent_capstone(body: AgentRequest):
    """Capstone agent — industry vertical + use cases + PPT spec."""
    if _AGENTS_LOADED:
        context = {
            "learner_name":       body.profile.get("name",""),
            "track":              body.track,
            "team_context":       body.extra.get("team_context", body.profile.get("team","")),
            "done_module_titles": body.extra.get("done_module_titles",[]),
            "weakest_modules":    body.extra.get("weakest_modules",[]),
        }
        result = run_capstone(context, graph=_GRAPHS.get("capstone"))
        return {**result}
    return {"response": _run_graph("capstone", body), "graph_used": "capstone" in _GRAPHS}

@app.post("/api/agents/practice")
def agent_practice(body: AgentRequest):
    """Practice agent — RAG-grounded scenario. Uses a learner-typed topic when
    given (dynamic, no predefined list); falls back to weakest-module
    targeting when no topic is supplied (original behaviour, unchanged)."""
    if _AGENTS_LOADED:
        context = {
            "learner_name":   body.profile.get("name",""),
            "track":          body.track,
            "topic":          body.extra.get("topic",""),
            "is_cross_skill": bool(body.extra.get("is_cross_skill", False)),
            "conf_scores":    body.extra.get("conf_scores",{}),
            "modules":        body.extra.get("modules",[]),
            "done_modules":   body.extra.get("done_modules",[]),
            "module_id":      body.extra.get("module_id"),
            "module_title":   body.module or body.extra.get("module_title",""),
        }
        result = run_practice(context, graph=_GRAPHS.get("practice"))
        return {"scenario": result["scenario"], "meta": result["meta"],
                "weak_module_title": result.get("weak_module_title","")}
    return {"response": _run_graph("practice", body), "graph_used": "practice" in _GRAPHS}

@app.post("/api/agents/practice/validate")
def agent_practice_validate(body: AgentRequest):
    """Validate My Understanding — RAG-grounded comparison of the learner's own
    explanation against the scenario they were just shown. Returns a 0-100
    confidence score plus scenario-specific guidance (never generic advice)."""
    if not _AGENTS_LOADED:
        raise HTTPException(
            status_code=503,
            detail="Practice validation agent is not loaded on this backend.",
        )
    context = {
        "scenario":              body.extra.get("scenario", {}),
        "learner_understanding": body.extra.get("learner_understanding", ""),
        "topic":                 body.extra.get("topic", ""),
    }
    result = run_validate_understanding(context)
    return result

@app.post("/api/agents/flashcard")
def agent_flashcard(body: AgentRequest, request: Request, user: dict = Depends(_optional_user)):
    """Flashcard agent — 8 curriculum-grounded reasoning cards via the Study Aid
    agent. Always calls run_study_aid() directly rather than branching on
    "study_aid" in _GRAPHS — that check only gated whether a compiled LangGraph
    object existed, but run_study_aid() already runs its own sequential fallback
    when graph=None, so the branch was pure risk: the alternative it fell back to
    (call_flashcards) returns cards shaped {"front","back"} while every consumer
    (this endpoint's own contract, the frontend, _extract_cards) expects
    {"q","a"} — that mismatch meant every card would render blank if this branch
    were ever taken.

    Rate-limited (not auth-gated — demo personas use this too) so an anonymous
    caller can't spam the shared Groq quota unbounded."""
    _rate_limit(_rate_limit_key(request, user, "flashcard"), STUDY_AID_RATE_LIMIT, STUDY_AID_RATE_WINDOW)
    if _AGENTS_LOADED:
        module = body.module or body.extra.get("topic", "AEP")
        # Confidence drives flashcard difficulty. Prefer the authoritative DB
        # score for this learner+module (can't be forged by the client); fall back
        # to a client-supplied value, else None (neutral difficulty).
        confidence = body.extra.get("confidence")
        email = (user.get("email") or "").lower()
        if email and module:
            try:
                db_conf = db_get_confidence(user.get("name") or "", module)
                if db_conf is not None:
                    confidence = db_conf
            except Exception:
                pass
        result = run_study_aid(
            {
                "module":     module,
                "module_id":  body.extra.get("module_id"),
                "track":      body.track,
                "confidence": confidence,
            },
            graph=_GRAPHS.get("study_aid"),
        )
        return {"response": str(result["cards"]), "cards": result["cards"], "meta": result["meta"],
                "used_fallback": result.get("used_fallback", False)}
    return {"response": _run_graph("flashcard", body), "graph_used": "flashcard" in _GRAPHS}

@app.post("/api/agents/notes")
def agent_notes(body: AgentRequest):
    """Study notes — structured notes per topic."""
    if _AGENTS_LOADED:
        result = call_study_notes(
            body.module or body.extra.get("topic","AEP"),
            body.track
        )
        return {"notes": result["notes"], "meta": result["meta"]}
    return {"response": _run_graph("flashcard", body)}

@app.post("/api/agents/session-evaluate")
def agent_session_evaluate(body: dict = Body(...)):
    """Evaluate a completed Socratic session → quality scores + confidence delta."""
    if _AGENTS_LOADED:
        result = call_session_evaluator(
            body.get("messages",[]),
            body.get("topic","")
        )
        return result
    return {"scores": {}, "confidence_delta": 0}

@app.post("/api/agents/rag")
def agent_rag(body: AgentRequest):
    """RAG agent — 5-step retrieval pipeline."""
    if _AGENTS_LOADED:
        query  = body.messages[-1]["content"] if body.messages else body.extra.get("query","")
        result = run_rag(query, graph=_GRAPHS.get("rag"))
        return result
    return {"response": _run_graph("rag", body), "graph_used": "rag" in _GRAPHS}

@app.get("/api/agents/status")
def agent_status():
    """Check which LangGraph agents are compiled and ready, plus which LLM
    provider keys are configured (booleans only — never the actual keys) so
    the Admin Integrations tab can show real status instead of stale claims."""
    try:
        import importlib.metadata as _ilm
        engine = f"langgraph=={_ilm.version('langgraph')}"
    except Exception:
        engine = "built-in StateGraph"
    return {
        "langgraph_available": LANGGRAPH_AVAILABLE,
        "graphs_compiled": list(_GRAPHS.keys()),
        "agents_ready": len(_GRAPHS),
        "total_agents": 8,
        "engine": engine,
        # Fallback order every agent's llm_call()/call_with_tools() actually
        # uses (see agents/config.py) — OpenAI primary, Claude 2nd, Groq 3rd.
        # Tool-calling is OpenAI->Groq only (Anthropic's tool-use schema isn't
        # wire-compatible with the OpenAI tools/tool_calls format every
        # agent's tool definitions are written in).
        "provider_order": ["openai", "anthropic", "groq"],
        "provider_keys_configured": {
            "openai":    bool(os.getenv("OPENAI_API_KEY", "")),
            "anthropic": bool(os.getenv("ANTHROPIC_API_KEY", "")),
            "groq":      bool(os.getenv("GROQ_API_KEY", "")),
        },
    }

@app.on_event("startup")
def create_all_tables():
    """Create every table on startup so they exist before any endpoint is called."""
    try:
        conn = psycopg2.connect(DATABASE_URL)
        conn.autocommit = True
        with conn.cursor() as c:
            # ── LLM call logs ───────────────────────────────────────────
            c.execute("""CREATE TABLE IF NOT EXISTS llm_logs (
                id SERIAL PRIMARY KEY, agent_name VARCHAR(50), model VARCHAR(80),
                input_tokens INT DEFAULT 0, output_tokens INT DEFAULT 0,
                latency_ms FLOAT DEFAULT 0, success BOOLEAN DEFAULT TRUE,
                error TEXT, created_at TIMESTAMP DEFAULT NOW())""")
            # ── Telemetry / behaviour events ────────────────────────────
            c.execute("""CREATE TABLE IF NOT EXISTS telemetry (
                id SERIAL PRIMARY KEY, persona VARCHAR(120), event_type VARCHAR(80),
                module VARCHAR(255) DEFAULT '', detail TEXT DEFAULT '',
                created_at TIMESTAMP DEFAULT NOW())""")
            # ── Socratic guardrail logs ─────────────────────────────────
            c.execute("""CREATE TABLE IF NOT EXISTS guardrail_logs (
                id SERIAL PRIMARY KEY, word_count INT, has_one_question BOOLEAN,
                avoids_direct_answer BOOLEAN, score INT DEFAULT 0,
                issue VARCHAR(255), response_preview VARCHAR(200),
                agent_name VARCHAR(50) DEFAULT 'socratic',
                created_at TIMESTAMP DEFAULT NOW())""")
            c.execute("ALTER TABLE guardrail_logs ADD COLUMN IF NOT EXISTS agent_name VARCHAR(50) DEFAULT 'socratic'")
            # ── Session summaries ────────────────────────────────────────
            c.execute("""CREATE TABLE IF NOT EXISTS session_summaries (
                id SERIAL PRIMARY KEY, user_name VARCHAR(120), session_type VARCHAR(50),
                module VARCHAR(255), summary TEXT, created_at TIMESTAMP DEFAULT NOW())""")
            # ── Confidence scores ────────────────────────────────────────
            c.execute("""CREATE TABLE IF NOT EXISTS confidence_scores (
                id SERIAL PRIMARY KEY, user_name VARCHAR(120), module VARCHAR(255),
                score FLOAT, created_at TIMESTAMP DEFAULT NOW())""")
            # ── Bandwidth log ────────────────────────────────────────────
            c.execute("""CREATE TABLE IF NOT EXISTS bw_logs (
                id SERIAL PRIMARY KEY, persona VARCHAR(120), bw INT,
                created_at TIMESTAMP DEFAULT NOW())""")
            # ── Module progress ──────────────────────────────────────────
            c.execute("""CREATE TABLE IF NOT EXISTS user_module_progress (
                id SERIAL PRIMARY KEY, member_name VARCHAR(120) NOT NULL,
                manager VARCHAR(120), track VARCHAR(50) DEFAULT 'rtcdp',
                module_id INT, module_title VARCHAR(255),
                via VARCHAR(30) DEFAULT 'normal',
                completed_at TIMESTAMP DEFAULT NOW(),
                UNIQUE(member_name, module_id, track))""")
            # ── Test-outs ────────────────────────────────────────────────
            c.execute("""CREATE TABLE IF NOT EXISTS module_test_outs (
                id SERIAL PRIMARY KEY, member_name VARCHAR(120) NOT NULL,
                manager VARCHAR(120), track VARCHAR(50) DEFAULT 'rtcdp',
                module_id INT, module_title VARCHAR(255),
                score INT DEFAULT 0, passed BOOLEAN DEFAULT FALSE,
                created_at TIMESTAMP DEFAULT NOW())""")
            # ── Points ledger ────────────────────────────────────────────
            c.execute("""CREATE TABLE IF NOT EXISTS points_ledger (
                id SERIAL PRIMARY KEY, member_name VARCHAR(120) NOT NULL,
                manager VARCHAR(120), points INT DEFAULT 0,
                reason VARCHAR(255), created_at TIMESTAMP DEFAULT NOW())""")
            # ── Notifications ────────────────────────────────────────────
            c.execute("""CREATE TABLE IF NOT EXISTS notifications (
                id SERIAL PRIMARY KEY, member_name VARCHAR(120) NOT NULL,
                manager VARCHAR(120), type VARCHAR(50), title VARCHAR(255),
                message TEXT, is_read BOOLEAN DEFAULT FALSE,
                created_at TIMESTAMP DEFAULT NOW())""")
            # ── Skill assessments (CAT) ──────────────────────────────────
            c.execute("""CREATE TABLE IF NOT EXISTS skill_assessments (
                id SERIAL PRIMARY KEY, member_name VARCHAR(120) NOT NULL,
                manager VARCHAR(120), skill VARCHAR(100),
                level VARCHAR(30), theta FLOAT, assessed_at TIMESTAMP DEFAULT NOW())""")
            # ── Project allocations + updates ────────────────────────────
            c.execute("""CREATE TABLE IF NOT EXISTS project_allocations (
                id SERIAL PRIMARY KEY, member_name VARCHAR(120) NOT NULL,
                manager VARCHAR(120), project_id VARCHAR(50), project_name VARCHAR(255),
                project_type VARCHAR(100), industry VARCHAR(100), phase VARCHAR(50),
                stage VARCHAR(50), start_date DATE, end_date DATE,
                hrs_per_week NUMERIC DEFAULT 0, use_cases TEXT,
                solutions_used VARCHAR(255), product_features VARCHAR(255),
                data_sources VARCHAR(255), destinations VARCHAR(255),
                num_audiences INT DEFAULT 0, region VARCHAR(50), ticket_ids VARCHAR(255),
                health_status VARCHAR(50) DEFAULT 'On track', renewal VARCHAR(20) DEFAULT 'TBD',
                comments TEXT, project_notes TEXT,
                created_at TIMESTAMP DEFAULT NOW(), updated_at TIMESTAMP DEFAULT NOW())""")
            c.execute("""CREATE TABLE IF NOT EXISTS allocation_updates (
                id SERIAL PRIMARY KEY,
                allocation_id INTEGER REFERENCES project_allocations(id) ON DELETE CASCADE,
                member_name VARCHAR(120) NOT NULL, comment TEXT NOT NULL,
                health_status VARCHAR(50), created_at TIMESTAMP DEFAULT NOW(),
                billable_hours NUMERIC DEFAULT 0, week_of DATE)""")
            c.execute("ALTER TABLE allocation_updates ADD COLUMN IF NOT EXISTS billable_hours NUMERIC DEFAULT 0")
            c.execute("ALTER TABLE allocation_updates ADD COLUMN IF NOT EXISTS week_of DATE")
            # ── curriculum_topics — module content populated by seed script ──────
            c.execute("""CREATE TABLE IF NOT EXISTS curriculum_topics (
                id SERIAL PRIMARY KEY,
                module_id INTEGER NOT NULL,
                track VARCHAR(50) NOT NULL DEFAULT 'rtcdp',
                topic_order INTEGER NOT NULL,
                title VARCHAR(255) NOT NULL,
                objective TEXT,
                activity TEXT,
                output TEXT,
                checkpoint TEXT,
                video_title VARCHAR(255),
                video_duration VARCHAR(30),
                el_url VARCHAR(500),
                content TEXT,
                created_at TIMESTAMP DEFAULT NOW(),
                UNIQUE(module_id, track, topic_order)
            )""")
            # `content` above was reserved but never populated. It's now used as a
            # cache of the last successful live GitHub/Experience-League fetch in
            # get_topic_content() — live fetch stays the default on every request;
            # this is served only when a fresh fetch fails, so a transient network
            # blip or GitHub rate limit doesn't leave the learner with nothing.
            c.execute("ALTER TABLE curriculum_topics ADD COLUMN IF NOT EXISTS cached_video_url VARCHAR(500)")
            c.execute("ALTER TABLE curriculum_topics ADD COLUMN IF NOT EXISTS cached_fetch_method VARCHAR(120)")
            c.execute("ALTER TABLE curriculum_topics ADD COLUMN IF NOT EXISTS cached_at TIMESTAMP")
            # ── Community threads/replies — replaces the previous NJ_THREADS/
            # THREADS hardcoded frontend arrays (fabricated posts from named
            # people with frozen "2h ago" timestamps). `space` separates the
            # New Joiner cohort ('nj') from the experienced-team community
            # ('exp'), same 2 spaces the frontend already had as separate UIs.
            c.execute("""CREATE TABLE IF NOT EXISTS community_threads (
                id SERIAL PRIMARY KEY,
                space VARCHAR(20) NOT NULL,
                author_name VARCHAR(150) NOT NULL,
                title TEXT NOT NULL,
                tag VARCHAR(60),
                product VARCHAR(40),
                created_at TIMESTAMP DEFAULT NOW())""")
            c.execute("""CREATE TABLE IF NOT EXISTS community_replies (
                id SERIAL PRIMARY KEY,
                thread_id INTEGER NOT NULL REFERENCES community_threads(id) ON DELETE CASCADE,
                author_name VARCHAR(150) NOT NULL,
                text TEXT NOT NULL,
                created_at TIMESTAMP DEFAULT NOW())""")
            # ── Team community upgrade: manager-scoped feed with visibility,
            # a post body, and @mentions. Added additively so the existing
            # nj/exp space feeds keep working (visibility defaults to 'team').
            #   visibility: 'private' (author only) | 'team' (same manager) | 'public' (all teams)
            for _col, _ddl in [
                ("author_email",  "VARCHAR(150)"),
                ("manager_email", "VARCHAR(150)"),
                ("visibility",    "VARCHAR(20) DEFAULT 'team'"),
                ("body",          "TEXT"),
                ("mentions",      "JSONB DEFAULT '[]'"),
            ]:
                c.execute(f"ALTER TABLE community_threads ADD COLUMN IF NOT EXISTS {_col} {_ddl}")
            # notifications: add link back to the thread so a click can deep-link.
            c.execute("ALTER TABLE notifications ADD COLUMN IF NOT EXISTS thread_id INTEGER")
            c.execute("ALTER TABLE notifications ADD COLUMN IF NOT EXISTS actor VARCHAR(150)")
            # ── Kudos reactions — one 👍 per person per thread. Scoring rewards
            # reactions RECEIVED (usefulness) rather than raw post volume, so the
            # leaderboard can't be farmed by low-value posting.
            c.execute("""CREATE TABLE IF NOT EXISTS community_reactions (
                id SERIAL PRIMARY KEY,
                thread_id INTEGER NOT NULL REFERENCES community_threads(id) ON DELETE CASCADE,
                member_name VARCHAR(150) NOT NULL,
                created_at TIMESTAMP DEFAULT NOW(),
                UNIQUE(thread_id, member_name))""")
            # ── Release notes cache — replaces the hardcoded, frozen RELEASE_NOTES
            # frontend object. Populated by parsing each product's REAL release-notes
            # page from its AdobeDocs GitHub repo (see fetch_release_notes()); refetched
            # when stale rather than on a schedule (no cron in this app).
            c.execute("""CREATE TABLE IF NOT EXISTS release_notes_cache (
                id SERIAL PRIMARY KEY,
                product VARCHAR(30) NOT NULL,
                period VARCHAR(80),
                title TEXT NOT NULL,
                description TEXT,
                source_url VARCHAR(500),
                fetched_at TIMESTAMP DEFAULT NOW(),
                UNIQUE(product, title))""")
            # ── Per-track capstones — a learner can have several parallel
            # ("solitaire") learning tracks; each has its OWN capstone. The
            # onboarding/primary capstone stays on onboarding_requests
            # (capstone_completed); THIS table records completion of each
            # additional cross-skill track's capstone, so progress is tracked
            # per track rather than as one global flag.
            c.execute("""CREATE TABLE IF NOT EXISTS track_capstones (
                id SERIAL PRIMARY KEY,
                member_name VARCHAR(150) NOT NULL,
                track VARCHAR(50) NOT NULL,
                status VARCHAR(20) DEFAULT 'completed',
                score NUMERIC,
                completed_at TIMESTAMP DEFAULT NOW(),
                UNIQUE(member_name, track))""")
            # ── doc_embeddings — populated by build_embeddings_index.py ─────────
            c.execute("""CREATE TABLE IF NOT EXISTS doc_embeddings (
                id SERIAL PRIMARY KEY,
                repo VARCHAR(120),
                file_path VARCHAR(500),
                el_url VARCHAR(500),
                title VARCHAR(255),
                track VARCHAR(50),
                chunk_index INTEGER,
                chunk_text TEXT,
                embedding TEXT,
                created_at TIMESTAMP DEFAULT NOW()
            )""")
            # ── CF Utilization ────────────────────────────────────────────
            c.execute("""CREATE TABLE IF NOT EXISTS user_utilization (
                id SERIAL PRIMARY KEY, member_name VARCHAR(120) NOT NULL,
                manager VARCHAR(120), week_of DATE NOT NULL,
                billable_hours NUMERIC DEFAULT 0, non_billable_cf_hours NUMERIC DEFAULT 0,
                ramp_credit NUMERIC DEFAULT 0, working_hours NUMERIC DEFAULT 40,
                holiday_hours NUMERIC DEFAULT 0, loa_hours NUMERIC DEFAULT 0,
                cf_target NUMERIC DEFAULT 75,
                created_at TIMESTAMP DEFAULT NOW(), updated_at TIMESTAMP DEFAULT NOW(),
                UNIQUE(member_name, week_of))""")
            # ── Conversation history ─────────────────────────────────────
            c.execute("""CREATE TABLE IF NOT EXISTS conversation_messages (
                id SERIAL PRIMARY KEY, member_name VARCHAR(120), manager VARCHAR(120),
                module VARCHAR(255), mode VARCHAR(50), role VARCHAR(20), content TEXT,
                created_at TIMESTAMP DEFAULT NOW())""")
            # ── AI content cache ─────────────────────────────────────────
            c.execute("""CREATE TABLE IF NOT EXISTS generated_content_cache (
                id SERIAL PRIMARY KEY, cache_key VARCHAR(255) UNIQUE, agent_name VARCHAR(50),
                content TEXT, created_at TIMESTAMP DEFAULT NOW())""")
            # ── Onboarding + manager accounts ────────────────────────────
            c.execute("""CREATE TABLE IF NOT EXISTS onboarding_requests (
                id SERIAL PRIMARY KEY, name VARCHAR(120) NOT NULL,
                preferred_name VARCHAR(80), email VARCHAR(255) UNIQUE NOT NULL,
                joining_date DATE, role VARCHAR(100) DEFAULT '',
                team VARCHAR(100), manager VARCHAR(120),
                status VARCHAR(20) DEFAULT 'pending',
                password_hash VARCHAR(64), capstone_completed BOOLEAN DEFAULT FALSE,
                capstone_completed_at TIMESTAMP, capstone_started_at TIMESTAMP,
                active_track VARCHAR(50) DEFAULT 'rtcdp',
                enrolled_tracks TEXT DEFAULT '[]',   -- JSON array of track IDs the user is working on
                username VARCHAR(60), avatar_emoji VARCHAR(10), avatar_color VARCHAR(10),
                actioned_by VARCHAR(120), actioned_at TIMESTAMP,
                created_at TIMESTAMP DEFAULT NOW())""")
            # ── Backward-compat columns for onboarding_requests ──────────
            # These are already in the CREATE TABLE above for new installs;
            # the ALTER TABLE IF NOT EXISTS guards cover existing databases.
            c.execute("ALTER TABLE onboarding_requests ADD COLUMN IF NOT EXISTS enrolled_tracks TEXT DEFAULT '[]'")
            c.execute("ALTER TABLE onboarding_requests ADD COLUMN IF NOT EXISTS password_hash VARCHAR(64)")
            c.execute("ALTER TABLE onboarding_requests ADD COLUMN IF NOT EXISTS capstone_completed BOOLEAN DEFAULT FALSE")
            c.execute("ALTER TABLE onboarding_requests ADD COLUMN IF NOT EXISTS capstone_completed_at TIMESTAMP")
            c.execute("ALTER TABLE onboarding_requests ADD COLUMN IF NOT EXISTS active_track VARCHAR(50) DEFAULT 'rtcdp'")
            c.execute("ALTER TABLE onboarding_requests ADD COLUMN IF NOT EXISTS role VARCHAR(100) DEFAULT ''")
            c.execute("ALTER TABLE onboarding_requests ADD COLUMN IF NOT EXISTS profile_confirmed BOOLEAN DEFAULT FALSE")
            c.execute("ALTER TABLE onboarding_requests ADD COLUMN IF NOT EXISTS capstone_started_at TIMESTAMP")
            c.execute("ALTER TABLE onboarding_requests ADD COLUMN IF NOT EXISTS username VARCHAR(60)")
            c.execute("ALTER TABLE onboarding_requests ADD COLUMN IF NOT EXISTS avatar_emoji VARCHAR(10)")
            c.execute("ALTER TABLE onboarding_requests ADD COLUMN IF NOT EXISTS avatar_color VARCHAR(10)")
            c.execute("""CREATE TABLE IF NOT EXISTS manager_accounts (
                id SERIAL PRIMARY KEY, name VARCHAR(120) NOT NULL,
                email VARCHAR(255) UNIQUE NOT NULL, password_hash VARCHAR(64),
                team VARCHAR(100), joining_date DATE, status VARCHAR(20) DEFAULT 'pending',
                username VARCHAR(60), avatar_emoji VARCHAR(10), avatar_color VARCHAR(10),
                actioned_by VARCHAR(120), actioned_at TIMESTAMP,
                created_at TIMESTAMP DEFAULT NOW())""")
            # ── Capstone submissions — learner's generated scenario + response + AI evaluation ──
            c.execute("""CREATE TABLE IF NOT EXISTS capstone_submissions (
                id SERIAL PRIMARY KEY,
                member_id INTEGER NOT NULL REFERENCES onboarding_requests(id) ON DELETE CASCADE,
                scenario JSONB, response_text TEXT, ai_evaluation JSONB,
                status VARCHAR(30) DEFAULT 'generated',
                generated_at TIMESTAMP DEFAULT NOW(), submitted_at TIMESTAMP,
                evaluated_at TIMESTAMP, reviewed_at TIMESTAMP, manager_notes TEXT,
                manager_review_history JSONB DEFAULT '[]')""")
            c.execute("ALTER TABLE capstone_submissions ADD COLUMN IF NOT EXISTS manager_review_history JSONB DEFAULT '[]'")
            # Capstones are time-boxed: a 7-day deadline is stamped at generation.
            c.execute("ALTER TABLE capstone_submissions ADD COLUMN IF NOT EXISTS due_at TIMESTAMP")
            # ── Manager hierarchy + role-based learning journey (Excel-uploaded
            # reference data grounding the cross-skilling agent's recommendations
            # in real org structure and role expectations, not LLM improvisation) ──
            c.execute("""CREATE TABLE IF NOT EXISTS manager_hierarchy (
                id SERIAL PRIMARY KEY,
                manager_name VARCHAR(120) UNIQUE NOT NULL,
                reports_to VARCHAR(120),
                track_focus VARCHAR(255),
                notes TEXT,
                updated_at TIMESTAMP DEFAULT NOW())""")
            c.execute("""CREATE TABLE IF NOT EXISTS role_learning_journey (
                id SERIAL PRIMARY KEY,
                role VARCHAR(80) NOT NULL,
                priority INTEGER NOT NULL,
                target_proficiency VARCHAR(20),
                tracks JSONB DEFAULT '[]',
                notes TEXT,
                updated_at TIMESTAMP DEFAULT NOW(),
                UNIQUE(role, priority))""")
            # role_aliases: maps free-text HR/profile role strings (e.g. 'Data
            # Analyst') to the matrix's canonical role codes (e.g. 'AEP - DA'), so
            # the cross-skilling role lens matches real directory roles instead of
            # silently returning nothing. Admin-editable via /api/admin/role-aliases.
            c.execute("""CREATE TABLE IF NOT EXISTS role_aliases (
                id SERIAL PRIMARY KEY,
                alias VARCHAR(150) UNIQUE NOT NULL,
                canonical_role VARCHAR(80) NOT NULL,
                updated_at TIMESTAMP DEFAULT NOW())""")
            # Seed sensible defaults once (only when empty). Admins can add/edit/
            # remove these in-app; seeding just makes the role lens work out of the box.
            c.execute("SELECT COUNT(*) FROM role_aliases")
            if (c.fetchone() or [0])[0] == 0:
                for _al, _canon in [
                    ("Data Analyst", "AEP - DA"), ("AEP Analyst", "AEP - DA"),
                    ("Data Architect", "AEP - DA"),
                    ("Data Engineer", "AED - DE"), ("Analytics Engineer", "AED - DE"),
                    ("AEP Developer", "RTCDP"), ("AEP Admin", "RTCDP"),
                    ("CDP Specialist", "RTCDP"),
                    ("Analytics Analyst", "AA-SDK"), ("Web SDK", "AA-SDK"),
                    ("Engineering Services", "ES"),
                ]:
                    c.execute("INSERT INTO role_aliases (alias, canonical_role) "
                              "VALUES (%s,%s) ON CONFLICT (alias) DO NOTHING", (_al, _canon))
            c.execute("""CREATE TABLE IF NOT EXISTS org_data_upload_batches (
                id SERIAL PRIMARY KEY,
                dataset VARCHAR(40) NOT NULL,
                filename VARCHAR(255),
                uploaded_by VARCHAR(255),
                uploaded_at TIMESTAMP DEFAULT NOW(),
                row_count INT DEFAULT 0,
                status VARCHAR(20) DEFAULT 'applied')""")
            # ── HR-provisioning directory (source of truth, from Excel upload) ────
            # employee_directory holds the canonical roster; app state (approval,
            # capstone, progress) stays in onboarding_requests keyed by email.
            c.execute("""CREATE TABLE IF NOT EXISTS directory_upload_batches (
                id SERIAL PRIMARY KEY,
                filename VARCHAR(255),
                uploaded_by VARCHAR(255),
                uploaded_at TIMESTAMP DEFAULT NOW(),
                row_count INT DEFAULT 0,
                checksum VARCHAR(64),
                status VARCHAR(20) DEFAULT 'validated',   -- validated | applied | failed
                n_insert INT DEFAULT 0, n_update INT DEFAULT 0,
                n_deactivate INT DEFAULT 0, n_reactivate INT DEFAULT 0,
                report JSONB)""")
            c.execute("""CREATE TABLE IF NOT EXISTS employee_directory (
                email VARCHAR(255) PRIMARY KEY,
                first_name VARCHAR(120), last_name VARCHAR(120),
                doj DATE, role VARCHAR(150), location VARCHAR(120),
                manager_name VARCHAR(160), manager_email VARCHAR(255),
                team VARCHAR(120), primary_skill VARCHAR(150), resource_email VARCHAR(255),
                is_active BOOLEAN DEFAULT TRUE,
                first_seen_batch_id INT REFERENCES directory_upload_batches(id),
                last_seen_batch_id INT REFERENCES directory_upload_batches(id),
                created_at TIMESTAMP DEFAULT NOW(), updated_at TIMESTAMP DEFAULT NOW())""")
            # Index for the manager-derivation query (who is a manager_email).
            c.execute("CREATE INDEX IF NOT EXISTS idx_empdir_manager_email ON employee_directory(LOWER(manager_email))")
            c.execute("""CREATE TABLE IF NOT EXISTS directory_change_log (
                id SERIAL PRIMARY KEY,
                batch_id INT REFERENCES directory_upload_batches(id),
                email VARCHAR(255),
                change_type VARCHAR(20),                  -- insert | update | deactivate | reactivate
                field_name VARCHAR(60), old_value TEXT, new_value TEXT,
                changed_at TIMESTAMP DEFAULT NOW())""")
            c.execute("CREATE INDEX IF NOT EXISTS idx_dirchange_batch ON directory_change_log(batch_id)")
            c.execute("""CREATE TABLE IF NOT EXISTS user_certifications (
                id             SERIAL PRIMARY KEY,
                user_id        INTEGER,
                email          TEXT NOT NULL,
                full_name      TEXT DEFAULT '',
                cert_name      TEXT NOT NULL,
                cert_type      TEXT DEFAULT '',
                status         TEXT NOT NULL DEFAULT 'Active',
                issued_date    DATE,
                expiry_date    DATE,
                days_remaining INTEGER,
                created_at     TIMESTAMP DEFAULT NOW(),
                updated_at     TIMESTAMP DEFAULT NOW(),
                UNIQUE(email, cert_name))""")
            c.execute("CREATE INDEX IF NOT EXISTS idx_usercerts_email ON user_certifications(email)")
            c.execute("CREATE INDEX IF NOT EXISTS idx_usercerts_user  ON user_certifications(user_id)")
            c.execute("""CREATE TABLE IF NOT EXISTS projects (
                id                    SERIAL PRIMARY KEY,
                manager_email         TEXT NOT NULL,
                title                 TEXT NOT NULL,
                sector                TEXT DEFAULT '',
                tag                   TEXT DEFAULT '',
                sprint                TEXT DEFAULT '',
                status                TEXT DEFAULT 'Planning',
                description           TEXT DEFAULT '',
                color                 TEXT DEFAULT '#1473E6',
                project_code          TEXT DEFAULT '',
                project_type          TEXT DEFAULT '',
                industry              TEXT DEFAULT '',
                phase                 TEXT DEFAULT '',
                stage                 TEXT DEFAULT '',
                start_date            DATE,
                end_date              DATE,
                solutions_used        TEXT DEFAULT '',
                health_status         TEXT DEFAULT '',
                weekly_comments       TEXT DEFAULT '',
                high_level_notes      TEXT DEFAULT '',
                renewal               TEXT DEFAULT '',
                region                TEXT DEFAULT '',
                use_cases             TEXT DEFAULT '',
                product_features      TEXT DEFAULT '',
                data_sources          TEXT DEFAULT '',
                destinations          TEXT DEFAULT '',
                num_audiences         INTEGER DEFAULT 0,
                ticket_ids            TEXT DEFAULT '',
                days_remaining        INTEGER,
                imported_from_tracker BOOLEAN DEFAULT FALSE,
                is_initiative         BOOLEAN DEFAULT FALSE,
                created_at            TIMESTAMP DEFAULT NOW(),
                updated_at            TIMESTAMP DEFAULT NOW())""")
            c.execute("CREATE INDEX IF NOT EXISTS idx_projects_mgr ON projects(LOWER(manager_email))")
            # Tracker-import enrichment columns — ADD COLUMN IF NOT EXISTS so
            # these land on startup for existing databases too (the CREATE
            # TABLE above only takes effect for a brand-new table).
            c.execute("ALTER TABLE projects ADD COLUMN IF NOT EXISTS product_features TEXT DEFAULT ''")
            c.execute("ALTER TABLE projects ADD COLUMN IF NOT EXISTS data_sources TEXT DEFAULT ''")
            c.execute("ALTER TABLE projects ADD COLUMN IF NOT EXISTS destinations TEXT DEFAULT ''")
            c.execute("ALTER TABLE projects ADD COLUMN IF NOT EXISTS num_audiences INTEGER DEFAULT 0")
            c.execute("ALTER TABLE projects ADD COLUMN IF NOT EXISTS ticket_ids TEXT DEFAULT ''")
            c.execute("ALTER TABLE projects ADD COLUMN IF NOT EXISTS days_remaining INTEGER")
            c.execute("""CREATE TABLE IF NOT EXISTS project_members (
                id              SERIAL PRIMARY KEY,
                project_id      INTEGER REFERENCES projects(id) ON DELETE CASCADE,
                member_email    TEXT NOT NULL,
                member_name     TEXT NOT NULL,
                hrs_per_week    NUMERIC(5,1) DEFAULT 0,
                role_on_project TEXT DEFAULT '',
                added_at        TIMESTAMP DEFAULT NOW(),
                UNIQUE(project_id, member_email))""")
            c.execute("""CREATE TABLE IF NOT EXISTS member_initiatives (
                id            SERIAL PRIMARY KEY,
                member_name   TEXT NOT NULL,
                member_email  TEXT DEFAULT '',
                initiative    TEXT NOT NULL,
                latest_update TEXT DEFAULT '',
                date_logged   DATE,
                manager_email TEXT DEFAULT '',
                created_at    TIMESTAMP DEFAULT NOW(),
                updated_at    TIMESTAMP DEFAULT NOW())""")
            c.execute("CREATE INDEX IF NOT EXISTS idx_initiatives_member ON member_initiatives(LOWER(member_name))")
            c.execute("""CREATE TABLE IF NOT EXISTS initiative_updates (
                id            SERIAL PRIMARY KEY,
                initiative_id INTEGER REFERENCES member_initiatives(id) ON DELETE CASCADE,
                update_text   TEXT NOT NULL,
                created_at    TIMESTAMP DEFAULT NOW())""")
            c.execute("""CREATE TABLE IF NOT EXISTS member_milestones (
                id            SERIAL PRIMARY KEY,
                member_name   TEXT NOT NULL,
                member_email  TEXT DEFAULT '',
                milestone_date DATE,
                note          TEXT NOT NULL,
                project_name  TEXT DEFAULT '',
                manager_email TEXT DEFAULT '',
                created_at    TIMESTAMP DEFAULT NOW())""")
            c.execute("""CREATE TABLE IF NOT EXISTS project_issues (
                id         SERIAL PRIMARY KEY,
                project_id INTEGER REFERENCES projects(id) ON DELETE CASCADE,
                title      TEXT NOT NULL,
                priority   TEXT DEFAULT 'Medium',
                status     TEXT DEFAULT 'Open',
                visibility TEXT DEFAULT 'Everyone',
                created_at TIMESTAMP DEFAULT NOW(),
                updated_at TIMESTAMP DEFAULT NOW())""")
            # Append-only weekly-note history — every "Update" posts a new
            # timestamped row instead of overwriting the previous note, so
            # nothing a member wrote is ever silently lost.
            c.execute("""CREATE TABLE IF NOT EXISTS project_weekly_updates (
                id           SERIAL PRIMARY KEY,
                project_id   INTEGER REFERENCES projects(id) ON DELETE CASCADE,
                member_email TEXT DEFAULT '',
                member_name  TEXT DEFAULT '',
                update_text  TEXT NOT NULL,
                created_at   TIMESTAMP DEFAULT NOW())""")
            c.execute("CREATE INDEX IF NOT EXISTS idx_projupdates_project ON project_weekly_updates(project_id)")
            print("✓ All tables ready")
        conn.close()
    except Exception as e:
        print(f"Startup table creation error: {e}")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:5173", "http://localhost:3000"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Register the Adobe IMS auth routes:
#   GET  /api/auth/ims/config    GET  /api/auth/ims/login
#   GET  /api/auth/ims/callback  GET  /api/auth/session   POST /api/auth/logout
app.include_router(ims_router)

# ────────────────────────────── Learning-agent routers (added) ──────────────────────────────
# Mount the (previously dormant) Curriculum quiz-lifecycle router and the new
# spec-named learning-agent router. Both are additive and namespaced, so they
# do not collide with or alter any existing route.
try:
    from agents.curriculum_routes import router as curriculum_router
    app.include_router(curriculum_router)
    from routes_learning import router as learning_router
    app.include_router(learning_router)
    print("✅ Mounted curriculum_router + learning_router")
except Exception as _lr_err:
    print(f"⚠ learning routers not mounted: {_lr_err}")

GROQ_KEY      = os.getenv("GROQ_API_KEY", "")
ANTHROPIC_KEY = os.getenv("ANTHROPIC_API_KEY", "")
GITHUB_TOKEN  = os.getenv("GITHUB_TOKEN", "")
DATABASE_URL  = os.getenv("DATABASE_URL", "")  # must be set in .env

# Manager alias map for verification — set only in this server's own .env,
# never user-editable or exposed as a UI control (a free-text "view as
# anyone" box would let any user snoop on another manager's real team data).
# MGR_ALIAS_MAP is a JSON object: {"tester_id": {"email": "...", "name": "..."}}.
# The tester_id is whatever identity the frontend has for the person
# currently testing — their own real email if logged in, or the literal
# string "mgr" for the anonymous demo-manager login. Whoever is testing gets
# redirected to view the mapped real manager's live data.
try:
    MGR_ALIAS_MAP = {
        str(k).strip().lower(): v
        for k, v in json.loads(os.getenv("MGR_ALIAS_MAP", "") or "{}").items()
    }
except Exception as _e:
    print(f"⚠ MGR_ALIAS_MAP in .env is not valid JSON, ignoring: {_e}")
    MGR_ALIAS_MAP = {}

@app.get("/api/config/manager-alias")
def get_manager_alias(as_email: str = ""):
    """Looks up the server-configured alias map for this tester's id. Returns
    the manager identity to view as, if this tester has a configured alias —
    otherwise `email: None` and the caller keeps using its own identity."""
    entry = MGR_ALIAS_MAP.get((as_email or "").strip().lower())
    if not entry or not entry.get("email"):
        return {"email": None}
    return {"email": entry["email"], "name": entry.get("name")}

# ── Community (threads/replies) — real backend replacing the previous
# hardcoded NJ_THREADS/THREADS frontend arrays. Points/streak are computed live
# from real rows, never stored/fabricated. `space` is "nj" (new-joiner cohort)
# or "exp" (experienced-team community) — the two UIs that already existed.
class CommunityThreadCreate(BaseModel):
    space: str = "team"
    author_name: str
    author_email: Optional[str] = None
    manager_email: Optional[str] = None
    title: str
    body: Optional[str] = None
    tag: str = "General"
    product: Optional[str] = None
    visibility: str = "team"           # private | team | public
    mentions: List[dict] = []          # [{name, email}] of people tagged

class CommunityReplyCreate(BaseModel):
    author_name: str
    author_email: Optional[str] = None
    text: str
    mentions: List[dict] = []


def _notify(cur, member_name, ntype, title, message, actor=None, thread_id=None, manager=None):
    """Insert one notification row. Never notifies a blank/None recipient."""
    if not member_name:
        return
    cur.execute("""INSERT INTO notifications (member_name, manager, type, title, message, actor, thread_id)
                   VALUES (%s,%s,%s,%s,%s,%s,%s)""",
                (member_name, manager, ntype, title, message, actor, thread_id))


def _session_identity(request: Request):
    """Return {name, email} from a valid session, or None. Used to stop an
    AUTHENTICATED user from posting/reacting as someone else — the client-sent
    author is overridden by the real session identity when one exists. Demo
    profiles (no session) keep their supplied name, so demo mode still works."""
    try:
        payload = get_current_user(request)
    except HTTPException:
        return None
    except Exception:
        return None
    prof = (payload or {}).get("profile") or {}
    nm = prof.get("name") or prof.get("displayName")
    return {"name": nm, "email": prof.get("email")} if nm else None

def _community_streak(cur, space: str, author_name: str) -> int:
    """Consecutive days (ending today or the most recent activity day) the
    author posted a thread or reply in this space. Real, computed from
    timestamps — not a fabricated counter."""
    cur.execute("""
        SELECT DISTINCT DATE(created_at) AS d FROM (
            SELECT created_at FROM community_threads WHERE space=%s AND author_name=%s
            UNION ALL
            SELECT r.created_at FROM community_replies r
              JOIN community_threads t ON t.id=r.thread_id
              WHERE t.space=%s AND r.author_name=%s
        ) x ORDER BY d DESC""", (space, author_name, space, author_name))
    days = [r[0] for r in cur.fetchall()]
    if not days:
        return 0
    from datetime import date, timedelta
    streak = 0
    expect = days[0]
    for d in days:
        if d == expect:
            streak += 1
            expect = expect - timedelta(days=1)
        elif d < expect:
            break
    return streak

@app.get("/api/community/threads")
def list_community_threads(space: str = None, tag: str = None, product: str = None,
                           as_name: str = None, as_email: str = None, my_manager: str = None):
    """Visibility-aware community feed.

    Legacy mode (space=nj|exp, no viewer identity): returns that space's threads
    unchanged, so the old NJ/EXP UIs keep working.

    Team-feed mode (viewer identity via as_name/as_email + my_manager): returns
    everything the viewer is allowed to see —
      • public   → all teams
      • team     → posts whose manager_email == the viewer's manager
      • private  → only the viewer's own posts
    """
    try:
        with get_db() as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                clauses, params = [], []
                viewer_feed = bool(as_name or as_email)
                if viewer_feed:
                    vis = ["visibility='public'"]
                    if my_manager:
                        vis.append("(visibility='team' AND lower(manager_email)=lower(%s))"); params.append(my_manager)
                    own = []
                    if as_email: own.append("lower(author_email)=lower(%s)"); params.append(as_email)
                    if as_name:  own.append("lower(author_name)=lower(%s)"); params.append(as_name)
                    if own:
                        vis.append("(visibility='private' AND ("+" OR ".join(own)+"))")
                    clauses.append("("+" OR ".join(vis)+")")
                elif space:
                    clauses.append("space=%s"); params.append(space)
                if tag and tag != "All":
                    clauses.append("tag=%s"); params.append(tag)
                if product and product != "All":
                    clauses.append("product=%s"); params.append(product)
                where = (" WHERE "+" AND ".join(clauses)) if clauses else ""
                cur.execute("SELECT id, author_name, author_email, manager_email, title, body, tag, "
                           "product, visibility, mentions, created_at "
                           f"FROM community_threads{where} ORDER BY created_at DESC LIMIT 200", tuple(params))
                threads = [dict(r) for r in cur.fetchall()]
                me = as_name or ""
                for t in threads:
                    t["created_at"] = str(t["created_at"])
                    cur.execute("SELECT author_name, text, created_at FROM community_replies "
                               "WHERE thread_id=%s ORDER BY created_at ASC", (t["id"],))
                    t["replies"] = [dict(r) for r in cur.fetchall()]
                    for rep in t["replies"]:
                        rep["created_at"] = str(rep["created_at"])
                    cur.execute("SELECT COUNT(*) AS c FROM community_reactions WHERE thread_id=%s", (t["id"],))
                    t["reactions"] = cur.fetchone()["c"]
                    if me:
                        cur.execute("SELECT 1 FROM community_reactions WHERE thread_id=%s AND member_name=%s", (t["id"], me))
                        t["reacted"] = bool(cur.fetchone())
                    else:
                        t["reacted"] = False
        return {"threads": threads}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/community/threads")
def create_community_thread(body: CommunityThreadCreate, request: Request):
    if not body.title.strip():
        raise HTTPException(status_code=422, detail="title is required.")
    vis = body.visibility if body.visibility in ("private", "team", "public") else "team"
    ident = _session_identity(request)
    author_name = (ident and ident["name"]) or body.author_name
    author_email = (ident and ident["email"]) or body.author_email
    try:
        with get_db() as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute("""INSERT INTO community_threads
                              (space, author_name, author_email, manager_email, title, body, tag, product, visibility, mentions)
                              VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                              RETURNING id, author_name, author_email, manager_email, title, body, tag, product, visibility, mentions, created_at""",
                           (body.space, author_name, author_email, body.manager_email,
                            body.title.strip(), (body.body or "").strip() or None, body.tag, body.product,
                            vis, json.dumps(body.mentions or [])))
                t = dict(cur.fetchone())
                tid = t["id"]
                t["created_at"] = str(t["created_at"])
                t["replies"] = []
                # ── Notifications ──
                # 1) @mentions (any visibility) — notify each tagged person.
                for m in (body.mentions or []):
                    nm = (m.get("name") if isinstance(m, dict) else None)
                    if nm and nm != author_name:
                        _notify(cur, nm, "mention", f"{author_name} mentioned you",
                                body.title.strip()[:140], actor=author_name, thread_id=tid, manager=body.manager_email)
                # 2) public post — light "new public post" notice to others. Capped
                # so a huge directory can't create thousands of rows per post.
                if vis == "public":
                    cur.execute("""SELECT DISTINCT TRIM(COALESCE(first_name,'')||' '||COALESCE(last_name,'')) AS nm
                                   FROM employee_directory WHERE is_active=TRUE LIMIT %s""",
                                (int(os.getenv("COMMUNITY_PUBLIC_NOTIFY_CAP", "150")),))
                    for (nm,) in cur.fetchall():
                        if nm and nm != author_name:
                            _notify(cur, nm, "public_post", f"{author_name} posted to all teams",
                                    body.title.strip()[:140], actor=author_name, thread_id=tid)
        return {"ok": True, "thread": t}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

class CommunityThreadEdit(BaseModel):
    editor_name: str
    title: Optional[str] = None
    body: Optional[str] = None
    tag: Optional[str] = None
    visibility: Optional[str] = None

def _can_modify(thread, editor_name, ident):
    """Only the post's author may edit/delete it. When a session exists, the
    session identity is authoritative; otherwise fall back to the supplied name."""
    who = (ident and ident["name"]) or editor_name
    return who and thread.get("author_name") == who

@app.put("/api/community/threads/{thread_id}")
def edit_community_thread(thread_id: int, body: CommunityThreadEdit, request: Request):
    ident = _session_identity(request)
    try:
        with get_db() as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute("SELECT author_name FROM community_threads WHERE id=%s", (thread_id,))
                thread = cur.fetchone()
                if not thread:
                    raise HTTPException(status_code=404, detail="Thread not found.")
                if not _can_modify(thread, body.editor_name, ident):
                    raise HTTPException(status_code=403, detail="Only the author can edit this post.")
                sets, params = [], []
                if body.title is not None and body.title.strip():
                    sets.append("title=%s"); params.append(body.title.strip())
                if body.body is not None:
                    sets.append("body=%s"); params.append(body.body.strip() or None)
                if body.tag is not None:
                    sets.append("tag=%s"); params.append(body.tag)
                if body.visibility in ("private", "team", "public"):
                    sets.append("visibility=%s"); params.append(body.visibility)
                if not sets:
                    return {"ok": True}
                params.append(thread_id)
                cur.execute(f"UPDATE community_threads SET {','.join(sets)} WHERE id=%s", tuple(params))
        return {"ok": True}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/api/community/threads/{thread_id}")
def delete_community_thread(thread_id: int, editor_name: str, request: Request):
    ident = _session_identity(request)
    try:
        with get_db() as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute("SELECT author_name FROM community_threads WHERE id=%s", (thread_id,))
                thread = cur.fetchone()
                if not thread:
                    raise HTTPException(status_code=404, detail="Thread not found.")
                if not _can_modify(thread, editor_name, ident):
                    raise HTTPException(status_code=403, detail="Only the author can delete this post.")
                cur.execute("DELETE FROM community_threads WHERE id=%s", (thread_id,))  # replies/reactions cascade
        return {"ok": True}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/community/threads/{thread_id}/replies")
def create_community_reply(thread_id: int, body: CommunityReplyCreate, request: Request):
    if not body.text.strip():
        raise HTTPException(status_code=422, detail="text is required.")
    ident = _session_identity(request)
    author_name = (ident and ident["name"]) or body.author_name
    try:
        with get_db() as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute("SELECT author_name, title, manager_email FROM community_threads WHERE id=%s", (thread_id,))
                thread = cur.fetchone()
                if not thread:
                    raise HTTPException(status_code=404, detail="Thread not found.")
                cur.execute("""INSERT INTO community_replies (thread_id, author_name, text)
                              VALUES (%s,%s,%s) RETURNING author_name, text, created_at""",
                           (thread_id, author_name, body.text.strip()))
                r = dict(cur.fetchone())
                r["created_at"] = str(r["created_at"])
                # Notify the post's author that someone replied (not on self-reply).
                if thread["author_name"] and thread["author_name"] != author_name:
                    _notify(cur, thread["author_name"], "reply", f"{author_name} replied to your post",
                            (thread["title"] or "")[:140], actor=author_name, thread_id=thread_id,
                            manager=thread["manager_email"])
                # Notify anyone @mentioned in the reply.
                for m in (body.mentions or []):
                    nm = (m.get("name") if isinstance(m, dict) else None)
                    if nm and nm != author_name:
                        _notify(cur, nm, "mention", f"{author_name} mentioned you in a reply",
                                (thread["title"] or "")[:140], actor=author_name, thread_id=thread_id)
        return {"ok": True, "reply": r}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# Notifications list/read endpoints already exist further down (get_notifications
# + PUT mark-read / read-all, used by NotificationsWidget). The community feature
# reuses them; new rows are written via _notify() with the standard columns those
# endpoints return, so nothing extra is needed here.

@app.get("/api/community/members")
def community_members(manager_email: str = None, q: str = None):
    """People available to @mention — the manager's team (same manager_email),
    plus the manager. Used by the compose box's mention autocomplete."""
    try:
        with get_db() as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                clauses, params = ["is_active=TRUE"], []
                if manager_email:
                    clauses.append("lower(manager_email)=lower(%s)"); params.append(manager_email)
                cur.execute(f"""SELECT DISTINCT TRIM(COALESCE(first_name,'')||' '||COALESCE(last_name,'')) AS name, email
                               FROM employee_directory WHERE {' AND '.join(clauses)}
                               ORDER BY name LIMIT 200""", tuple(params))
                people = [dict(r) for r in cur.fetchall() if r["name"]]
                if q:
                    ql = q.lower()
                    people = [x for x in people if ql in (x["name"] or "").lower()]
        return {"members": people}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# ── Community scoring — quality over quantity ────────────────────────────────
# Posting is worth little on its own (can be farmed); the real driver is
# reactions RECEIVED — i.e. other people found your post useful. Kept small and
# simple so it motivates helpfulness without turning the feed into point-farming.
PTS_POST, PTS_REPLY, PTS_KUDOS_RECEIVED = 2, 3, 5

class CommunityReact(BaseModel):
    member_name: str

@app.post("/api/community/threads/{thread_id}/react")
def toggle_community_reaction(thread_id: int, body: CommunityReact, request: Request):
    """Toggle a 👍 kudos on a post. Returns the new count + whether the caller
    now has one active. Notifies the post author on a fresh kudos."""
    ident = _session_identity(request)
    member = (ident and ident["name"]) or body.member_name
    if not member:
        raise HTTPException(status_code=422, detail="member_name required.")
    body_member_name = member
    try:
        with get_db() as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute("SELECT author_name, title, manager_email FROM community_threads WHERE id=%s", (thread_id,))
                thread = cur.fetchone()
                if not thread:
                    raise HTTPException(status_code=404, detail="Thread not found.")
                cur.execute("DELETE FROM community_reactions WHERE thread_id=%s AND member_name=%s RETURNING id",
                            (thread_id, body_member_name))
                removed = cur.fetchone()
                reacted = False
                if not removed:
                    cur.execute("INSERT INTO community_reactions (thread_id, member_name) VALUES (%s,%s) ON CONFLICT DO NOTHING",
                                (thread_id, body_member_name))
                    reacted = True
                    if thread["author_name"] and thread["author_name"] != body_member_name:
                        _notify(cur, thread["author_name"], "kudos", f"{body_member_name} gave your post kudos 👍",
                                (thread["title"] or "")[:140], actor=body_member_name, thread_id=thread_id,
                                manager=thread["manager_email"])
                cur.execute("SELECT COUNT(*) AS c FROM community_reactions WHERE thread_id=%s", (thread_id,))
                count = cur.fetchone()["c"]
        return {"ok": True, "reacted": reacted, "reactions": count}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

def _score(posts, replies, kudos_received):
    return posts * PTS_POST + replies * PTS_REPLY + kudos_received * PTS_KUDOS_RECEIVED

@app.get("/api/community/stats")
def community_stats(space: str, author_name: str):
    try:
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT COUNT(*) FROM community_threads WHERE space=%s AND author_name=%s",
                           (space, author_name))
                posts = cur.fetchone()[0]
                cur.execute("""SELECT COUNT(*) FROM community_replies r
                              JOIN community_threads t ON t.id=r.thread_id
                              WHERE t.space=%s AND r.author_name=%s""", (space, author_name))
                replies = cur.fetchone()[0]
                # kudos received across ALL this author's posts (any space) — usefulness signal
                cur.execute("""SELECT COUNT(*) FROM community_reactions x
                              JOIN community_threads t ON t.id=x.thread_id
                              WHERE t.author_name=%s""", (author_name,))
                kudos = cur.fetchone()[0]
                streak = _community_streak(cur, space, author_name)
        return {"points": _score(posts, replies, kudos), "posts": posts, "replies": replies,
                "kudos": kudos, "streak": streak}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/community/leaderboard")
def community_leaderboard(space: str):
    """Real leaderboard ranked by the quality-weighted score (reactions received
    dominate). Only people who have actually participated appear."""
    try:
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT author_name, COUNT(*) AS posts FROM community_threads WHERE space=%s GROUP BY author_name", (space,))
                posts_by = {r[0]: r[1] for r in cur.fetchall()}
                cur.execute("""SELECT r.author_name, COUNT(*) AS replies FROM community_replies r
                    JOIN community_threads t ON t.id=r.thread_id WHERE t.space=%s GROUP BY r.author_name""", (space,))
                replies_by = {r[0]: r[1] for r in cur.fetchall()}
                cur.execute("""SELECT t.author_name, COUNT(*) AS k FROM community_reactions x
                    JOIN community_threads t ON t.id=x.thread_id WHERE t.space=%s GROUP BY t.author_name""", (space,))
                kudos_by = {r[0]: r[1] for r in cur.fetchall()}
                names = set(posts_by) | set(replies_by) | set(kudos_by)
                board = []
                for n in names:
                    p, r, k = posts_by.get(n, 0), replies_by.get(n, 0), kudos_by.get(n, 0)
                    board.append({"name": n, "posts": p, "replies": r, "kudos": k,
                                 "points": _score(p, r, k), "streak": _community_streak(cur, space, n)})
                board.sort(key=lambda x: x["points"], reverse=True)
        return {"leaderboard": board}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ── Remediation — a targeted catch-up path when a learner is struggling ───────
@app.get("/api/curriculum/remediation")
def curriculum_remediation(member_name: str, track: str = "rtcdp"):
    """Detect weak areas (failed test-outs + low confidence) and return a
    prioritised catch-up plan: which modules to revisit, why, and the
    recommended actions (revisit lesson / Socratic drill / re-take test-out).
    on_track=True with an empty plan means nothing needs remediation."""
    LOW_CONF = float(os.getenv("REMEDIATION_CONF_THRESHOLD", "0.6"))
    PASS_PCT = int(os.getenv("REMEDIATION_TESTOUT_PASS", "70"))
    try:
        with get_db() as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                # Most recent FAILED test-out per module (strongest weakness signal).
                cur.execute("""SELECT DISTINCT ON (module_id) module_id, module_title, score
                               FROM module_test_outs
                               WHERE member_name=%s AND track=%s AND passed=FALSE
                               ORDER BY module_id, created_at DESC""", (member_name, track))
                failed = {r["module_id"]: r for r in cur.fetchall()}
                # Latest low-confidence score per module (keyed by module title/name).
                cur.execute("""SELECT DISTINCT ON (module) module, score
                               FROM confidence_scores WHERE user_name=%s
                               ORDER BY module, created_at DESC""", (member_name,))
                low_conf = {r["module"]: r["score"] for r in cur.fetchall()
                            if r["score"] is not None and r["score"] < LOW_CONF and r["module"]}
                # Map module titles -> module_id for this track (to attach actions).
                cur.execute("SELECT DISTINCT module_id, title FROM curriculum_topics WHERE track=%s", (track,))
                title_to_mid = {}
                for r in cur.fetchall():
                    title_to_mid.setdefault((r["title"] or "").lower(), r["module_id"])

        plan, seen = [], set()
        for mid, r in failed.items():
            sc = r.get("score") or 0
            plan.append({"module_id": mid, "module_title": r["module_title"],
                         "kind": "failed_testout", "score": sc,
                         "reason": f"Scored {sc}% on the test-out (needs {PASS_PCT}%+).",
                         "severity": max(1, PASS_PCT - sc),
                         "actions": ["revisit", "socratic", "retest"]})
            seen.add(mid)
        for mod_title, conf in low_conf.items():
            mid = title_to_mid.get((mod_title or "").lower())
            if mid in seen:
                continue
            plan.append({"module_id": mid, "module_title": mod_title,
                         "kind": "low_confidence", "confidence": round(conf, 2),
                         "reason": f"Confidence is {round(conf*100)}% (below {round(LOW_CONF*100)}%).",
                         "severity": max(1, round((LOW_CONF - conf) * 100)),
                         "actions": ["revisit", "socratic"]})
        plan.sort(key=lambda x: x["severity"], reverse=True)
        return {"member_name": member_name, "track": track,
                "weak_count": len(plan), "on_track": len(plan) == 0, "plan": plan}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


import re
import html
import datetime as _dt
import requests
import threading

# ── Release notes — real, parsed from each product's actual AdobeDocs release-
# notes page (replaces the hardcoded, frozen RELEASE_NOTES frontend object).
# Source pages verified to be real markdown/HTML on GitHub as of 2026; formats
# differ per product (markdown tables, raw HTML tables, version+bullet lists),
# so the parser tries several patterns per section rather than assuming one.
RELEASE_NOTES_SOURCES = {
    # AEP publishes one file per month, and the current month's file often isn't
    # published yet (docs lag a release) — so this yields several candidates,
    # newest first, and fetch_release_notes tries each until one exists.
    "AEP": {"repo": "AdobeDocs/experience-platform.en",
            "path": lambda: [
                f"help/release-notes/{(_dt.datetime.now() - _dt.timedelta(days=30*i)).year}/"
                f"{(_dt.datetime.now() - _dt.timedelta(days=30*i)).strftime('%B').lower()}-"
                f"{(_dt.datetime.now() - _dt.timedelta(days=30*i)).year}.md"
                for i in range(4)
            ],
            "el_url": "https://experienceleague.adobe.com/en/docs/experience-platform/release-notes/latest"},
    "AJO": {"repo": "AdobeDocs/journey-optimizer.en",
            "path": lambda: f"help/using/rn/release-notes-{_dt.datetime.now().year}.md",
            "el_url": "https://experienceleague.adobe.com/en/docs/journey-optimizer/using/whats-new/release-notes"},
    "CJA": {"repo": "AdobeDocs/analytics-platform.en",
            "path": lambda: f"help/release-notes/{_dt.datetime.now().year}.md",
            "el_url": "https://experienceleague.adobe.com/en/docs/analytics-platform/using/releases/latest"},
    "WebSDK": {"repo": "AdobeDocs/experience-platform.en",
               "path": lambda: "help/tags/extensions/client/web-sdk/web-sdk-ext-release-notes.md",
               "el_url": "https://experienceleague.adobe.com/en/docs/experience-platform/web-sdk/release-notes"},
    # Analytics: no current-year combined file existed at build time in analytics.en;
    # falls back to the same 2026.md pattern other products use, degrading to "no
    # entries parsed" (not a crash) if the file genuinely isn't there this year.
    "Analytics": {"repo": "AdobeDocs/analytics.en",
                  "path": lambda: f"help/release-notes/{_dt.datetime.now().year}.md",
                  "el_url": "https://experienceleague.adobe.com/en/docs/analytics/release-notes/latest"},
}

def _rn_clean(s: str) -> str:
    s = re.sub(r'<[^>]+>', ' ', s or '')
    s = re.sub(r'\[!DNL\s*([^\]]+)\]', r'\1', s)
    s = re.sub(r'\[!UICONTROL\s*([^\]]+)\]', r'\1', s)
    s = re.sub(r'\[([^\]]+)\]\([^)]+\)', r'\1', s)
    s = re.sub(r'\{#[^}]+\}', '', s)  # strip markdown heading anchors {#id}
    s = html.unescape(s)
    return re.sub(r'\s+', ' ', s).strip()

def _rn_parse_section(period: str, body: str, max_entries: int = 6):
    """Try several real-world release-note entry formats against one section
    of text (already split by its '## <period>' heading) and return
    [(title, description)], newest-first as they appear in the doc."""
    entries = []
    for m in re.finditer(r'\|\s*\*\*(.+?)\*\*\s*(?:<br/?>)?\s*(.*?)\s*\|', body):
        title, rest = _rn_clean(m.group(1)), _rn_clean(m.group(2))
        if title and len(title) < 120:
            entries.append((title, rest[:260]))
    for m in re.finditer(r'^\|\s*([A-Z][^|]{2,60}?)\s*\|\s*([^|]{15,400}?)\s*\|\s*$', body, re.MULTILINE):
        title, desc = _rn_clean(m.group(1)), _rn_clean(m.group(2))
        if title and title.lower() not in ('feature', 'description'):
            entries.append((title, desc[:260]))
    for m in re.finditer(r'<th>\s*<strong>(.+?)</strong>.*?</th>.*?<td>(.*?)</td>', body, re.DOTALL):
        title = _rn_clean(m.group(1))
        pm = re.search(r'<p>(.*?)</p>', m.group(2), re.DOTALL)
        desc = _rn_clean(pm.group(1)) if pm else _rn_clean(m.group(2))
        if title:
            entries.append((title, desc[:260]))
    for m in re.finditer(r'^##\s+(Version[^\n]+)\n+((?:^-\s.+\n?)+)', body, re.MULTILINE):
        title = _rn_clean(m.group(1))
        bullets = re.findall(r'^-\s(.+)$', m.group(2), re.MULTILINE)
        entries.append((title, ' '.join(_rn_clean(b) for b in bullets[:2])[:260]))
    # WebSDK-style docs: each release IS its own top-level "## Version X - Date"
    # heading (the section splitter already consumed it into `period`), with a
    # plain bullet list as the body — no nested title to find inside `body`.
    if not entries and re.match(r'^Version\s', period) and re.search(r'^-\s', body, re.MULTILINE):
        bullets = re.findall(r'^-\s(.+)$', body, re.MULTILINE)
        if bullets:
            # Use just "Month Year" as the period badge (distinct from the
            # full "Version X - Date" title) so the two don't read as duplicates.
            dm = re.search(r'([A-Z][a-z]+)\s+\d{1,2},?\s*(\d{4})', period)
            short_period = f"{dm.group(1)} {dm.group(2)}" if dm else period
            entries.append((period, ' '.join(_rn_clean(b) for b in bullets[:2])[:260], short_period))
    seen, out = set(), []
    for entry in entries:
        t, d = entry[0], entry[1]
        this_period = entry[2] if len(entry) > 2 else period
        k = t.lower()
        if k in seen:
            continue
        seen.add(k)
        out.append((this_period, t, d))
        if len(out) >= max_entries:
            break
    return out

def fetch_release_notes(product: str, max_entries: int = 8) -> list:
    """Fetch + parse ONE product's real release-notes doc. Returns
    [(period, title, description, source_url)]. Never raises — returns []
    on any failure so a bad fetch degrades gracefully instead of 500ing."""
    src = RELEASE_NOTES_SOURCES.get(product)
    if not src:
        return []
    try:
        repo = src["repo"]
        paths = src["path"]()
        paths = paths if isinstance(paths, list) else [paths]
        gh_headers = {"User-Agent": "Nexus-Platform/1.0"}
        if GITHUB_TOKEN:
            gh_headers["Authorization"] = f"token {GITHUB_TOKEN}"
        text, matched_path = None, None
        for path in paths:
            for branch in ("main", "master"):
                r = requests.get(f"https://raw.githubusercontent.com/{repo}/{branch}/{path}",
                                 headers=gh_headers, timeout=20)
                if r.status_code == 200:
                    text, matched_path = r.text, path
                    break
            if text:
                break
        if not text:
            return []
        # ── Month resolution ──────────────────────────────────────────────────
        # Products like AEP publish ONE file per month (…/august-2026.md) whose
        # top-level "## headings" are feature CATEGORIES, not months — so the
        # month has to come from the filename, otherwise the feed groups by
        # "Destinations"/"Segmentation Service" instead of "August 2026".
        # Products that publish one file per YEAR use "## Month Year" headings,
        # so there the heading IS the month and we keep it.
        _MONTHS = ("january","february","march","april","may","june","july",
                   "august","september","october","november","december")
        mm = re.search(r'(' + '|'.join(_MONTHS) + r')[-_ ](\d{4})', (matched_path or '').lower())
        file_month = f"{mm.group(1).capitalize()} {mm.group(2)}" if mm else None
        # A real, month-specific Experience League link when the GitHub path
        # mirrors the EL structure (AEP) — beats the generic ".../latest".
        url = src["el_url"]
        if file_month and matched_path and "experience-platform" in repo:
            sub = matched_path[5:] if matched_path.startswith("help/") else matched_path
            if sub.endswith(".md"):
                sub = sub[:-3]
            url = f"https://experienceleague.adobe.com/en/docs/experience-platform/{sub}"
        parts = re.split(r'^##\s+(.+)$', text, flags=re.MULTILINE)
        sections = list(zip(parts[1::2], parts[2::2])) if len(parts) > 1 else [("", text)]
        out = []
        for heading, body in sections:
            period = _rn_clean(heading) or "Latest"
            for period2, title, desc in _rn_parse_section(period, body, max_entries=max_entries):
                # file_month (per-month files) always wins so the feed is month-wise;
                # otherwise use the per-section period (a "Month Year" heading).
                out.append((file_month or period2, title, desc, url))
            if len(out) >= max_entries:
                break
        return out[:max_entries]
    except Exception as e:
        print(f"[release-notes] fetch failed for {product}: {e}")
        return []

def _release_notes_refresh(product: str) -> int:
    entries = fetch_release_notes(product)
    if not entries:
        return 0
    try:
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("DELETE FROM release_notes_cache WHERE product=%s", (product,))
                for period, title, desc, url in entries:
                    cur.execute("""INSERT INTO release_notes_cache (product, period, title, description, source_url)
                                  VALUES (%s,%s,%s,%s,%s) ON CONFLICT (product, title) DO NOTHING""",
                               (product, period, title, desc, url))
        return len(entries)
    except Exception as e:
        print(f"[release-notes] cache write failed for {product}: {e}")
        return 0

@app.get("/api/release-notes")
def get_release_notes(product: str = None, max_age_hours: int = 24):
    """Real release notes, parsed from each product's actual Experience League
    doc (via its AdobeDocs GitHub repo). Each GitHub fetch can take several
    seconds (or time out) on networks that throttle/block raw.githubusercontent.com,
    so this endpoint NEVER blocks the request on that: it only does a synchronous
    refresh the very first time a product has zero cached rows (so the UI has
    something to show), and otherwise serves whatever is cached immediately while
    kicking off a background refresh thread for any stale product — matching the
    same fire-and-forget pattern used for RAGAS scoring."""
    products = [product] if product else list(RELEASE_NOTES_SOURCES.keys())
    try:
        with get_db() as conn:
            stale_products = []
            for p in products:
                if p not in RELEASE_NOTES_SOURCES:
                    continue
                with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                    cur.execute("SELECT MAX(fetched_at) AS latest FROM release_notes_cache WHERE product=%s", (p,))
                    latest = cur.fetchone()["latest"]
                if not latest:
                    # Never fetched before — block once so the UI isn't empty.
                    _release_notes_refresh(p)
                elif (_dt.datetime.now() - latest).total_seconds() > max_age_hours * 3600:
                    stale_products.append(p)
            for p in stale_products:
                threading.Thread(target=_release_notes_refresh, args=(p,), daemon=True).start()
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute(f"""SELECT product, period, title, description, source_url, fetched_at
                               FROM release_notes_cache WHERE product = ANY(%s)
                               ORDER BY product, id""", (products,))
                rows = [dict(r) for r in cur.fetchall()]
        for r in rows:
            r["fetched_at"] = str(r["fetched_at"])
        return {"entries": rows}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ── Email (SMTP) ─────────────────────────────────────────────────────────────
# Requires SMTP_HOST, SMTP_PORT, SMTP_USER, SMTP_PASS, SMTP_FROM in .env.
# For Gmail: SMTP_HOST=smtp.gmail.com, SMTP_PORT=587, SMTP_USER=you@gmail.com,
# SMTP_PASS=<app password, not your regular password>, SMTP_FROM=you@gmail.com
SMTP_HOST = os.getenv("SMTP_HOST", "")
SMTP_PORT = int(os.getenv("SMTP_PORT", "587"))
SMTP_USER = os.getenv("SMTP_USER", "")
SMTP_PASS = os.getenv("SMTP_PASS", "")
SMTP_FROM = os.getenv("SMTP_FROM", SMTP_USER)
EMAIL_ENABLED = bool(SMTP_HOST and SMTP_USER and SMTP_PASS)

def send_email(to_email: str, subject: str, body_html: str) -> dict:
    """Send an email via SMTP. Returns {ok, error}. No-ops safely if not configured."""
    if not EMAIL_ENABLED:
        return {"ok": False, "error": "Email not configured. Set SMTP_HOST, SMTP_USER, SMTP_PASS in .env"}
    try:
        import smtplib
        from email.mime.text import MIMEText
        from email.mime.multipart import MIMEMultipart
        msg = MIMEMultipart("alternative")
        msg["Subject"] = subject
        msg["From"] = SMTP_FROM
        msg["To"] = to_email
        msg.attach(MIMEText(body_html, "html"))
        with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=10) as server:
            server.starttls()
            server.login(SMTP_USER, SMTP_PASS)
            server.sendmail(SMTP_FROM, [to_email], msg.as_string())
        return {"ok": True, "error": None}
    except Exception as e:
        return {"ok": False, "error": str(e)}

def email_template(title: str, message: str, cta_label: str = None, cta_url: str = None) -> str:
    cta = f'<tr><td style="padding-top:20px"><a href="{cta_url}" style="background:#1473E6;color:#fff;padding:10px 22px;border-radius:8px;text-decoration:none;font-weight:600;font-size:14px;">{cta_label}</a></td></tr>' if cta_label else ""
    return f"""
    <table style="font-family:Arial,sans-serif;max-width:480px;margin:0 auto;">
      <tr><td style="padding-bottom:6px;"><span style="background:#FA0F00;color:#fff;font-weight:800;border-radius:6px;padding:4px 9px;font-size:13px;">N</span>
        <span style="font-weight:700;font-size:15px;color:#1A1A1A;margin-left:6px;">Nexus</span></td></tr>
      <tr><td style="padding:14px 0 6px;"><h2 style="margin:0;font-size:18px;color:#1A1A1A;">{title}</h2></td></tr>
      <tr><td style="font-size:14px;color:#444;line-height:1.6;">{message}</td></tr>
      {cta}
      <tr><td style="padding-top:24px;font-size:11px;color:#999;">Adobe Internal · Nexus Learning Platform</td></tr>
    </table>
    """

# ── Database ──────────────────────────────────────────────────────────────────

@contextmanager
def get_db():
    conn = psycopg2.connect(DATABASE_URL)
    try:
        yield conn
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()

def db_log_llm(agent_name, model, input_tokens, output_tokens, latency_ms, success, error=None):
    try:
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "INSERT INTO llm_logs (agent_name, model, input_tokens, output_tokens, latency_ms, success, error) VALUES (%s,%s,%s,%s,%s,%s,%s)",
                    (agent_name, model, input_tokens, output_tokens, latency_ms, success, error)
                )
    except Exception as e:
        print(f"DB log error: {e}")

def db_log_telemetry(persona, event_type, module, detail):
    try:
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "INSERT INTO telemetry (persona, event_type, module, detail) VALUES (%s,%s,%s,%s)",
                    (persona, event_type, module, detail)
                )
    except Exception as e:
        print(f"DB telemetry error: {e}")

def db_log_guardrail(word_count, has_one_question, avoids_direct_answer, score, issue, response_preview):
    try:
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "INSERT INTO guardrail_logs (word_count, has_one_question, avoids_direct_answer, score, issue, response_preview) VALUES (%s,%s,%s,%s,%s,%s)",
                    (word_count, has_one_question, avoids_direct_answer, score, issue, response_preview[:100] if response_preview else None)
                )
    except Exception as e:
        print(f"DB guardrail error: {e}")

def db_save_summary(user_name, session_type, module, summary):
    try:
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "INSERT INTO session_summaries (user_name, session_type, module, summary) VALUES (%s,%s,%s,%s)",
                    (user_name, session_type, module, summary)
                )
    except Exception as e:
        print(f"DB summary error: {e}")

def db_update_confidence(user_name, module, score):
    try:
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    """INSERT INTO confidence_scores (user_name, module, score)
                       VALUES (%s,%s,%s)
                       ON CONFLICT DO NOTHING""",
                    (user_name, module, score)
                )
    except Exception as e:
        print(f"DB confidence error: {e}")

def db_get_confidence(user_name, module):
    """Return the latest stored confidence score for (user, module), or None."""
    try:
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    """SELECT score FROM confidence_scores
                       WHERE user_name=%s AND module=%s
                       ORDER BY id DESC LIMIT 1""",
                    (user_name, module)
                )
                row = cur.fetchone()
                return float(row[0]) if row and row[0] is not None else None
    except Exception as e:
        print(f"DB confidence read error: {e}")
        return None

# ── Models ────────────────────────────────────────────────────────────────────

class Message(BaseModel):
    role: str
    content: Any  # str or list (multimodal)

class AgentRequest(BaseModel):
    messages: List[Message]
    system: str
    max_tokens: int = 1000
    temperature: float = 0.7
    agent_name: str = "Agent"
    prefer_groq: bool = True  # use Groq if available, else Anthropic

class RAGRequest(BaseModel):
    query: str
    module: str = ""
    track: Optional[str] = None

class JudgeRequest(BaseModel):
    response: str

# ── Health ────────────────────────────────────────────────────────────────────

@app.get("/")
def health():
    db_ok = False
    try:
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT 1")
                db_ok = True
    except Exception:
        pass
    # Actually test the PRIMARY provider (OpenAI) with a minimal call so the health
    # check reflects real connectivity; Groq is tested too as the fallback provider.
    import requests as _hr
    OPENAI_KEY = os.getenv("OPENAI_API_KEY", "")
    openai_ok = False
    openai_model = OPENAI_MODEL
    if OPENAI_KEY:
        try:
            r = _hr.post(OPENAI_URL,
                headers={"Authorization": f"Bearer {OPENAI_KEY}", "Content-Type": "application/json"},
                json={"model": OPENAI_MODEL, "max_tokens": 5,
                      "messages": [{"role": "user", "content": "hi"}]},
                timeout=8)
            openai_ok = r.status_code == 200
            if not openai_ok:
                err = r.json().get("error", {})
                openai_model = err.get("message", f"HTTP {r.status_code}")
        except Exception as e:
            openai_model = str(e)[:80]

    groq_ok = False
    groq_model = GROQ_MODEL
    if GROQ_KEY:
        try:
            r = _hr.post(GROQ_URL,
                headers={"Authorization": f"Bearer {GROQ_KEY}", "Content-Type": "application/json"},
                json={"model": GROQ_MODEL, "max_tokens": 5, "include_reasoning": False,
                      "messages": [{"role": "user", "content": "hi"}]},
                timeout=8)
            groq_ok = r.status_code == 200
            if not groq_ok:
                err = r.json().get("error", {})
                groq_model = err.get("message", f"HTTP {r.status_code}")
        except Exception as e:
            groq_model = str(e)[:80]
    return {
        "status": "ok",
        "openai": openai_ok,
        "openai_model": openai_model,
        "groq": groq_ok,
        "groq_model": groq_model,
        "anthropic": bool(ANTHROPIC_KEY),
        "github":   bool(GITHUB_TOKEN),
        "database": db_ok,
    }

# ── Admin: all users ─────────────────────────────────────────────────────────

@app.get("/api/admin/users")
def list_all_users(_user: dict = Depends(require_persona("admin"))):
    """Return every registered account for the Admin users table.
    Protected: requires a valid IMS session with the admin persona."""
    try:
        with get_db() as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute("""
                    SELECT
                        o.id,
                        o.name,
                        o.preferred_name,
                        o.email,
                        o.role,
                        o.team,
                        o.manager,
                        o.status,
                        o.joining_date,
                        o.active_track,
                        COALESCE(o.capstone_completed, FALSE) AS capstone_completed,
                        o.created_at,
                        -- module progress count
                        (SELECT COUNT(*) FROM user_module_progress p
                         WHERE p.member_name = o.name) AS modules_done,
                        -- points total
                        (SELECT COALESCE(SUM(points),0) FROM points_ledger pl
                         WHERE pl.member_name = o.name) AS total_points,
                        -- last activity
                        (SELECT MAX(created_at) FROM telemetry t
                         WHERE t.persona NOT IN ('nj','exp','mgr','admin')
                           AND t.created_at > o.created_at - INTERVAL '1 day'
                         LIMIT 1) AS last_activity
                    FROM onboarding_requests o
                    ORDER BY o.created_at DESC
                """)
                rows = cur.fetchall()
        # Serialise dates
        users = []
        for r in rows:
            u = dict(r)
            for k in ('joining_date','created_at','last_activity'):
                if u.get(k): u[k] = str(u[k])
            users.append(u)
        return {"users": users, "total": len(users)}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/api/admin/users/{user_id}")
def delete_user(user_id: int, _user: dict = Depends(require_persona("admin"))):
    """Hard-delete a user account (admin only). Protected: requires admin IMS session."""
    try:
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("DELETE FROM onboarding_requests WHERE id=%s", (user_id,))
        return {"ok": True}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.put("/api/admin/users/{user_id}/status")
def set_user_status(user_id: int, body: dict = Body(...), _user: dict = Depends(require_persona("admin"))):
    """Approve or decline a user from the admin users table. Protected: requires admin IMS session."""
    new_status = body.get("status","pending")
    try:
        with get_db() as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute(
                    "UPDATE onboarding_requests SET status=%s WHERE id=%s RETURNING name,email",
                    (new_status, user_id))
                row = cur.fetchone()
        if row and new_status in ("approved","declined"):
            # Create in-app notification for the user
            try:
                with get_db() as conn2:
                    with conn2.cursor() as c2:
                        msg = ("Your Nexus account has been approved. Sign in now."
                               if new_status=="approved"
                               else "Your onboarding request was not approved. Contact your manager.")
                        c2.execute("""INSERT INTO notifications
                                      (member_name,manager,type,title,message)
                                      VALUES (%s,%s,%s,%s,%s)""",
                                   (row["name"],None,new_status,
                                    "Account approved" if new_status=="approved" else "Account declined",
                                    msg))
            except Exception:
                pass
        return {"ok": True, "status": new_status}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# ── Admin: HR directory provisioning (Excel roster upload) ────────────────────
# An admin uploads an .xlsx roster; we parse + validate, then apply a DELTA
# (insert / update / soft-delete / reactivate) against employee_directory in ONE
# transaction, recording a batch summary + field-level change log. No blind
# overwrite: departures are soft-deleted (is_active=FALSE), history preserved.
import re as _re_dir

# Excel header (normalized: lowercased, non-alphanumerics stripped) → DB column.
DIRECTORY_HEADER_MAP = {
    "email": "email", "firstname": "first_name", "lastname": "last_name",
    "dateofjoining": "doj", "role": "role", "location": "location",
    "manager": "manager_name", "manageremail": "manager_email",
    "team": "team", "primaryskill": "primary_skill", "resourceemail": "resource_email",
}
DIRECTORY_REQUIRED_HEADERS = {"email", "dateofjoining"}
DIRECTORY_COMPARE_FIELDS = ["first_name", "last_name", "doj", "role", "location",
                            "manager_name", "manager_email", "team", "primary_skill", "resource_email"]
DIRECTORY_MAX_ROWS = 20000
_EMAIL_RE = _re_dir.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")


def _norm_header(h) -> str:
    return _re_dir.sub(r"[^a-z0-9]", "", str(h or "").strip().lower())


def _parse_doj(v):
    """Coerce an Excel cell into a datetime.date, or None if unparseable."""
    from datetime import datetime, date
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y", "%d/%m/%Y", "%Y/%m/%d", "%m-%d-%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _norm_val(v) -> str:
    """Normalized string form for change-detection + change-log storage."""
    from datetime import date, datetime
    if v is None:
        return ""
    if isinstance(v, (date, datetime)):
        return v.isoformat()[:10]
    return str(v).strip()


def _read_directory_xlsx(raw: bytes):
    """Parse an uploaded .xlsx into (rows, errors). data_only=True → cached
    values only, so no formulas execute. Each row is a dict keyed by DB column."""
    import io
    import openpyxl
    errors = []
    try:
        wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
    except Exception as e:
        return [], [f"Could not read the file as .xlsx: {e}"]
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return [], ["The sheet is empty."]
    idx_to_field, present = {}, set()
    for i, h in enumerate(header_row):
        key = _norm_header(h)
        if key in DIRECTORY_HEADER_MAP:
            idx_to_field[i] = DIRECTORY_HEADER_MAP[key]
            present.add(key)
    missing = DIRECTORY_REQUIRED_HEADERS - present
    if missing:
        return [], [f"Missing required column(s): {', '.join(sorted(missing))}"]
    rows = []
    for rownum, raw_row in enumerate(rows_iter, start=2):
        if raw_row is None or all(c is None or str(c).strip() == "" for c in raw_row):
            continue  # skip blank lines
        rec = {f: None for f in DIRECTORY_HEADER_MAP.values()}
        for i, field in idx_to_field.items():
            rec[field] = raw_row[i] if i < len(raw_row) else None
        rec["email"] = str(rec["email"]).strip().lower() if rec["email"] else ""
        rec["manager_email"] = str(rec["manager_email"]).strip().lower() if rec["manager_email"] else None
        rec["doj"] = _parse_doj(rec["doj"])
        for f in ("first_name", "last_name", "role", "location", "manager_name",
                  "team", "primary_skill", "resource_email"):
            if rec[f] is not None:
                rec[f] = str(rec[f]).strip() or None
        rec["_row"] = rownum
        if not rec["email"] or not _EMAIL_RE.match(rec["email"]):
            errors.append(f"Row {rownum}: invalid or missing email ({rec['email']!r})")
            continue
        if len(rows) >= DIRECTORY_MAX_ROWS:
            errors.append(f"Row limit exceeded (max {DIRECTORY_MAX_ROWS}).")
            break
        rows.append(rec)
    return rows, errors


@app.post("/api/admin/directory/upload")
async def upload_directory(file: UploadFile = File(...),
                           user: dict = Depends(require_persona("admin"))):
    """Validate + apply an Excel roster as an atomic delta. Rejects the whole
    file (no writes) on any validation error or duplicate email."""
    import hashlib
    filename = file.filename or "upload.xlsx"
    if not filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Please upload a .xlsx file.")
    raw = await file.read()
    if not raw:
        raise HTTPException(status_code=400, detail="The uploaded file is empty.")
    if len(raw) > 15 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="File too large (max 15 MB).")
    checksum = hashlib.sha256(raw).hexdigest()

    rows, errors = _read_directory_xlsx(raw)
    # Duplicate emails within the same file → reject (safer than last-wins).
    seen_file, dupes = {}, []
    for r in rows:
        if r["email"] in seen_file:
            dupes.append(r["email"])
        seen_file[r["email"]] = r
    if dupes:
        errors.append("Duplicate email(s) in file: " + ", ".join(sorted(set(dupes))))
    if errors:
        raise HTTPException(status_code=422, detail={
            "message": "Upload rejected — fix the file and re-upload. No changes were made.",
            "errors": errors[:50]})
    if not rows:
        raise HTTPException(status_code=422, detail={"message": "No valid rows found.", "errors": []})

    uploaded_by = user.get("email") or "admin"
    n_insert = n_update = n_deactivate = n_reactivate = 0
    try:
        with get_db() as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                # 1) batch row first (its id is the FK for change-log rows)
                cur.execute("""INSERT INTO directory_upload_batches
                    (filename, uploaded_by, row_count, checksum, status)
                    VALUES (%s,%s,%s,%s,'validated') RETURNING id""",
                    (filename, uploaded_by, len(rows), checksum))
                batch_id = cur.fetchone()["id"]
                # 2) snapshot current directory
                cur.execute("SELECT * FROM employee_directory")
                current = {row["email"]: row for row in cur.fetchall()}
                file_emails = set(seen_file.keys())
                # 3) insert / update / reactivate
                for email, rec in seen_file.items():
                    existing = current.get(email)
                    if existing is None:
                        cur.execute("""INSERT INTO employee_directory
                            (email, first_name, last_name, doj, role, location,
                             manager_name, manager_email, team, primary_skill, resource_email,
                             is_active, first_seen_batch_id, last_seen_batch_id, updated_at)
                            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,TRUE,%s,%s,NOW())""",
                            (email, rec["first_name"], rec["last_name"], rec["doj"], rec["role"],
                             rec["location"], rec["manager_name"], rec["manager_email"], rec["team"],
                             rec["primary_skill"], rec["resource_email"], batch_id, batch_id))
                        cur.execute("INSERT INTO directory_change_log (batch_id,email,change_type) VALUES (%s,%s,'insert')",
                                    (batch_id, email))
                        n_insert += 1
                    else:
                        changes = [(f, existing.get(f), rec.get(f)) for f in DIRECTORY_COMPARE_FIELDS
                                   if _norm_val(existing.get(f)) != _norm_val(rec.get(f))]
                        reactivated = not existing["is_active"]
                        if changes or reactivated:
                            cur.execute("""UPDATE employee_directory SET
                                first_name=%s,last_name=%s,doj=%s,role=%s,location=%s,
                                manager_name=%s,manager_email=%s,team=%s,primary_skill=%s,resource_email=%s,
                                is_active=TRUE,last_seen_batch_id=%s,updated_at=NOW() WHERE email=%s""",
                                (rec["first_name"], rec["last_name"], rec["doj"], rec["role"], rec["location"],
                                 rec["manager_name"], rec["manager_email"], rec["team"], rec["primary_skill"],
                                 rec["resource_email"], batch_id, email))
                            for f, oldv, newv in changes:
                                cur.execute("""INSERT INTO directory_change_log
                                    (batch_id,email,change_type,field_name,old_value,new_value)
                                    VALUES (%s,%s,'update',%s,%s,%s)""",
                                    (batch_id, email, f, _norm_val(oldv), _norm_val(newv)))
                            if reactivated:
                                cur.execute("INSERT INTO directory_change_log (batch_id,email,change_type) VALUES (%s,%s,'reactivate')",
                                            (batch_id, email))
                                n_reactivate += 1
                            if changes:
                                n_update += 1
                        else:
                            cur.execute("UPDATE employee_directory SET last_seen_batch_id=%s WHERE email=%s",
                                        (batch_id, email))
                # 4) soft-delete active emails absent from the new file
                for email, existing in current.items():
                    if existing["is_active"] and email not in file_emails:
                        cur.execute("UPDATE employee_directory SET is_active=FALSE, updated_at=NOW() WHERE email=%s", (email,))
                        cur.execute("INSERT INTO directory_change_log (batch_id,email,change_type) VALUES (%s,%s,'deactivate')",
                                    (batch_id, email))
                        n_deactivate += 1
                # 5) finalize batch summary
                report = {"inserted": n_insert, "updated": n_update,
                          "deactivated": n_deactivate, "reactivated": n_reactivate}
                cur.execute("""UPDATE directory_upload_batches SET status='applied',
                    n_insert=%s,n_update=%s,n_deactivate=%s,n_reactivate=%s,report=%s WHERE id=%s""",
                    (n_insert, n_update, n_deactivate, n_reactivate, json.dumps(report), batch_id))
        return {"ok": True, "batch_id": batch_id, "filename": filename, "row_count": len(rows), **report}
    except HTTPException:
        raise
    except Exception as e:
        # Whole transaction rolled back → directory unchanged.
        raise HTTPException(status_code=500, detail=f"Upload failed, no changes applied: {e}")


@app.delete("/api/admin/directory/wipe")
def wipe_directory(confirm: str = "", user: dict = Depends(require_persona("admin"))):
    """Hard reset — permanently deletes the entire HR roster (employee_directory,
    its change log, and upload batch history). Irreversible; requires the
    caller to pass ?confirm=WIPE to guard against accidental calls."""
    if confirm != "WIPE":
        raise HTTPException(400, 'Pass ?confirm=WIPE to actually delete everything.')
    conn = psycopg2.connect(DATABASE_URL)
    try:
        with conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as c:
                c.execute("SELECT COUNT(*) AS n FROM employee_directory")
                n_members = c.fetchone()["n"]
                c.execute("DELETE FROM directory_change_log")
                c.execute("DELETE FROM employee_directory")
                c.execute("DELETE FROM directory_upload_batches")
    finally:
        conn.close()
    return {"ok": True, "deleted_members": n_members}


@app.get("/api/admin/directory")
def list_directory(active: Optional[str] = None, q: Optional[str] = None,
                   _user: dict = Depends(require_persona("admin"))):
    """List the employee directory. ?active=true|false filters; ?q= searches name/email/team."""
    clauses, params = [], []
    if active == "true":
        clauses.append("is_active=TRUE")
    elif active == "false":
        clauses.append("is_active=FALSE")
    if q:
        like = f"%{q.strip().lower()}%"
        clauses.append("(LOWER(email) LIKE %s OR LOWER(COALESCE(first_name,'')) LIKE %s OR LOWER(COALESCE(last_name,'')) LIKE %s OR LOWER(COALESCE(team,'')) LIKE %s)")
        params += [like, like, like, like]
    where = ("WHERE " + " AND ".join(clauses)) if clauses else ""
    with get_db() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(f"SELECT * FROM employee_directory {where} ORDER BY is_active DESC, last_name, first_name LIMIT 2000", params)
            rows = cur.fetchall()
    out = []
    for r in rows:
        d = dict(r)
        for k in ("doj", "created_at", "updated_at"):
            if d.get(k):
                d[k] = str(d[k])
        out.append(d)
    return {"employees": out, "total": len(out)}


@app.get("/api/admin/directory/batches")
def list_directory_batches(_user: dict = Depends(require_persona("admin"))):
    """Audit history: recent Excel upload batches with their summary counts."""
    with get_db() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute("SELECT * FROM directory_upload_batches ORDER BY id DESC LIMIT 100")
            rows = cur.fetchall()
    out = []
    for r in rows:
        d = dict(r)
        if d.get("uploaded_at"):
            d["uploaded_at"] = str(d["uploaded_at"])
        out.append(d)
    return {"batches": out}


@app.get("/api/directory/my-team")
def get_my_team(manager_email: str = None, manager_name: str = None):
    """Return all active employees who report to this manager, joined with
    their onboarding_requests app-state and certifications."""
    if not manager_email and not manager_name:
        raise HTTPException(400, "Provide manager_email or manager_name")
    conn = psycopg2.connect(DATABASE_URL)
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as c:
            # Match on email OR name — whichever is populated in the directory
            clauses, params = [], []
            if manager_email:
                clauses.append("LOWER(manager_email)=LOWER(%s)")
                params.append(manager_email.strip())
            if manager_name:
                clauses.append("LOWER(manager_name)=LOWER(%s)")
                params.append(manager_name.strip())
            where = "(" + " OR ".join(clauses) + ") AND is_active=TRUE" if clauses else "is_active=TRUE"
            c.execute(f"SELECT * FROM employee_directory WHERE {where} ORDER BY last_name, first_name", params)
            dir_rows = c.fetchall()

            # Get app-state (persona, track, capstone, preferred_name) for each
            emails = [r["email"] for r in dir_rows]
            app_state = {}
            if emails:
                c.execute("""SELECT email, id, preferred_name, joining_date, role AS app_role,
                                    COALESCE(active_track,'rtcdp') AS active_track,
                                    COALESCE(capstone_completed,FALSE) AS capstone_completed,
                                    status, username, avatar_emoji, avatar_color
                             FROM onboarding_requests WHERE LOWER(email)=ANY(%s)""",
                          ([e.lower() for e in emails],))
                for row in c.fetchall():
                    app_state[row["email"].lower()] = dict(row)

            # Get certifications
            cert_map = {}
            if emails:
                c.execute("""SELECT LOWER(email) AS email, cert_name, cert_type, status,
                                    TO_CHAR(issued_date,'YYYY-MM-DD') AS issued_date,
                                    TO_CHAR(expiry_date,'YYYY-MM-DD') AS expiry_date,
                                    days_remaining
                             FROM user_certifications WHERE LOWER(email)=ANY(%s)
                             ORDER BY email, expiry_date NULLS LAST""",
                          ([e.lower() for e in emails],))
                for row in c.fetchall():
                    cert_map.setdefault(row["email"], []).append(dict(row))
    finally:
        conn.close()

    members = []
    for r in dir_rows:
        d = dict(r)
        email_l = d["email"].lower()
        for k in ("doj","created_at","updated_at"):
            if d.get(k): d[k] = str(d[k])
        d["full_name"] = f"{d.get('first_name') or ''} {d.get('last_name') or ''}".strip()
        app = app_state.get(email_l, {})
        d["app_id"]            = app.get("id")
        d["preferred_name"]    = app.get("preferred_name")
        d["active_track"]      = app.get("active_track","rtcdp")
        d["capstone_completed"]= app.get("capstone_completed",False)
        d["app_status"]        = app.get("status")          # approved / pending
        d["username"]          = app.get("username")
        d["certs"]             = cert_map.get(email_l, [])
        members.append(d)

    return {"members": members, "total": len(members)}


@app.get("/api/admin/directory/batches/{batch_id}")
def get_directory_batch(batch_id: int, _user: dict = Depends(require_persona("admin"))):
    """A single batch plus its field-level change log (the audit detail view)."""
    with get_db() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute("SELECT * FROM directory_upload_batches WHERE id=%s", (batch_id,))
            batch = cur.fetchone()
            if not batch:
                raise HTTPException(status_code=404, detail="Batch not found.")
            cur.execute("SELECT * FROM directory_change_log WHERE batch_id=%s ORDER BY id LIMIT 5000", (batch_id,))
            changes = cur.fetchall()
    b = dict(batch)
    if b.get("uploaded_at"):
        b["uploaded_at"] = str(b["uploaded_at"])
    ch = []
    for c in changes:
        d = dict(c)
        if d.get("changed_at"):
            d["changed_at"] = str(d["changed_at"])
        ch.append(d)
    return {"batch": b, "changes": ch}


# ── Manager hierarchy + role-based learning journey (Excel upload, admin) ──────
# Simpler than employee_directory's upload: these are small reference/lookup
# tables (not living employee records), so each upload does a full validate-then
# -replace rather than a diff/change-log. A batch row is still recorded for
# traceability (who uploaded what, when, how many rows).

MGR_HIERARCHY_HEADER_MAP = {
    "managername": "manager_name", "manager": "manager_name",
    "reportsto": "reports_to",
    "trackfocus": "track_focus", "track": "track_focus",
    "notes": "notes",
}
MGR_HIERARCHY_REQUIRED = {"managername", "manager"}  # at least one must be present — checked below

LEARNING_JOURNEY_HEADER_MAP = {
    "role": "role", "priority": "priority",
    # "Priority Meaning" is deliberately NOT mapped — it's derivable from the
    # priority number alone (see PRIORITY_MEANING in crossskill.py) and mapping
    # it to the same target as "Notes" would let whichever column comes second
    # silently overwrite the first.
    "targetproficiency": "target_proficiency",
    "tracks": "tracks", "track": "tracks",
    "notes": "notes",
}
LEARNING_JOURNEY_REQUIRED = {"role", "priority"}


def _read_simple_xlsx(raw: bytes, header_map: dict, required_norm_keys: set, sheet_name: str = None):
    """Generic small-reference-table parser (mirrors _read_directory_xlsx's
    header-matching approach, without the diff/audit machinery that table needs)."""
    import io
    import openpyxl
    try:
        wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
    except Exception as e:
        return [], [f"Could not read the file as .xlsx: {e}"]
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.worksheets[0]
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return [], ["The sheet is empty."]
    idx_to_field, present = {}, set()
    for i, h in enumerate(header_row):
        key = _norm_header(h)
        if key in header_map:
            idx_to_field[i] = header_map[key]
            present.add(key)
    if required_norm_keys and not (required_norm_keys & present):
        return [], [f"Missing required column — need one of: {', '.join(sorted(required_norm_keys))}"]
    rows = []
    for rownum, raw_row in enumerate(rows_iter, start=2):
        if raw_row is None or all(c is None or str(c).strip() == "" for c in raw_row):
            continue
        rec = {}
        for i, field in idx_to_field.items():
            val = raw_row[i] if i < len(raw_row) else None
            if isinstance(val, str):
                val = val.strip() or None
            rec[field] = val
        rec["_row"] = rownum
        rows.append(rec)
    return rows, []


@app.post("/api/admin/manager-hierarchy/upload")
async def upload_manager_hierarchy(file: UploadFile = File(...),
                                   user: dict = Depends(require_persona("admin"))):
    """Replace the manager_hierarchy table from an uploaded .xlsx (Manager Name,
    Reports To, Track Focus, Notes columns)."""
    raw = await file.read()
    rows, errors = _read_simple_xlsx(raw, MGR_HIERARCHY_HEADER_MAP, MGR_HIERARCHY_REQUIRED)
    if errors:
        raise HTTPException(status_code=400, detail="; ".join(errors))
    clean = [r for r in rows if r.get("manager_name")]
    if not clean:
        raise HTTPException(status_code=400, detail="No rows with a Manager Name were found.")
    try:
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("DELETE FROM manager_hierarchy")
                for r in clean:
                    cur.execute(
                        """INSERT INTO manager_hierarchy (manager_name, reports_to, track_focus, notes)
                           VALUES (%s,%s,%s,%s)
                           ON CONFLICT (manager_name) DO UPDATE SET
                             reports_to=EXCLUDED.reports_to, track_focus=EXCLUDED.track_focus,
                             notes=EXCLUDED.notes, updated_at=NOW()""",
                        (r["manager_name"], r.get("reports_to"), r.get("track_focus"), r.get("notes")))
                cur.execute(
                    """INSERT INTO org_data_upload_batches (dataset, filename, uploaded_by, row_count)
                       VALUES ('manager_hierarchy', %s, %s, %s)""",
                    (file.filename, user.get("name") or user.get("email"), len(clean)))
        return {"ok": True, "rows_applied": len(clean)}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/admin/manager-hierarchy")
def list_manager_hierarchy(user: dict = Depends(get_current_user)):
    try:
        with get_db() as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute("SELECT manager_name, reports_to, track_focus, notes, updated_at FROM manager_hierarchy ORDER BY manager_name")
                rows = [dict(r) for r in cur.fetchall()]
        for r in rows:
            if r.get("updated_at"): r["updated_at"] = str(r["updated_at"])
        return {"managers": rows}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/admin/learning-journey/upload")
async def upload_learning_journey(file: UploadFile = File(...),
                                  user: dict = Depends(require_persona("admin"))):
    """Replace the role_learning_journey table from an uploaded .xlsx (Role,
    Priority, Target Proficiency, Track(s), Notes columns — Track(s) as a
    comma-separated list)."""
    raw = await file.read()
    rows, errors = _read_simple_xlsx(raw, LEARNING_JOURNEY_HEADER_MAP, LEARNING_JOURNEY_REQUIRED)
    if errors:
        raise HTTPException(status_code=400, detail="; ".join(errors))
    clean = []
    for r in rows:
        role = r.get("role")
        try:
            priority = int(r.get("priority"))
        except (TypeError, ValueError):
            continue
        if not role or not (1 <= priority <= 5):
            continue
        tracks_raw = r.get("tracks") or ""
        tracks = [t.strip() for t in str(tracks_raw).split(",") if t.strip()]
        clean.append({"role": str(role).strip(), "priority": priority,
                      "target_proficiency": r.get("target_proficiency"),
                      "tracks": tracks, "notes": r.get("notes")})
    if not clean:
        raise HTTPException(status_code=400, detail="No valid rows found (need Role and Priority 1-5).")
    try:
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("DELETE FROM role_learning_journey")
                for r in clean:
                    cur.execute(
                        """INSERT INTO role_learning_journey (role, priority, target_proficiency, tracks, notes)
                           VALUES (%s,%s,%s,%s,%s)
                           ON CONFLICT (role, priority) DO UPDATE SET
                             target_proficiency=EXCLUDED.target_proficiency, tracks=EXCLUDED.tracks,
                             notes=EXCLUDED.notes, updated_at=NOW()""",
                        (r["role"], r["priority"], r["target_proficiency"], json.dumps(r["tracks"]), r["notes"]))
                cur.execute(
                    """INSERT INTO org_data_upload_batches (dataset, filename, uploaded_by, row_count)
                       VALUES ('role_learning_journey', %s, %s, %s)""",
                    (file.filename, user.get("name") or user.get("email"), len(clean)))
        return {"ok": True, "rows_applied": len(clean)}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/admin/learning-journey")
def list_learning_journey(user: dict = Depends(get_current_user)):
    try:
        with get_db() as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute("SELECT role, priority, target_proficiency, tracks, notes, updated_at FROM role_learning_journey ORDER BY role, priority")
                rows = [dict(r) for r in cur.fetchall()]
        for r in rows:
            if r.get("updated_at"): r["updated_at"] = str(r["updated_at"])
        return {"journey": rows}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ── In-app single-row editing of the matrix (complements the bulk /upload) ─────
class LearningJourneyRow(BaseModel):
    role: str
    priority: int
    target_proficiency: str = ""
    tracks: List[str] = []
    notes: str = ""


@app.post("/api/admin/learning-journey")
def upsert_learning_journey_row(body: LearningJourneyRow,
                                user: dict = Depends(require_persona("admin"))):
    """Create or update one role×priority row from the admin grid. Upserts on the
    (role, priority) unique key, so editing a cell in-app doesn't need a re-upload."""
    role = (body.role or "").strip()
    if not role or not (1 <= body.priority <= 5):
        raise HTTPException(status_code=422, detail="role and priority (1-5) are required.")
    tracks = [t.strip() for t in (body.tracks or []) if t and t.strip()]
    try:
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    """INSERT INTO role_learning_journey (role, priority, target_proficiency, tracks, notes)
                       VALUES (%s,%s,%s,%s,%s)
                       ON CONFLICT (role, priority) DO UPDATE SET
                         target_proficiency=EXCLUDED.target_proficiency, tracks=EXCLUDED.tracks,
                         notes=EXCLUDED.notes, updated_at=NOW()""",
                    (role, body.priority, (body.target_proficiency or "").strip() or None,
                     json.dumps(tracks), (body.notes or "").strip() or None))
        return {"ok": True, "role": role, "priority": body.priority}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.delete("/api/admin/learning-journey/{role}/{priority}")
def delete_learning_journey_row(role: str, priority: int,
                                user: dict = Depends(require_persona("admin"))):
    """Delete one role×priority row from the matrix."""
    try:
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("DELETE FROM role_learning_journey WHERE role=%s AND priority=%s",
                            (role, priority))
                n = cur.rowcount
        return {"ok": True, "deleted": n}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ── Role aliases: HR/profile role string → canonical matrix role code ──────────
class RoleAliasRow(BaseModel):
    alias: str
    canonical_role: str


@app.get("/api/admin/role-aliases")
def list_role_aliases(user: dict = Depends(get_current_user)):
    try:
        with get_db() as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute("SELECT alias, canonical_role, updated_at FROM role_aliases ORDER BY canonical_role, alias")
                rows = [dict(r) for r in cur.fetchall()]
        for r in rows:
            if r.get("updated_at"): r["updated_at"] = str(r["updated_at"])
        return {"aliases": rows}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/admin/role-aliases")
def upsert_role_alias(body: RoleAliasRow, user: dict = Depends(require_persona("admin"))):
    """Map a free-text HR/profile role to a canonical role_learning_journey code
    (e.g. 'Data Analyst' → 'AEP - DA'). Consulted by the cross-skilling agent."""
    alias = (body.alias or "").strip()
    canon = (body.canonical_role or "").strip()
    if not alias or not canon:
        raise HTTPException(status_code=422, detail="alias and canonical_role are required.")
    try:
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("""CREATE TABLE IF NOT EXISTS role_aliases (
                    id SERIAL PRIMARY KEY, alias VARCHAR(150) UNIQUE NOT NULL,
                    canonical_role VARCHAR(80) NOT NULL, updated_at TIMESTAMP DEFAULT NOW())""")
                cur.execute(
                    """INSERT INTO role_aliases (alias, canonical_role) VALUES (%s,%s)
                       ON CONFLICT (alias) DO UPDATE SET
                         canonical_role=EXCLUDED.canonical_role, updated_at=NOW()""",
                    (alias, canon))
        return {"ok": True, "alias": alias, "canonical_role": canon}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.delete("/api/admin/role-aliases/{alias}")
def delete_role_alias(alias: str, user: dict = Depends(require_persona("admin"))):
    try:
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("DELETE FROM role_aliases WHERE LOWER(alias)=LOWER(%s)", (alias,))
                n = cur.rowcount
        return {"ok": True, "deleted": n}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ── Data validation report (admin diagnostics tab) ─────────────────────────────
def _offtrack_reason(track: str, url: str):
    """Conservative heuristic: does a topic's Experience League URL point at a
    different product than its track? Returns a reason string or None."""
    u = (url or "").lower()
    if not u:
        return None
    if track == "analytics":
        if "/rtcdp" in u or "platform-learn" in u:
            return "links to a CDP / Platform doc"
        if "experience-platform" in u and "query" not in u:
            return "links to an Experience Platform doc"
    else:  # rtcdp and other platform tracks
        if "docs/analytics/" in u or "analytics-platform" in u or "analytics-learn" in u:
            return "links to an Adobe Analytics doc"
    return None


@app.get("/api/admin/validate")
def validate_platform_data(user: dict = Depends(require_persona("admin"))):
    """Read-only data-integrity report across curriculum, the RAG index, and org
    reference data (teams / roles → track / journey resolution). Powers the admin
    'Data Validation' tab. Everything is a check, nothing is mutated."""
    import time as _t
    report = {"generated_at": _t.strftime("%Y-%m-%d %H:%M:%S")}

    # ── Curriculum integrity per track ────────────────────────────────────────
    rows = []
    try:
        with get_db() as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute("""SELECT track, module_id, topic_order, title, el_url
                               FROM curriculum_topics
                               ORDER BY track, module_id, topic_order""")
                rows = [dict(r) for r in cur.fetchall()]
    except Exception as e:
        report["curriculum_error"] = str(e)

    tracks = {}
    for r in rows:
        t = r["track"] or "(none)"
        tk = tracks.setdefault(t, {"modules": {}, "topic_count": 0,
                                   "missing_el_url": 0, "offtrack": []})
        tk["topic_count"] += 1
        tk["modules"][r["module_id"]] = tk["modules"].get(r["module_id"], 0) + 1
        if not (r["el_url"] or "").strip():
            tk["missing_el_url"] += 1
        off = _offtrack_reason(r["track"] or "", r["el_url"])
        if off:
            tk["offtrack"].append({"module_id": r["module_id"], "topic_order": r["topic_order"],
                                   "title": r["title"], "el_url": r["el_url"], "reason": off})

    curriculum_tracks = []
    for t in sorted(tracks):
        tk = tracks[t]
        mod_ids = sorted(tk["modules"])
        issues = []
        if tk["missing_el_url"]:
            issues.append({"level": "warn",
                           "detail": f"{tk['missing_el_url']} topic(s) have no Experience League URL — those lessons can't fetch content or video."})
        if tk["offtrack"]:
            issues.append({"level": "warn",
                           "detail": f"{len(tk['offtrack'])} topic(s) link to another product's docs — review below."})
        gaps = [m for m in range(1, (mod_ids[-1] if mod_ids else 0) + 1) if m not in tk["modules"]]
        if gaps:
            issues.append({"level": "warn", "detail": f"No topics for module(s) {gaps} (gap in the module sequence)."})
        curriculum_tracks.append({
            "track": t, "topic_count": tk["topic_count"],
            "modules": [{"module_id": m, "topics": tk["modules"][m]} for m in mod_ids],
            "missing_el_url": tk["missing_el_url"], "offtrack": tk["offtrack"], "issues": issues,
        })
    # Canonical track set — so tracks with NO content are visible and flagged,
    # not silently absent. Union of the cross-skill vocabulary and whatever's in
    # the DB, so every track a learner could be routed to shows up here.
    try:
        from agents.crossskill import AVAILABLE_TRACKS
        canonical = {k: v.get("name", k) for k, v in AVAILABLE_TRACKS.items()}
    except Exception:
        canonical = {}
    present = {ct["track"]: ct for ct in curriculum_tracks}
    coverage = []
    for code in sorted(set(canonical) | set(present)):
        ct = present.get(code)
        coverage.append({
            "track": code,
            "label": canonical.get(code, code),
            "db_topics": ct["topic_count"] if ct else 0,
            "db_modules": len(ct["modules"]) if ct else 0,
        })

    report["curriculum"] = {"tracks": curriculum_tracks,
                            "tracks_present": sorted(tracks),
                            "coverage": coverage,
                            "total_topics": len(rows)}

    # ── Embeddings / RAG index ────────────────────────────────────────────────
    emb = {"per_track": {}, "total": 0}
    try:
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT track, COUNT(*) FROM doc_embeddings GROUP BY track")
                for tr, n in cur.fetchall():
                    emb["per_track"][tr or "(none)"] = n
                    emb["total"] += n
    except Exception as e:
        emb["error"] = str(e)
    try:
        from agents import vector_store as _vs
        emb["pgvector_available"] = _vs.is_available()
        emb["pgvector_count"] = _vs.store_count() if _vs.is_available() else 0
    except Exception as e:
        emb["pgvector_available"] = False
        emb["pgvector_error"] = str(e)
    report["embeddings"] = emb

    # ── Org reference data + role/team resolution ─────────────────────────────
    org = {}
    roles = []
    try:
        with get_db() as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute("SELECT COUNT(*) n FROM employee_directory")
                org["directory_count"] = cur.fetchone()["n"]
                cur.execute("SELECT team, COUNT(*) n FROM employee_directory WHERE team IS NOT NULL AND team<>'' GROUP BY team ORDER BY n DESC")
                org["teams"] = [dict(r) for r in cur.fetchall()]
                cur.execute("SELECT role, COUNT(*) n FROM employee_directory WHERE role IS NOT NULL AND role<>'' GROUP BY role ORDER BY n DESC")
                roles = [dict(r) for r in cur.fetchall()]
                cur.execute("SELECT COUNT(DISTINCT role) n FROM role_learning_journey")
                org["role_journey_role_count"] = cur.fetchone()["n"]
                cur.execute("SELECT COUNT(*) n FROM role_aliases")
                org["role_aliases_count"] = cur.fetchone()["n"]
                cur.execute("SELECT COUNT(*) n FROM manager_hierarchy")
                org["manager_hierarchy_count"] = cur.fetchone()["n"]
    except Exception as e:
        org["error"] = str(e)

    # Resolve each distinct directory role through the SAME path the cross-skill
    # agent uses, so this reflects what learners actually get.
    try:
        from agents.crossskill import _fetch_role_journey, _resolve_role_alias
    except Exception:
        _fetch_role_journey = _resolve_role_alias = None

    resolution, resolved_head, unresolved_head = [], 0, 0
    for r in roles:
        role, n = r["role"], r["n"]
        matched, via = None, None
        if _fetch_role_journey:
            try:
                journey = _fetch_role_journey(role)
                if journey:
                    matched = journey[0]["role"]
                    via = "alias" if (_resolve_role_alias and _resolve_role_alias(role)) else "fuzzy match"
            except Exception:
                pass
        if matched:
            resolved_head += n
        else:
            unresolved_head += n
        resolution.append({"role": role, "count": n, "resolves": bool(matched),
                           "matched_role": matched, "via": via})
    org["role_resolution"] = resolution
    org["resolved_headcount"] = resolved_head
    org["unresolved_headcount"] = unresolved_head

    # Manager-based resolution: person → their manager → the manager's team focus
    # → matrix role. This is the org's actual model — a generic title like
    # "Technical Consultant" is Data or Journey depending on who they report to,
    # so the manager's track_focus decides the track, not the role string.
    def _tf_to_role(tf):
        t = (tf or "").strip().lower()
        if not t:
            return None
        if "engineering services" in t or t == "es":
            return "ES"
        if t == "de" or "data engineer" in t:
            return "AEP - DE"
        if t == "aa" or "analytics" in t or "aa-sdk" in t or "web sdk" in t:
            return "AA-SDK"
        if t == "da" or "architect" in t:          # data architects
            return "AEP - DA"
        if "rtcdp" in t:
            return "RTCDP"
        return None  # broad / ambiguous (e.g. "Data and Journeys")

    mgr_res = {"by_manager": [], "resolved_headcount": 0, "ambiguous_headcount": 0,
               "unmapped_manager_headcount": 0}
    try:
        with get_db() as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                # Exclude the managers themselves from the learner population —
                # anyone whose own name appears in manager_hierarchy is a manager,
                # not a learner (their reports are the learners).
                cur.execute("""
                    SELECT ed.manager_name AS mgr, COUNT(*) AS n, mh.track_focus AS tf
                    FROM employee_directory ed
                    LEFT JOIN manager_hierarchy mh
                      ON LOWER(TRIM(ed.manager_name)) = LOWER(TRIM(mh.manager_name))
                    WHERE ed.manager_name IS NOT NULL AND ed.manager_name <> ''
                      AND LOWER(TRIM(COALESCE(ed.first_name,'') || ' ' || COALESCE(ed.last_name,'')))
                          NOT IN (SELECT LOWER(TRIM(manager_name)) FROM manager_hierarchy)
                    GROUP BY ed.manager_name, mh.track_focus
                    ORDER BY n DESC""")
                for r in cur.fetchall():
                    n, tf = r["n"], r["tf"]
                    if tf is None:
                        status, role = "manager_missing", None
                        mgr_res["unmapped_manager_headcount"] += n
                    else:
                        role = _tf_to_role(tf)
                        if role:
                            status = "resolved"; mgr_res["resolved_headcount"] += n
                        else:
                            status = "ambiguous"; mgr_res["ambiguous_headcount"] += n
                    mgr_res["by_manager"].append({
                        "manager": r["mgr"], "people": n, "track_focus": tf,
                        "matrix_role": role, "status": status})
    except Exception as e:
        mgr_res["error"] = str(e)
    org["manager_resolution"] = mgr_res

    report["org"] = org

    return report


# ── Agent proxy ───────────────────────────────────────────────────────────────

@app.post("/api/agent")
async def call_agent(req: AgentRequest):
    """Generic LLM proxy for agents that build their own system prompt on the
    frontend (test-out quizzes, capstone hints, practice scenarios, etc.).

    Routes through the provider-agnostic llm_call helper so a missing Groq key
    or a transient 429 transparently fails over to Anthropic instead of dead-
    ending the learner with "Could not generate… check your Groq key". The call
    is synchronous inside a threadpool so we don't block the event loop.
    """
    msgs = [{"role": m.role,
             "content": m.content if isinstance(m.content, str) else str(m.content)}
            for m in req.messages]
    try:
        text = await run_in_threadpool(
            _llm_call, msgs, req.system,
            max_tokens=req.max_tokens,
            prefer=("groq" if getattr(req, "prefer_groq", True) else "anthropic"),
        )
    except Exception as e:
        db_log_llm(req.agent_name, "none", 0, 0, 0, False)
        raise HTTPException(status_code=503, detail=f"LLM generation failed: {e}")

    result = {"text": text, "model": "auto", "input_tokens": 0, "output_tokens": 0}
    db_log_llm(req.agent_name, result["model"], 0, 0, 0, True)
    return result


# ── Notes generation (direct — no proxy complexity) ───────────────────────────
@app.post("/api/notes/generate")
def generate_notes(body: dict = Body(...)):
    """Generate structured study notes for a topic. No auth required."""
    topic   = body.get("topic", "AEP Topic")
    module  = body.get("module", "")
    doc     = body.get("doc_content", "")[:3000]
    track   = body.get("track", "rtcdp")

    system = f"""You are a learning coach converting AEP documentation into structured study notes.
Topic: "{topic}" (Module: {module}, Track: {track.upper()})
Return ONLY valid JSON matching this schema:
{{
  "summary": "3-sentence executive summary",
  "concepts": [{{"title":"...","explanation":"...","example":"..."}}],
  "terms": [{{"term":"...","definition":"..."}}],
  "steps": [{{"n":1,"title":"...","detail":"..."}}],
  "warnings": ["..."],
  "takeaways": ["..."]
}}
Rules: 3-5 concepts, 4-8 terms, steps only if procedural, 1-3 warnings, 3-5 takeaways.
No markdown in values. Write for a learner, not a developer."""

    messages = [{"role": "user", "content": f"Topic: {topic}\nModule: {module}\n\nDoc excerpt:\n{doc}\n\nGenerate study notes."}]
    try:
        raw = _llm_call(messages, system, max_tokens=1000)
        import json as _json
        parsed = _json.loads(raw.replace("```json","").replace("```","").strip())
        return {"ok": True, "notes": parsed}
    except Exception as e:
        return {"ok": False, "error": str(e), "raw": raw if 'raw' in dir() else ""}


# ── Quiz generation (direct) ──────────────────────────────────────────────────
@app.post("/api/quiz/generate")
def generate_quiz_questions(body: dict = Body(...)):
    """Generate 10 MCQ questions for a topic. No auth required."""
    topic      = body.get("topic", "AEP Topic")
    module     = body.get("module", "")
    track      = body.get("track", "rtcdp")
    confidence = float(body.get("confidence", 0.5))

    difficulty = "beginner" if confidence < 0.5 else "advanced" if confidence > 0.7 else "intermediate"

    system = f"""You are generating a 10-question quiz for an AEP learner.
Topic: "{topic}" (Module: {module}, Track: {track.upper()})
Difficulty: {difficulty} (learner confidence: {confidence:.0%})
Return ONLY valid JSON:
{{
  "questions": [
    {{"id":1,"question":"...","options":["A. ...","B. ...","C. ...","D. ..."],"correct":"A","explanation":"..."}}
  ]
}}
Generate exactly 10 questions. Mix definitions, scenarios, and comparisons.
CRITICAL: Never confuse AA vs CJA, RTCDP vs AEP, AJO vs Campaign, Web SDK vs AA."""

    messages = [{"role": "user", "content": f"Generate 10 quiz questions for: {topic} ({track.upper()})"}]
    try:
        raw = _llm_call(messages, system, max_tokens=1400)
        import json as _json
        parsed = _json.loads(raw.replace("```json","").replace("```","").strip())
        return {"ok": True, **parsed}
    except Exception as e:
        return {"ok": False, "error": str(e), "questions": []}


# ── RAG — GitHub AdobeDocs fetch ──────────────────────────────────────────────

REPO_MAP = [
    (["ajo","journey","orchestrat","notification","suppression"], "AdobeDocs/journey-optimizer.en"),
    (["cja","customer journey analytics","attribution","stitching"], "AdobeDocs/customer-journey-analytics.en"),
    (["analytics","report suite","fallout"], "AdobeDocs/analytics.en"),
    (["marketo","email","campaign","lead"], "AdobeDocs/marketo.en"),
]

def pick_repo(query: str, module: str) -> str:
    text = (query + " " + module).lower()
    for keys, repo in REPO_MAP:
        if any(k in text for k in keys):
            return repo
    return "AdobeDocs/experience-platform.en"

def strip_markdown(md: str) -> str:
    import re
    md = re.sub(r'^---[\s\S]*?---\n', '', md)
    md = re.sub(r'!\[.*?\]\(.*?\)', '', md)
    md = re.sub(r'\[([^\]]+)\]\([^)]+\)', r'\1', md)
    md = re.sub(r'`{3}[\s\S]*?`{3}', '', md)
    md = re.sub(r'`[^`]+`', '', md)
    md = re.sub(r'#{1,6}\s+', '', md)
    md = re.sub(r'\*{1,2}([^*]+)\*{1,2}', r'\1', md)
    md = re.sub(r'^\s*[-*>]\s+', '', md, flags=re.MULTILINE)
    md = re.sub(r'\n{3,}', '\n\n', md)
    return md.strip()[:900]

@app.post("/api/rag")
async def retrieve_docs(req: RAGRequest):
    """
    Retrieves relevant AdobeDocs context for the Socratic Agent.
    Tries real semantic (embeddings) search first — if the index has been built
    via build_embeddings_index.py. Falls back to GitHub keyword search if the
    index is empty or the match quality is weak, so this works either way.

    vector_search() is synchronous (psycopg2 / fastembed / pgvector are all
    blocking calls) — run it in a thread pool so it can't stall the single
    event loop for every other request while it's in flight (first-call model
    load or a slow DB round-trip would otherwise freeze the whole server).
    """
    try:
        vec_docs = await run_in_threadpool(vector_search, req.query, track=req.track, top_k=3)
        if vec_docs:
            return {"docs": vec_docs, "source": "embeddings"}
    except Exception as e:
        print(f"vector search failed, falling back to keyword search: {e}")

    repo = pick_repo(req.query, req.module)
    headers = {"Accept": "application/vnd.github.v3+json"}
    if GITHUB_TOKEN:
        headers["Authorization"] = f"token {GITHUB_TOKEN}"

    import re
    stopwords = {"what","when","does","this","that","with","from","have","will","your","about","how"}
    words = list({w for w in re.split(r'\W+', (req.query + " " + req.module).lower())
                  if len(w) > 3 and w not in stopwords})[:4]

    if not words:
        return {"docs": [], "source": "empty_query"}

    q = " ".join(words) + f" repo:{repo} extension:md"

    async with httpx.AsyncClient(timeout=15.0) as client:
        try:
            search = await client.get(
                "https://api.github.com/search/code",
                params={"q": q, "per_page": 5},
                headers=headers,
            )
            data = search.json()
            items = data.get("items", [])
            if not items:
                return {"docs": [], "source": "no_results"}

            docs = []
            for item in items[:2]:
                raw_url = f"https://raw.githubusercontent.com/{repo}/main/{item['path']}"
                raw = await client.get(raw_url, headers=headers)
                if raw.status_code != 200:
                    continue
                content = strip_markdown(raw.text)
                if len(content) < 60:
                    continue
                docs.append({
                    "id":      item["sha"][:7],
                    "title":   item["name"].replace(".md", "").replace("-", " "),
                    "content": content,
                    "url":     item["html_url"],
                    "repo":    repo.split("/")[1],
                })
            return {"docs": docs, "source": "github"}

        except Exception as e:
            return {"docs": [], "source": "error", "error": str(e)}

# ── Judge — LLM-as-judge for Socratic responses ───────────────────────────────

@app.post("/api/judge")
async def judge_response(req: JudgeRequest):
    """
    Evaluate a Socratic Agent response for quality and rule compliance.
    """
    judge_req = AgentRequest(
        messages=[Message(role="user", content=f'Evaluate: "{req.response}"')],
        system="""You are an AI safety evaluator for a Socratic tutoring agent.
Evaluate the response below strictly. Return ONLY valid JSON, nothing else.
{"wordCount":N,"hasOneQuestion":bool,"avoidsDirectAnswer":bool,"isSocratic":bool,"score":1-10,"issue":null|"brief description"}
Rules: Good = asks exactly 1 question, never states the answer, under 55 words, guides reasoning.""",
        max_tokens=150,
        temperature=0.1,
        agent_name="Judge",
    )
    result = await call_agent(judge_req)
    import json
    try:
        text = result["text"].replace("```json", "").replace("```", "").strip()
        return json.loads(text)
    except Exception:
        wc = len(req.response.split())
        has_q = req.response.count("?") == 1
        return {
            "wordCount": wc,
            "hasOneQuestion": has_q,
            "avoidsDirectAnswer": True,
            "isSocratic": has_q,
            "score": 7 if (has_q and wc <= 55) else 4,
            "issue": None if has_q else "Missing question",
        }

# ── Telemetry endpoint ────────────────────────────────────────────────────────

class TelemetryEvent(BaseModel):
    persona: str
    event_type: str
    module: str = ""
    detail: str = ""

@app.post("/api/telemetry")
def log_telemetry(event: TelemetryEvent):
    db_log_telemetry(event.persona, event.event_type, event.module, event.detail)
    return {"ok": True}

# ── Session summary endpoint ──────────────────────────────────────────────────

class SummaryRequest(BaseModel):
    persona: str = "nj"
    module: str = ""
    mode: str = "socratic"
    message_count: int = 0
    summary: str
    created_at: Optional[str] = None
    # legacy fields
    user_name: Optional[str] = None
    session_type: Optional[str] = None

@app.post("/api/summary")
def save_summary(req: SummaryRequest):
    user = req.user_name or req.persona
    stype = req.session_type or req.mode
    db_save_summary(user, stype, req.module, req.summary)
    return {"ok": True}

@app.get("/api/summary/latest")
def get_latest_summary(persona: str = "nj", module: str = ""):
    try:
        with get_db() as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                if module:
                    cur.execute(
                        "SELECT summary, created_at FROM session_summaries WHERE user_name=%s AND module=%s ORDER BY created_at DESC LIMIT 1",
                        (persona, module)
                    )
                else:
                    cur.execute(
                        "SELECT summary, created_at FROM session_summaries WHERE user_name=%s ORDER BY created_at DESC LIMIT 1",
                        (persona,)
                    )
                row = cur.fetchone()
        if row:
            return {"summary": row["summary"], "created_at": str(row["created_at"])}
        return {"summary": None}
    except Exception as e:
        return {"summary": None, "error": str(e)}

# ── Bandwidth (BW) update ─────────────────────────────────────────────────────

class BWUpdate(BaseModel):
    persona: str
    bw: int  # 0-100

@app.post("/api/bw")
def update_bw(req: BWUpdate):
    try:
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "INSERT INTO telemetry (persona, event_type, module, detail) VALUES (%s, %s, %s, %s)",
                    (req.persona, "bw_update", "profile", f"bw={req.bw}")
                )
            conn.commit()
        return {"ok": True, "bw": req.bw}
    except Exception as e:
        return {"ok": False, "error": str(e)}

@app.get("/api/bw/latest")
def get_latest_bw(persona: str = "nj"):
    try:
        with get_db() as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute(
                    "SELECT detail, created_at FROM telemetry WHERE persona=%s AND event_type='bw_update' ORDER BY created_at DESC LIMIT 1",
                    (persona,)
                )
                row = cur.fetchone()
        if row:
            # detail is "bw=85"
            bw = int(row["detail"].split("=")[1])
            return {"bw": bw, "updated_at": str(row["created_at"])}
        return {"bw": None}
    except Exception as e:
        return {"bw": None, "error": str(e)}

# ── DB stats endpoint (for Admin dashboard) ───────────────────────────────────

@app.get("/api/stats")
def get_stats():
    try:
        with get_db() as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute("SELECT COUNT(*) as total, SUM(input_tokens+output_tokens) as tokens, AVG(latency_ms) as avg_latency FROM llm_logs")
                llm = cur.fetchone()
                cur.execute("SELECT COUNT(*) as total FROM telemetry")
                tel = cur.fetchone()
                cur.execute("SELECT COUNT(*) as total, AVG(score) as avg_score FROM guardrail_logs")
                guard = cur.fetchone()
                cur.execute("""SELECT agent_name,
                                      COUNT(*) as calls,
                                      COALESCE(SUM(input_tokens+output_tokens),0) as tokens,
                                      ROUND(AVG(latency_ms)) as avg_latency,
                                      MAX(model) as model,
                                      ROUND(100.0*SUM(CASE WHEN success THEN 1 ELSE 0 END)/NULLIF(COUNT(*),0)) as success_rate
                               FROM llm_logs GROUP BY agent_name ORDER BY calls DESC""")
                by_agent = cur.fetchall()
        return {
            "llm": dict(llm),
            "telemetry": dict(tel),
            "guardrails": dict(guard),
            "by_agent": [dict(r) for r in by_agent],
        }
    except Exception as e:
        return {"error": str(e)}

# ── Guardrail save (just stores pre-computed result) ──────────────────────────

class GuardrailSave(BaseModel):
    word_count: int = 0
    has_one_question: bool = False
    avoids_direct_answer: bool = True
    score: int = 0
    issue: Optional[str] = None
    response_preview: str = ""

@app.post("/api/guardrail/save")
def save_guardrail(req: GuardrailSave):
    db_log_guardrail(
        req.word_count, req.has_one_question,
        req.avoids_direct_answer, req.score,
        req.issue, req.response_preview
    )
    return {"ok": True}

# ── Onboarding / Approval endpoints ──────────────────────────────────────────

class OnboardingRequest(BaseModel):
    name: str
    preferred_name: str
    email: str
    joining_date: Optional[str] = None   # DATE — coerced to None if empty
    role: Optional[str] = None
    team: str
    manager: str
    password: Optional[str] = None   # hashed on server

class LoginRequest(BaseModel):
    email: str
    password: str

class ApprovalAction(BaseModel):
    action: str  # "approve" | "decline"
    manager_name: Optional[str] = None

class CapstoneAction(BaseModel):
    actioned_by: Optional[str] = None
    notes: Optional[str] = None  # optional comment left on approval

class TrackSelection(BaseModel):
    email: str
    track: str

class AllocationRow(BaseModel):
    member_name: str
    manager: str
    project_id: Optional[str] = ""
    project_name: Optional[str] = ""
    project_type: Optional[str] = ""
    industry: Optional[str] = ""
    phase: Optional[str] = ""
    stage: Optional[str] = ""
    start_date: Optional[str] = None
    end_date: Optional[str] = None
    hrs_per_week: Optional[float] = 0
    use_cases: Optional[str] = ""
    solutions_used: Optional[str] = ""
    product_features: Optional[str] = ""
    data_sources: Optional[str] = ""
    destinations: Optional[str] = ""
    num_audiences: Optional[int] = 0
    region: Optional[str] = ""
    ticket_ids: Optional[str] = ""
    health_status: Optional[str] = "On track"
    renewal: Optional[str] = "TBD"
    comments: Optional[str] = ""
    project_notes: Optional[str] = ""


@app.post("/api/onboarding")
def submit_onboarding(req: OnboardingRequest):
    """New joiner submits onboarding request — saved as pending, password hashed with SHA-256."""
    import hashlib
    pwd_hash = hashlib.sha256(req.password.encode()).hexdigest() if req.password else None
    email_clean = req.email.strip().lower()
    # Coerce empty date string → None (PostgreSQL DATE rejects empty strings)
    joining_date = req.joining_date if req.joining_date and req.joining_date.strip() else None
    try:
        with get_db() as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                # Ensure columns exist
                cur.execute("ALTER TABLE onboarding_requests ADD COLUMN IF NOT EXISTS password_hash VARCHAR(64)")
                cur.execute("ALTER TABLE onboarding_requests ADD COLUMN IF NOT EXISTS capstone_completed BOOLEAN DEFAULT FALSE")
                cur.execute("ALTER TABLE onboarding_requests ADD COLUMN IF NOT EXISTS capstone_completed_at TIMESTAMP")
                cur.execute("ALTER TABLE onboarding_requests ADD COLUMN IF NOT EXISTS active_track VARCHAR(50) DEFAULT 'rtcdp'")
                cur.execute("ALTER TABLE onboarding_requests ADD COLUMN IF NOT EXISTS role VARCHAR(100) DEFAULT ''")
                cur.execute("ALTER TABLE onboarding_requests ADD COLUMN IF NOT EXISTS profile_confirmed BOOLEAN DEFAULT FALSE")
                # Project tracker enrichment columns
                cur.execute("ALTER TABLE projects ADD COLUMN IF NOT EXISTS project_code TEXT DEFAULT ''")
                cur.execute("ALTER TABLE projects ADD COLUMN IF NOT EXISTS project_type TEXT DEFAULT ''")
                cur.execute("ALTER TABLE projects ADD COLUMN IF NOT EXISTS industry TEXT DEFAULT ''")
                cur.execute("ALTER TABLE projects ADD COLUMN IF NOT EXISTS phase TEXT DEFAULT ''")
                cur.execute("ALTER TABLE projects ADD COLUMN IF NOT EXISTS stage TEXT DEFAULT ''")
                cur.execute("ALTER TABLE projects ADD COLUMN IF NOT EXISTS start_date DATE")
                cur.execute("ALTER TABLE projects ADD COLUMN IF NOT EXISTS end_date DATE")
                cur.execute("ALTER TABLE projects ADD COLUMN IF NOT EXISTS solutions_used TEXT DEFAULT ''")
                cur.execute("ALTER TABLE projects ADD COLUMN IF NOT EXISTS health_status TEXT DEFAULT ''")
                cur.execute("ALTER TABLE projects ADD COLUMN IF NOT EXISTS weekly_comments TEXT DEFAULT ''")
                cur.execute("ALTER TABLE projects ADD COLUMN IF NOT EXISTS high_level_notes TEXT DEFAULT ''")
                cur.execute("ALTER TABLE projects ADD COLUMN IF NOT EXISTS renewal TEXT DEFAULT ''")
                cur.execute("ALTER TABLE projects ADD COLUMN IF NOT EXISTS region TEXT DEFAULT ''")
                cur.execute("ALTER TABLE projects ADD COLUMN IF NOT EXISTS use_cases TEXT DEFAULT ''")
                cur.execute("ALTER TABLE projects ADD COLUMN IF NOT EXISTS product_features TEXT DEFAULT ''")
                cur.execute("ALTER TABLE projects ADD COLUMN IF NOT EXISTS data_sources TEXT DEFAULT ''")
                cur.execute("ALTER TABLE projects ADD COLUMN IF NOT EXISTS destinations TEXT DEFAULT ''")
                cur.execute("ALTER TABLE projects ADD COLUMN IF NOT EXISTS num_audiences INTEGER DEFAULT 0")
                cur.execute("ALTER TABLE projects ADD COLUMN IF NOT EXISTS ticket_ids TEXT DEFAULT ''")
                cur.execute("ALTER TABLE projects ADD COLUMN IF NOT EXISTS days_remaining INTEGER")
                cur.execute("ALTER TABLE projects ADD COLUMN IF NOT EXISTS imported_from_tracker BOOLEAN DEFAULT FALSE")
                # hrs_per_week on project_members
                cur.execute("ALTER TABLE project_members ADD COLUMN IF NOT EXISTS hrs_per_week NUMERIC(5,1) DEFAULT 0")
                cur.execute("ALTER TABLE project_members ADD COLUMN IF NOT EXISTS role_on_project TEXT DEFAULT ''")
                cur.execute(
                    """INSERT INTO onboarding_requests
                       (name, preferred_name, email, joining_date, role, team, manager, status, password_hash)
                       VALUES (%s,%s,%s,%s,%s,%s,%s,'pending',%s)
                       ON CONFLICT (email) DO UPDATE
                       SET name=%s, preferred_name=%s, joining_date=%s, role=%s,
                           team=%s, manager=%s, status='pending',
                           password_hash=COALESCE(%s, onboarding_requests.password_hash),
                           created_at=NOW()
                       RETURNING id""",
                    (req.name, req.preferred_name, email_clean, joining_date, req.role or '',
                     req.team, req.manager, pwd_hash,
                     req.name, req.preferred_name, joining_date, req.role or '',
                     req.team, req.manager, pwd_hash)
                )
                row = cur.fetchone()
        return {"ok": True, "id": row["id"]}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/auth/login")
def login(req: LoginRequest):
    """Authenticate an approved user with email + password."""
    import hashlib
    pwd_hash = hashlib.sha256(req.password.encode()).hexdigest()
    # Ensure columns exist — separate connection so DDL commits before SELECT
    try:
        _conn = psycopg2.connect(DATABASE_URL)
        _conn.autocommit = True
        with _conn.cursor() as _c:
            _c.execute("ALTER TABLE onboarding_requests ADD COLUMN IF NOT EXISTS capstone_completed BOOLEAN DEFAULT FALSE")
            _c.execute("ALTER TABLE onboarding_requests ADD COLUMN IF NOT EXISTS capstone_completed_at TIMESTAMP")
            _c.execute("ALTER TABLE onboarding_requests ADD COLUMN IF NOT EXISTS active_track VARCHAR(50) DEFAULT 'rtcdp'")
            _c.execute("ALTER TABLE onboarding_requests ADD COLUMN IF NOT EXISTS capstone_started_at TIMESTAMP")
            _c.execute("ALTER TABLE onboarding_requests ADD COLUMN IF NOT EXISTS username VARCHAR(60)")
            _c.execute("ALTER TABLE onboarding_requests ADD COLUMN IF NOT EXISTS avatar_emoji VARCHAR(10)")
            _c.execute("ALTER TABLE onboarding_requests ADD COLUMN IF NOT EXISTS avatar_color VARCHAR(10)")
        _conn.close()
    except Exception:
        pass

    try:
        with get_db() as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute(
                    """SELECT id, name, preferred_name, email, team, manager,
                              joining_date, status, password_hash, capstone_started_at,
                              username, avatar_emoji, avatar_color,
                              COALESCE(capstone_completed, FALSE) as capstone_completed,
                              COALESCE(active_track, 'rtcdp') as active_track
                       FROM onboarding_requests
                       WHERE LOWER(email)=LOWER(%s)
                       ORDER BY created_at DESC LIMIT 1""",
                    (req.email.strip(),)
                )
                row = cur.fetchone()
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

    if not row:
        raise HTTPException(status_code=404, detail="No account found for this email. Please complete registration first.")

    if row["status"] == "pending":
        return {"ok": False, "status": "pending",
                "message": "Your account is awaiting manager approval. Check back soon.",
                "name": row["name"]}

    if row["status"] == "declined":
        return {"ok": False, "status": "declined",
                "message": "Your onboarding request was not approved. Please contact your manager."}

    if row["status"] != "approved":
        return {"ok": False, "status": row["status"], "message": "Account not yet approved."}

    # Approved — verify password
    if not row["password_hash"]:
        # No password set — allow login without password check (legacy / admin-created accounts)
        pass
    elif row["password_hash"] != pwd_hash:
        raise HTTPException(status_code=401, detail="Incorrect password.")

    # Determine persona from the shared DOJ + capstone rule (see ims_auth.classify_persona).
    capstone_done = bool(row.get("capstone_completed", False))
    active_track = row.get("active_track", "rtcdp") or "rtcdp"
    capstone_started_at = row.get("capstone_started_at")
    persona, tenure = classify_persona(row.get("joining_date"), capstone_done)

    return {
        "ok": True,
        "status": "approved",
        "profile": {
            "id": row["id"],
            "name": row["name"],
            "preferred_name": row["preferred_name"],
            "email": row["email"],
            "team": row["team"],
            "manager": row["manager"],
            "joining_date": str(row["joining_date"]),
            "tenure": tenure,
            "persona": persona,
            "capstone_completed": capstone_done,
            "capstone_started_at": str(capstone_started_at) if capstone_started_at else None,
            "active_track": active_track,
            "username": row.get("username"),
            "avatar_emoji": row.get("avatar_emoji"),
            "avatar_color": row.get("avatar_color"),
        }
    }

@app.get("/api/onboarding/pending")
def get_pending_approvals(manager: Optional[str] = None):
    """Manager fetches all pending requests (optionally filtered by manager name)."""
    try:
        with get_db() as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                if manager:
                    cur.execute(
                        "SELECT * FROM onboarding_requests WHERE status='pending' AND manager=%s ORDER BY created_at DESC",
                        (manager,)
                    )
                else:
                    cur.execute(
                        "SELECT * FROM onboarding_requests WHERE status='pending' ORDER BY created_at DESC"
                    )
                rows = cur.fetchall()
        return {"requests": [dict(r) for r in rows]}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.put("/api/onboarding/{request_id}")
def action_onboarding(request_id: int, body: ApprovalAction):
    """Manager approves or declines a request. Sends an email to the learner if SMTP is configured."""
    if body.action not in ("approve", "decline"):
        raise HTTPException(status_code=400, detail="action must be approve or decline")
    new_status = body.action + "d"
    try:
        with get_db() as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute(
                    "UPDATE onboarding_requests SET status=%s, actioned_by=%s, actioned_at=NOW() WHERE id=%s RETURNING name, email, team",
                    (new_status, body.manager_name, request_id)
                )
                row = cur.fetchone()
        email_result = {"ok": False, "error": "not attempted"}
        if row:
            if new_status == "approved":
                if row.get("email"):
                    html = email_template(
                        "You're approved!",
                        f"Hi {row['name']}, your manager has approved your Nexus access for the <b>{row['team']}</b> team. "
                        f"Sign in with your registered email and password to start your enablement track.",
                        "Sign in to Nexus", "http://localhost:5173"
                    )
                    email_result = send_email(row["email"], "Nexus access approved", html)
                create_notification(row["name"], body.manager_name or row.get("team") or "", "approval",
                    "You're approved!", f"Your manager approved your access for the {row['team']} team. Sign in to start your enablement track.")
                award_points(row["name"], body.manager_name or row.get("team") or "", 20, "Welcome bonus — onboarding approved")
            else:
                if row.get("email"):
                    html = email_template(
                        "Update on your request",
                        f"Hi {row['name']}, your Nexus onboarding request was not approved at this time. Please reach out to your manager for details."
                    )
                    email_result = send_email(row["email"], "Nexus onboarding update", html)
                create_notification(row["name"], body.manager_name or row.get("team") or "", "decline",
                    "Update on your request", "Your onboarding request was not approved at this time. Reach out to your manager for details.")
        return {"ok": True, "status": new_status, "email_sent": email_result.get("ok", False)}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/onboarding/status/{email}")
def get_onboarding_status(email: str):
    """Learner polls to check if their request was approved."""
    try:
        with get_db() as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute(
                    "SELECT status, team, manager, actioned_at FROM onboarding_requests WHERE LOWER(email)=LOWER(%s) ORDER BY created_at DESC LIMIT 1",
                    (email.strip(),)
                )
                row = cur.fetchone()
        if not row:
            return {"status": "not_found"}
        return dict(row)
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# ── Curriculum endpoints ───────────────────────────────────────────────────────


# ── Curriculum endpoints ─────────────────────────────────────────────────────

@app.get("/api/curriculum/{module_id}")
def get_curriculum(module_id: int, track: str = "rtcdp"):
    """Fetch all topics for a given module and track."""
    try:
        with get_db() as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute(
                    "SELECT * FROM curriculum_topics WHERE module_id=%s AND track=%s ORDER BY topic_order",
                    (module_id, track)
                )
                rows = cur.fetchall()
        return {"module_id": module_id, "track": track, "topics": [dict(r) for r in rows]}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/curriculum")
def get_all_curriculum(track: str = "rtcdp"):
    """Fetch all topics grouped by module for a given track."""
    try:
        with get_db() as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute(
                    "SELECT * FROM curriculum_topics WHERE track=%s ORDER BY module_id, topic_order",
                    (track,)
                )
                rows = cur.fetchall()
        result = {}
        for r in rows:
            mid = r["module_id"]
            result.setdefault(mid, []).append(dict(r))
        return {"modules": result, "total": len(rows), "track": track}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# ── EL Content fetch — returns actual AdobeDocs markdown for a topic ──────────

def el_url_to_github_path(el_url: str):
    """Map an EL URL to the corresponding AdobeDocs GitHub repo + file path."""
    import re
    url = el_url.replace("https://experienceleague.adobe.com","").split("?")[0].split("#")[0].rstrip("/")

    # platform-learn tutorials
    m = re.search(r'/docs/platform-learn/tutorials/(.+?)(?:\.html)?$', url)
    if m:
        return "AdobeDocs/platform-learn.en", f"help/platform/{m.group(1)}.md"

    # analytics-platform (CJA)
    m = re.search(r'/(?:en/)?docs/analytics-platform/(.+?)(?:\.html)?$', url)
    if m:
        return "AdobeDocs/analytics-platform.en", f"help/{m.group(1)}.md"

    # adobe analytics (legacy + new)
    m = re.search(r'/(?:en/)?docs/analytics/(.+?)(?:\.html)?$', url)
    if m:
        return "AdobeDocs/analytics.en", f"help/{m.group(1)}.md"

    # analytics learn tutorials
    m = re.search(r'/docs/analytics-learn/(.+?)(?:\.html)?$', url)
    if m:
        return "AdobeDocs/analytics-learn.en", f"help/{m.group(1)}.md"

    # experience-platform docs
    m = re.search(r'/(?:en/)?docs/experience-platform/(.+?)(?:\.html)?$', url)
    if m:
        return "AdobeDocs/experience-platform.en", f"help/{m.group(1)}.md"

    # federated audience composition
    m = re.search(r'/(?:en/)?docs/federated-audience-composition/(.+?)(?:\.html)?$', url)
    if m:
        return "AdobeDocs/federated-audience-composition.en", f"help/{m.group(1)}.md"

    # Adobe Target
    m = re.search(r'/(?:en/)?docs/target/(.+?)(?:\.html)?$', url)
    if m:
        return "AdobeDocs/target.en", f"help/{m.group(1)}.md"

    # Marketo Engage — repo's docs live under help/marketo/, not help/ directly
    m = re.search(r'/(?:en/)?docs/marketo/(.+?)(?:\.html)?$', url)
    if m:
        return "AdobeDocs/marketo.en", f"help/marketo/{m.group(1)}.md"

    # Adobe Campaign Classic
    m = re.search(r'/(?:en/)?docs/campaign-classic/(.+?)(?:\.html)?$', url)
    if m:
        return "AdobeDocs/campaign-classic.en", f"help/{m.group(1)}.md"

    # Adobe Journey Optimizer B2B Edition — a SEPARATE repo from the standard
    # edition (journey-optimizer-b2b.en, not journey-optimizer.en), with its own
    # URL segment. Must be checked before the plain journey-optimizer case below.
    m = re.search(r'/(?:en/)?docs/journey-optimizer-b2b/(.+?)(?:\.html)?$', url)
    if m:
        return "AdobeDocs/journey-optimizer-b2b.en", f"help/{m.group(1)}.md"

    # Adobe Journey Optimizer (standard edition)
    m = re.search(r'/(?:en/)?docs/journey-optimizer/(.+?)(?:\.html)?$', url)
    if m:
        return "AdobeDocs/journey-optimizer.en", f"help/{m.group(1)}.md"

    return "AdobeDocs/experience-platform.en", None

def _cache_topic_content(module_id: int, topic_order: int, track: str, result: dict) -> None:
    """Best-effort: persist a successful live fetch so it can be served if a later
    fetch for the same topic fails. Never raises — a caching failure should not
    break the response that's already succeeding."""
    try:
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    """UPDATE curriculum_topics
                       SET content=%s, cached_video_url=%s, cached_fetch_method=%s, cached_at=NOW()
                       WHERE module_id=%s AND topic_order=%s AND track=%s""",
                    (result.get("content"), result.get("video_url"), result.get("fetch_method"),
                     module_id, topic_order, track))
    except Exception as e:
        print(f"[get_topic_content] cache write skipped: {e}")


def _get_cached_topic_content(module_id: int, topic_order: int, track: str) -> dict | None:
    """Last-resort fallback when every live fetch attempt fails this request."""
    try:
        with get_db() as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute(
                    """SELECT title, el_url, content, cached_video_url, cached_fetch_method, cached_at
                       FROM curriculum_topics
                       WHERE module_id=%s AND topic_order=%s AND track=%s AND content IS NOT NULL""",
                    (module_id, topic_order, track))
                row = cur.fetchone()
    except Exception as e:
        print(f"[get_topic_content] cache read skipped: {e}")
        return None
    if not row:
        return None
    return {
        "content": row["content"], "format": "markdown", "source": "cache",
        "fetch_method": f"Cached copy from a previous successful fetch ({row['cached_fetch_method']})",
        "cached_at": str(row["cached_at"]) if row["cached_at"] else None,
        "el_url": row.get("el_url") or "", "title": row["title"],
        "video_url": row.get("cached_video_url"),
    }


@app.get("/api/content/{module_id}/{topic_order}")
async def get_topic_content(module_id: int, topic_order: int, track: str = "rtcdp"):
    """Fetch content for a topic — GitHub AdobeDocs with EL URL as hint, falls back to search.

    Live fetch is always tried first and is never skipped in favor of the cache.
    Only when every live attempt below fails does this fall back to the last
    successful fetch cached on curriculum_topics (see _cache_topic_content) —
    the DB is a reliability fallback, not a replacement for the live fetch."""
    import re
    try:
        with get_db() as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute(
                    "SELECT title, objective, activity, output, checkpoint, el_url FROM curriculum_topics WHERE module_id=%s AND topic_order=%s AND track=%s",
                    (module_id, topic_order, track)
                )
                topic = cur.fetchone()
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

    if not topic:
        return {"content": None, "source": "not_found"}

    title = topic["title"]
    el_url = topic.get("el_url") or ""

    gh_headers = {"Accept": "application/vnd.github.v3+json", "User-Agent": "Nexus-Platform/1.0"}
    if GITHUB_TOKEN:
        gh_headers["Authorization"] = f"token {GITHUB_TOKEN}"

    # Track -> single-repo products (no ambiguity, no keyword heuristics needed)
    SINGLE_REPO_TRACKS = {
        "target": "AdobeDocs/target.en",
        "marketo": "AdobeDocs/marketo.en",
        "campaign": "AdobeDocs/campaign-classic.en",
        "ajo": "AdobeDocs/journey-optimizer.en",
        "cja": "AdobeDocs/analytics-platform.en",
    }

    # Choose repo based on track, EL URL, or title
    if track in SINGLE_REPO_TRACKS:
        repo = SINGLE_REPO_TRACKS[track]
        # B2B-specific AJO content lives in the separate B2B-edition repo.
        if track == "ajo" and "b2b" in title.lower():
            repo = "AdobeDocs/journey-optimizer-b2b.en"
    elif track == "analytics":
        if "analytics-platform" in el_url or any(k in title.lower() for k in ["cja","customer journey analytics","connection","data view","stitching"]):
            repo = "AdobeDocs/analytics-platform.en"
        elif "analytics-learn" in el_url:
            repo = "AdobeDocs/analytics-learn.en"
        else:
            repo = "AdobeDocs/analytics.en"
    else:
        repo = "AdobeDocs/experience-platform.en"
        if "platform-learn" in el_url or any(k in title.lower() for k in ["ingest","source","monitor","destination","activate","profile","segment","audience","union","merge","identity"]):
            repo = "AdobeDocs/platform-learn.en"
        if any(k in title.lower() for k in ["governance","sandbox","query service","rtcdp","computed","look-alike","playbook","partner data","customer ai","fac","fae","guardrail","ttl"]):
            repo = "AdobeDocs/experience-platform.en"
        if "federated" in title.lower() or "federated-audience" in el_url:
            repo = "AdobeDocs/federated-audience-composition.en"

    # The URL itself already encodes which product/repo it belongs to — that's
    # exactly what el_url_to_github_path derives. It's a more reliable signal
    # than the title-keyword heuristic above (which was guessing "platform-learn"
    # for anything with "destination" in the title, including AEP's own
    # Destination SDK docs, which actually live in experience-platform.en).
    # Try the URL-derived repo FIRST; the heuristic `repo` is now only a fallback.
    url_repo, file_path = el_url_to_github_path(el_url) if el_url else (None, None)
    direct_repos = [r for r in ([url_repo, repo] if url_repo else [repo]) if r]
    # de-dupe while preserving order
    seen = set(); direct_repos = [r for r in direct_repos if not (r in seen or seen.add(r))]

    async with httpx.AsyncClient(timeout=20.0, follow_redirects=True) as client:
        debug = []
        # ── Try direct path from EL URL ──────────────────────────────────────
        if el_url and file_path:
            for try_repo in direct_repos:
                for branch in ["main", "master"]:
                    raw_url = f"https://raw.githubusercontent.com/{try_repo}/{branch}/{file_path}"
                    try:
                        r = await client.get(raw_url, headers=gh_headers)
                        debug.append(f"direct:{try_repo}:{branch}={r.status_code}:len={len(r.text)}")
                        if r.status_code == 200 and is_relevant_content(r.text, title):
                            video_url = extract_video_url(r.text)
                            result = {"content": r.text, "format": "markdown", "source": "github",
                                      "fetch_method": "Direct GitHub path from Experience League URL",
                                      "repo": try_repo, "file_path": file_path, "github_url": raw_url,
                                      "el_url": el_url, "title": title, "video_url": video_url}
                            _cache_topic_content(module_id, topic_order, track, result)
                            return result
                    except Exception as e:
                        debug.append(f"direct:{try_repo}:{branch}=err:{str(e)[:40]}")

        # ── Search GitHub by filename from EL URL ────────────────────────────
        if el_url:
            url_parts = el_url.rstrip('/').split('/')
            fname = url_parts[-1].replace('.html', '.md')
            if track == "analytics":
                extra_repos = ["AdobeDocs/analytics.en", "AdobeDocs/analytics-platform.en"]
            elif track in SINGLE_REPO_TRACKS:
                extra_repos = [SINGLE_REPO_TRACKS[track]]
            else:
                extra_repos = ["AdobeDocs/platform-learn.en", "AdobeDocs/experience-platform.en"]
            search_order = direct_repos + [repo] + extra_repos
            seen2 = set(); search_order = [r for r in search_order if not (r in seen2 or seen2.add(r))]
            for search_repo in search_order:
                try:
                    q = f"filename:{fname} repo:{search_repo}"
                    s = await client.get("https://api.github.com/search/code", params={"q": q, "per_page": 3}, headers=gh_headers)
                    sdata = s.json()
                    debug.append(f"search:{search_repo}={s.status_code}:hits={sdata.get('total_count',0)}")
                    items = sdata.get("items", [])
                    for item in items[:3]:
                        r = await client.get(f"https://raw.githubusercontent.com/{search_repo}/main/{item['path']}", headers=gh_headers)
                        debug.append(f"fetch:{item['path'][:40]}={r.status_code}")
                        if r.status_code == 200 and is_relevant_content(r.text, title):
                            video_url = extract_video_url(r.text)
                            result = {"content": r.text, "format": "markdown", "source": "github_search",
                                      "fetch_method": "GitHub code search by filename from Experience League URL",
                                      "repo": search_repo, "file_path": item['path'],
                                      "github_url": f"https://github.com/{search_repo}/blob/main/{item['path']}",
                                      "el_url": el_url, "title": title, "video_url": video_url}
                            _cache_topic_content(module_id, topic_order, track, result)
                            return result
                except Exception as e:
                    debug.append(f"search_err:{str(e)[:40]}")

        # ── Search by topic title keywords ───────────────────────────────────
        clean = re.sub(r'[^a-zA-Z0-9 ]', '', title.lower())
        keywords = '-'.join(clean.split()[:4])
        if track == "analytics":
            kw_repos = ["AdobeDocs/analytics.en", "AdobeDocs/analytics-platform.en", "AdobeDocs/analytics-learn.en"]
        elif track in SINGLE_REPO_TRACKS:
            kw_repos = [SINGLE_REPO_TRACKS[track]]
        else:
            kw_repos = ["AdobeDocs/platform-learn.en", "AdobeDocs/experience-platform.en"]
        kw_order = direct_repos + kw_repos
        seen3 = set(); kw_order = [r for r in kw_order if not (r in seen3 or seen3.add(r))]
        for search_repo in kw_order:
            try:
                q = f"filename:{keywords}.md repo:{search_repo}"
                s = await client.get("https://api.github.com/search/code", params={"q": q, "per_page": 3}, headers=gh_headers)
                sdata = s.json()
                debug.append(f"kw:{search_repo}={s.status_code}:hits={sdata.get('total_count',0)}")
                items = sdata.get("items", [])
                for item in items[:3]:
                    r = await client.get(f"https://raw.githubusercontent.com/{search_repo}/main/{item['path']}", headers=gh_headers)
                    if r.status_code == 200 and is_relevant_content(r.text, title):
                        video_url = extract_video_url(r.text)
                        result = {"content": r.text, "format": "markdown", "source": "github_search",
                                  "fetch_method": "GitHub code search by topic-title keywords",
                                  "repo": search_repo, "file_path": item['path'],
                                  "github_url": f"https://github.com/{search_repo}/blob/main/{item['path']}",
                                  "el_url": el_url, "title": title, "video_url": video_url}
                        _cache_topic_content(module_id, topic_order, track, result)
                        return result
            except Exception as e:
                debug.append(f"kw_err:{str(e)[:40]}")

    # Every live attempt above failed — fall back to the last successful fetch,
    # if one was ever cached for this exact topic/track. Still "not_found" (no
    # live attempt succeeded) is the correct outcome only when nothing is cached.
    cached = _get_cached_topic_content(module_id, topic_order, track)
    if cached:
        return cached

    return {"content": None, "source": "not_found", "el_url": el_url, "title": title,
            "repo": repo, "fetch_method": "No matching document found on GitHub AdobeDocs", "debug": debug}

def extract_video_url(text: str) -> str | None:
    """Extract video.tv.adobe.com URL from AdobeDocs markdown."""
    import re
    m = re.search(r'https://video\.tv\.adobe\.com/v/[^\s\)\"\'>]+', text)
    if not m:
        return None
    raw = m.group(0).rstrip(')')
    # Extract just the video ID
    vid_id = re.search(r'/v/([^\s\?/\"\']+)', raw)
    if not vid_id:
        return None
    return f"https://video.tv.adobe.com/v/{vid_id.group(1)}?quality=12&learn=on&hidetitle=true"

def is_relevant_content(text: str, title: str, min_len: int = 50) -> bool:
    """Check fetched content is about the topic and not a stub."""
    import re
    body = re.sub(r'^---[\s\S]*?---\n?', '', text).strip()
    if len(body) < min_len and 'video.tv.adobe.com' not in text:
        return False
    fm_title = re.search(r'^title:\s*(.+)$', text, re.MULTILINE)
    h1_title = re.search(r'^#\s+(.+)$', text, re.MULTILINE)
    content_title = (fm_title or h1_title)
    if not content_title:
        return True
    ct = content_title.group(1).lower()
    topic_words = set(w for w in re.split(r'\W+', title.lower()) if len(w) > 3)
    ct_words = set(w for w in re.split(r'\W+', ct) if len(w) > 3)
    # Prefix-tolerant match so plurals / minor variants count as the same concept
    # (e.g. "Sandboxes"↔"sandbox", "Datasets"↔"dataset"). Guarded at len>=5 to
    # avoid loose matches on short common words.
    def _rel(a, b):
        if a == b:
            return True
        return len(a) >= 5 and len(b) >= 5 and (a.startswith(b) or b.startswith(a))
    return any(_rel(tw, cw) for tw in topic_words for cw in ct_words)
    """Check if fetched markdown is real content, not just a metadata stub."""
    import re
    body = re.sub(r'^---[\s\S]*?---\n?', '', text).strip()
    # Accept if body is long enough, OR if it contains a video embed (tutorial files are short by design)
    has_video = bool(re.search(r'video\.tv\.adobe\.com', text))
    return len(body) >= min_len or has_video



@app.put("/api/onboarding/{request_id}/capstone")
def mark_capstone(request_id: int, body: CapstoneAction):
    """Manager marks a learner's capstone as complete."""
    try:
        with get_db() as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute("ALTER TABLE onboarding_requests ADD COLUMN IF NOT EXISTS capstone_completed BOOLEAN DEFAULT FALSE")
                cur.execute("ALTER TABLE onboarding_requests ADD COLUMN IF NOT EXISTS capstone_completed_at TIMESTAMP")
                cur.execute(
                    "UPDATE onboarding_requests SET capstone_completed=TRUE, capstone_completed_at=NOW(), actioned_by=%s WHERE id=%s RETURNING id, name, email, manager",
                    (body.actioned_by, request_id)
                )
                row = cur.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Request not found")
        manager_name = row.get("manager") or body.actioned_by or ""
        award_points(row["name"], manager_name, 200, "Capstone completed")
        create_notification(row["name"], manager_name, "capstone_complete",
            "Capstone complete! 🎉",
            "Your manager marked your capstone complete. Your Skills Dashboard is now unlocked — pick a cross-skill track or go deeper in your current one.")
        # Best-effort — reflect the manager's decision on the latest capstone submission
        # (if one exists) without changing this endpoint's own semantics or return shape.
        # manager_notes itself is untouched here (approval never sets/clears it) — the
        # optional approval comment only goes into the permanent history array below.
        try:
            from datetime import datetime, timezone
            history_entry = json.dumps([{
                "decision": "approved",
                "notes": (body.notes or "").strip() or None,
                "reviewed_at": datetime.now(timezone.utc).isoformat(),
                "manager_name": body.actioned_by or manager_name,
            }])
            with get_db() as _conn:
                with _conn.cursor() as _cur:
                    _cur.execute(
                        """UPDATE capstone_submissions
                           SET status='manager_approved', reviewed_at=NOW(),
                               manager_review_history = COALESCE(manager_review_history,'[]'::jsonb) || %s::jsonb
                           WHERE id=(SELECT id FROM capstone_submissions WHERE member_id=%s ORDER BY generated_at DESC LIMIT 1)""",
                        (history_entry, request_id)
                    )
        except Exception:
            pass
        return {"ok": True, "id": row["id"], "name": row["name"], "email": row["email"]}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.put("/api/onboarding/{request_id}/track")
def set_active_track(request_id: int, body: TrackSelection):
    """Learner selects their cross-skill or upskill track."""
    try:
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("ALTER TABLE onboarding_requests ADD COLUMN IF NOT EXISTS active_track VARCHAR(50) DEFAULT 'rtcdp'")
                cur.execute(
                    "UPDATE onboarding_requests SET active_track=%s WHERE id=%s",
                    (body.track, request_id)
                )
        return {"ok": True, "track": body.track}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/onboarding/team")
def get_team_members(manager: str):
    """Manager fetches all approved team members with capstone status."""
    try:
        with get_db() as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute("ALTER TABLE onboarding_requests ADD COLUMN IF NOT EXISTS capstone_started_at TIMESTAMP")
                cur.execute(
                    """SELECT id, name, preferred_name, email, team, joining_date,
                              COALESCE(capstone_completed, FALSE) as capstone_completed,
                              capstone_completed_at, capstone_started_at,
                              COALESCE(active_track, 'rtcdp') as active_track
                       FROM onboarding_requests
                       WHERE status='approved' AND manager=%s
                       ORDER BY joining_date""",
                    (manager,)
                )
                rows = cur.fetchall()
        for r in rows:
            if r.get("capstone_started_at"): r["capstone_started_at"]=str(r["capstone_started_at"])
            if r.get("capstone_completed_at"): r["capstone_completed_at"]=str(r["capstone_completed_at"])
            if r.get("joining_date"): r["joining_date"]=str(r["joining_date"])
        return {"members": [dict(r) for r in rows]}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ── Capstone submissions ───────────────────────────────────────────────────────
# GENERATE/EVALUATE themselves are NOT dedicated backend endpoints — the frontend
# calls the existing generic /api/agent proxy with a mode-prefixed system prompt
# (agent_name="Capstone"), the same pattern already scaffolded for the Evaluation
# Agent. Two reasons this beats routing through the old dedicated LangGraph
# /api/agents/capstone endpoint: (1) _groq_call() inside LangGraph nodes never
# writes to llm_logs, so the Admin per-agent token view would silently show zero
# calls for Capstone; (2) that graph has no EVALUATE mode and returns a task-list
# shape, not the {title, client_context, skills_tested, deliverable,
# evaluation_criteria} / {pass, score, feedback, strengths, gaps, recommendation}
# contract. These endpoints only persist the scenario/response/evaluation so a
# manager can review them alongside the existing "mark complete" action.

class CapstoneGenerateBody(BaseModel):
    member_id: int
    scenario: dict

class CapstoneSubmitBody(BaseModel):
    response_text: str

class CapstoneEvaluateBody(BaseModel):
    ai_evaluation: dict

@app.post("/api/capstone/generate")
def capstone_generate(body: CapstoneGenerateBody):
    """Persist a freshly-generated capstone scenario for a learner, stamping a
    7-day deadline (due_at) from now. Capstones are human-graded, time-boxed."""
    try:
        with get_db() as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute(
                    """INSERT INTO capstone_submissions (member_id, scenario, status, due_at)
                       VALUES (%s,%s,'generated', NOW() + INTERVAL '7 days')
                       RETURNING id, due_at""",
                    (body.member_id, json.dumps(body.scenario))
                )
                row = cur.fetchone()
        return {"ok": True, "id": row["id"], "due_at": str(row["due_at"])}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


class CapstoneAssistantBody(BaseModel):
    messages: List[dict] = []
    track: str = "rtcdp"
    profile: dict = {}
    extra: dict = {}


@app.post("/api/agents/capstone-assistant")
def capstone_assistant(body: CapstoneAssistantBody):
    """Capstone AI assistant — guides a learner through THEIR capstone. It is a
    coach: it clarifies tasks, points to the right Adobe capabilities/docs, and
    asks guiding questions, but never writes the submission for them. Grounded in
    the learner's own generated capstone scenario (passed in extra.scenario)."""
    scenario = body.extra.get("scenario") or {}
    tasks = scenario.get("tasks", [])
    task_lines = "\n".join(
        f"- [{t.get('level','')}] {t.get('title','')}: {t.get('description','')}" for t in tasks
    )
    deck = ", ".join(s.get("slide", "") for s in scenario.get("deck_requirements", []))
    sys = (
        "You are the Capstone Assistant, a supportive Adobe SME coaching a learner "
        "through their FINAL capstone. Guide with hints, clarifications, and the "
        "right Adobe products/features/docs to consult. Ask a guiding question when "
        "useful. NEVER write their deliverable or hand them a full answer — this is "
        "graded work. Keep replies under 130 words.\n\n"
        f"THEIR CAPSTONE: {scenario.get('title','(capstone)')}\n"
        f"Objective: {scenario.get('objective','')}\n"
        f"Company/context: {scenario.get('company_brief','')}\n"
        f"Tasks:\n{task_lines}\n"
        f"Required deck: {deck}\n\n"
        f"{PRODUCT_DISTINCTIONS}"
    )
    msgs = [{"role": m.get("role", "user"),
             "content": m.get("content") if isinstance(m.get("content"), str) else str(m.get("content"))}
            for m in body.messages if m.get("role") in ("user", "assistant")]
    try:
        text = _llm_call(msgs, sys, max_tokens=350)
    except Exception as e:
        raise HTTPException(status_code=503, detail=f"Capstone assistant unavailable: {e}")
    return {"response": text, "grounded": bool(tasks)}

@app.put("/api/capstone/{submission_id}/submit")
def capstone_submit(submission_id: int, body: CapstoneSubmitBody):
    """Learner submits their written response to the generated scenario."""
    try:
        with get_db() as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute(
                    """UPDATE capstone_submissions SET response_text=%s, status='submitted', submitted_at=NOW(), manager_notes=NULL
                       WHERE id=%s RETURNING id""",
                    (body.response_text, submission_id)
                )
                row = cur.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Submission not found")
        return {"ok": True}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.put("/api/capstone/{submission_id}/evaluate")
def capstone_evaluate(submission_id: int, body: CapstoneEvaluateBody):
    """Store the advisory AI evaluation of a learner's submitted response."""
    try:
        with get_db() as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute(
                    """UPDATE capstone_submissions SET ai_evaluation=%s, status='ai_evaluated', evaluated_at=NOW()
                       WHERE id=%s RETURNING id""",
                    (json.dumps(body.ai_evaluation), submission_id)
                )
                row = cur.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Submission not found")
        return {"ok": True}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

class CapstoneRejectBody(BaseModel):
    manager_notes: str
    actioned_by: Optional[str] = None

@app.put("/api/capstone/{submission_id}/reject")
def capstone_reject(submission_id: int, body: CapstoneRejectBody):
    """Manager sends a submission back for changes. A note is mandatory — this is
    the manager's only channel to explain what needs to change. Learner keeps the
    same scenario and can edit their response_text and resubmit."""
    notes = (body.manager_notes or "").strip()
    if not notes:
        raise HTTPException(status_code=400, detail="manager_notes is required to reject a submission")
    from datetime import datetime, timezone
    history_entry = json.dumps([{
        "decision": "rejected",
        "notes": notes,
        "reviewed_at": datetime.now(timezone.utc).isoformat(),
        "manager_name": body.actioned_by or "",
    }])
    try:
        with get_db() as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute(
                    """UPDATE capstone_submissions
                       SET status='manager_rejected', reviewed_at=NOW(), manager_notes=%s,
                           manager_review_history = COALESCE(manager_review_history,'[]'::jsonb) || %s::jsonb
                       WHERE id=%s RETURNING id, member_id""",
                    (notes, history_entry, submission_id)
                )
                row = cur.fetchone()
                if not row:
                    raise HTTPException(status_code=404, detail="Submission not found")
                cur.execute(
                    "SELECT name, manager FROM onboarding_requests WHERE id=%s",
                    (row["member_id"],)
                )
                member = cur.fetchone()
        if member:
            create_notification(member["name"], member.get("manager") or "", "capstone_rejected",
                "Your capstone needs changes",
                f"Your manager reviewed your capstone submission and asked for changes: \"{notes}\". Head back to the Capstone tab to revise and resubmit.")
        return {"ok": True}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/capstone/{member_id}")
def capstone_latest(member_id: int):
    """Fetch a learner's latest capstone submission — used by both the learner
    (to resume across reloads) and the manager (to review before marking complete)."""
    try:
        with get_db() as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute(
                    """SELECT id, member_id, scenario, response_text, ai_evaluation, status,
                              generated_at, submitted_at, evaluated_at, reviewed_at, manager_notes, due_at,
                              COALESCE(manager_review_history,'[]'::jsonb) AS manager_review_history
                       FROM capstone_submissions WHERE member_id=%s
                       ORDER BY generated_at DESC LIMIT 1""",
                    (member_id,)
                )
                row = cur.fetchone()
        if not row:
            return {"submission": None}
        for k in ("generated_at","submitted_at","evaluated_at","reviewed_at","due_at"):
            if row.get(k): row[k] = str(row[k])
        return {"submission": dict(row)}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ── Demo mode endpoint ────────────────────────────────────────────────────────

@app.get("/api/demo/status")
def demo_status():
    """Quick health check with demo mode info."""
    return {
        "demo_email": "demo@adobe.com",
        "demo_persona": "demo",
        "all_modules": "unlocked",
        "confidence": 1.0,
        "note": "Use demo@adobe.com on the login screen to access Demo Mode"
    }


class UtilizationEntry(BaseModel):
    member_name: str
    manager: str
    week_of: str                   # ISO date: Monday of the week (YYYY-MM-DD)
    billable_hours: float = 0
    non_billable_cf_hours: float = 0
    ramp_credit: float = 0
    working_hours: float = 40
    holiday_hours: float = 0
    loa_hours: float = 0
    cf_target: float = 75

class WeeklyUpdate(BaseModel):
    member_name: str
    comment: str
    health_status: Optional[str] = None
    billable_hours: float = 0
    week_of: Optional[str] = None   # ISO date YYYY-MM-DD

# ── Weekly Tracker — Project Allocations (individual self-entry) ──────────────
# No roster table. The set of "team members" is derived dynamically from whoever
# has actually logged a project allocation — nothing is pre-seeded or hardcoded.

DEFAULT_WEEKLY_CAPACITY_HRS = 40

def _ensure_tracker_tables():
    """Idempotent table creation, run in a dedicated autocommit connection."""
    try:
        conn = psycopg2.connect(DATABASE_URL)
        conn.autocommit = True
        with conn.cursor() as c:
            c.execute("""
                CREATE TABLE IF NOT EXISTS project_allocations (
                    id SERIAL PRIMARY KEY,
                    member_name VARCHAR(120) NOT NULL,
                    manager VARCHAR(120),
                    project_id VARCHAR(50),
                    project_name VARCHAR(255),
                    project_type VARCHAR(100),
                    industry VARCHAR(100),
                    phase VARCHAR(50),
                    stage VARCHAR(50),
                    start_date DATE,
                    end_date DATE,
                    hrs_per_week NUMERIC DEFAULT 0,
                    use_cases TEXT,
                    solutions_used VARCHAR(255),
                    product_features VARCHAR(255),
                    data_sources VARCHAR(255),
                    destinations VARCHAR(255),
                    num_audiences INTEGER DEFAULT 0,
                    region VARCHAR(50),
                    ticket_ids VARCHAR(255),
                    health_status VARCHAR(50) DEFAULT 'On track',
                    renewal VARCHAR(20) DEFAULT 'TBD',
                    comments TEXT,
                    project_notes TEXT,
                    created_at TIMESTAMP DEFAULT NOW(),
                    updated_at TIMESTAMP DEFAULT NOW()
                )
            """)
            c.execute("""
                CREATE TABLE IF NOT EXISTS allocation_updates (
                    id SERIAL PRIMARY KEY,
                    allocation_id INTEGER REFERENCES project_allocations(id) ON DELETE CASCADE,
                    member_name VARCHAR(120) NOT NULL,
                    comment TEXT NOT NULL,
                    health_status VARCHAR(50),
                    created_at TIMESTAMP DEFAULT NOW()
                )
            """)
        conn.close()
    except Exception:
        pass

@app.get("/api/allocations")
def get_allocations(manager: str, member: Optional[str] = None):
    """List allocations for a manager's team, optionally filtered to one member."""
    _ensure_tracker_tables()
    try:
        with get_db() as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                if member:
                    cur.execute(
                        "SELECT * FROM project_allocations WHERE manager=%s AND member_name=%s ORDER BY end_date",
                        (manager, member)
                    )
                else:
                    cur.execute(
                        "SELECT * FROM project_allocations WHERE manager=%s ORDER BY member_name, end_date",
                        (manager,)
                    )
                rows = cur.fetchall()
        for r in rows:
            if r.get("start_date"): r["start_date"] = str(r["start_date"])
            if r.get("end_date"): r["end_date"] = str(r["end_date"])
        return {"allocations": [dict(r) for r in rows]}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/allocations")
def add_allocation(body: AllocationRow):
    """A team member logs a new project they're working on."""
    _ensure_tracker_tables()
    # Postgres DATE columns reject an empty string — coerce blank dates to NULL
    # (a blank form field arrives as "", not None). Mirrors the onboarding fix.
    start_date = body.start_date if (body.start_date or "").strip() else None
    end_date   = body.end_date if (body.end_date or "").strip() else None
    try:
        with get_db() as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute(
                    """INSERT INTO project_allocations
                       (member_name, manager, project_id, project_name, project_type, industry,
                        phase, stage, start_date, end_date, hrs_per_week, use_cases, solutions_used,
                        product_features, data_sources, destinations, num_audiences, region,
                        ticket_ids, health_status, renewal, comments, project_notes)
                       VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                       RETURNING id""",
                    (body.member_name, body.manager, body.project_id, body.project_name, body.project_type,
                     body.industry, body.phase, body.stage, start_date, end_date, body.hrs_per_week,
                     body.use_cases, body.solutions_used, body.product_features, body.data_sources,
                     body.destinations, body.num_audiences, body.region, body.ticket_ids,
                     body.health_status, body.renewal, body.comments, body.project_notes)
                )
                row = cur.fetchone()
        return {"ok": True, "id": row["id"]}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.put("/api/allocations/{alloc_id}")
def update_allocation(alloc_id: int, body: AllocationRow):
    """Owner edits their own project row (or manager corrects it)."""
    start_date = body.start_date if (body.start_date or "").strip() else None
    end_date   = body.end_date if (body.end_date or "").strip() else None
    try:
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    """UPDATE project_allocations SET
                       member_name=%s, project_id=%s, project_name=%s, project_type=%s, industry=%s,
                       phase=%s, stage=%s, start_date=%s, end_date=%s, hrs_per_week=%s, use_cases=%s,
                       solutions_used=%s, product_features=%s, data_sources=%s, destinations=%s,
                       num_audiences=%s, region=%s, ticket_ids=%s, health_status=%s, renewal=%s,
                       comments=%s, project_notes=%s, updated_at=NOW()
                       WHERE id=%s""",
                    (body.member_name, body.project_id, body.project_name, body.project_type, body.industry,
                     body.phase, body.stage, start_date, end_date, body.hrs_per_week, body.use_cases,
                     body.solutions_used, body.product_features, body.data_sources, body.destinations,
                     body.num_audiences, body.region, body.ticket_ids, body.health_status, body.renewal,
                     body.comments, body.project_notes, alloc_id)
                )
        return {"ok": True}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/api/allocations/{alloc_id}")
def delete_allocation(alloc_id: int):
    try:
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("DELETE FROM project_allocations WHERE id=%s", (alloc_id,))
        return {"ok": True}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/allocations/{alloc_id}/updates")
def get_weekly_updates(alloc_id: int):
    """Full weekly update history for one project, newest first."""
    _ensure_tracker_tables()
    try:
        with get_db() as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute(
                    "SELECT * FROM allocation_updates WHERE allocation_id=%s ORDER BY created_at DESC",
                    (alloc_id,)
                )
                rows = cur.fetchall()
        return {"updates": [dict(r) for r in rows]}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/allocations/{alloc_id}/updates")
def add_weekly_update(alloc_id: int, body: WeeklyUpdate):
    """Member posts this week's update for a specific project. Also refreshes
    the live health_status on the parent allocation if provided."""
    _ensure_tracker_tables()
    try:
        with get_db() as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                from datetime import date
                week_of=body.week_of or str(date.today()-__import__("datetime").timedelta(days=date.today().weekday()))
                cur.execute(
                    "INSERT INTO allocation_updates (allocation_id, member_name, comment, health_status, billable_hours, week_of) VALUES (%s,%s,%s,%s,%s,%s) RETURNING id, created_at",
                    (alloc_id, body.member_name, body.comment, body.health_status, body.billable_hours, week_of)
                )
                row = cur.fetchone()
                if body.health_status:
                    cur.execute(
                        "UPDATE project_allocations SET health_status=%s, updated_at=NOW() WHERE id=%s",
                        (body.health_status, alloc_id)
                    )
                cur.execute("SELECT manager FROM project_allocations WHERE id=%s", (alloc_id,))
                alloc_row = cur.fetchone()
        if alloc_row and alloc_row.get("manager"):
            award_points(body.member_name, alloc_row["manager"], 10, "Posted a weekly update")
        return {"ok": True, "id": row["id"], "created_at": str(row["created_at"])}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# ── Billing summary — total hours billed from project updates ────────────────

# ── Multi-track enrollment ───────────────────────────────────────────────────

class TrackEnrollRequest(BaseModel):
    member: str
    track: str
    member_id: Optional[int] = None
    email: Optional[str] = None

def _resolve_member_where(cur, member_id=None, email=None, member_name=None):
    """Resolve an onboarding_requests row by the most reliable identifier
    available: id > email > name. Name matching is fragile — the session's
    display name can drift from the directory record (e.g. IMS name vs a
    manually-entered onboarding name), so callers that HAVE the learner's real
    id/email (any authenticated session) should always pass them; name is only
    a last-resort fallback for demo personas with no directory row at all.
    Returns the row (with 'id') or None."""
    if member_id is not None:
        cur.execute("SELECT * FROM onboarding_requests WHERE id=%s", (member_id,))
        row = cur.fetchone()
        if row:
            return row
    if email:
        cur.execute("SELECT * FROM onboarding_requests WHERE LOWER(email)=LOWER(%s) ORDER BY created_at DESC LIMIT 1", (email,))
        row = cur.fetchone()
        if row:
            return row
    if member_name:
        cur.execute("SELECT * FROM onboarding_requests WHERE LOWER(name)=LOWER(%s) ORDER BY created_at DESC LIMIT 1", (member_name,))
        row = cur.fetchone()
        if row:
            return row
    return None

@app.post("/api/tracks/enroll")
def enroll_track(body: TrackEnrollRequest):
    """Add a track to the learner's enrolled tracks list."""
    import json
    try:
        with get_db() as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                row = _resolve_member_where(cur, body.member_id, body.email, body.member)
                if not row: raise HTTPException(status_code=404,detail="Member not found")
                tracks=json.loads(row["enrolled_tracks"] or "[]")
                if body.track not in tracks: tracks.append(body.track)
                cur.execute("UPDATE onboarding_requests SET enrolled_tracks=%s WHERE id=%s",(json.dumps(tracks),row["id"]))
        return {"ok":True,"enrolled_tracks":tracks}
    except HTTPException: raise
    except Exception as e: raise HTTPException(status_code=500,detail=str(e))

class CrossSkillConfirmRequest(BaseModel):
    member: str = ""
    track: str
    include_prerequisites: bool = True
    member_id: Optional[int] = None
    email: Optional[str] = None


@app.post("/api/cross-skilling/confirm")
def confirm_cross_skill(body: CrossSkillConfirmRequest):
    """Learner confirms the AI-suggested track → start it. One call: enrol the
    member (so the track shows in their list) and build the ordered Curriculum
    learning path. This is the single target for the 'Confirm and start track'
    button after /api/agents/advisor recommends a track."""
    track = (body.track or "").strip()
    if not track:
        raise HTTPException(status_code=422, detail="track is required.")
    enrolled = None
    if body.member_id or body.email or body.member.strip():
        try:
            with get_db() as conn:
                with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                    row = _resolve_member_where(cur, body.member_id, body.email, body.member)
                    if row:
                        tracks = json.loads(row["enrolled_tracks"] or "[]")
                        if track not in tracks:
                            tracks.append(track)
                        cur.execute("UPDATE onboarding_requests SET enrolled_tracks=%s WHERE id=%s", (json.dumps(tracks), row["id"]))
                        enrolled = tracks
        except Exception as e:
            print(f"[cross-skilling.confirm] enrol skipped: {e}")
    try:
        from services import generate_learning_path
        path = generate_learning_path(track, include_prerequisites=body.include_prerequisites)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Could not build learning path: {e}")
    return {"ok": True, "track": track, "enrolled_tracks": enrolled, "learning_path": path}


@app.post("/api/tracks/unenroll")
def unenroll_track(body: TrackEnrollRequest):
    """Remove a track from the learner's enrolled tracks list. The learner's
    primary/active track cannot be removed (it's their core enablement path),
    only cross-skill tracks they added themselves."""
    import json
    try:
        with get_db() as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                row = _resolve_member_where(cur, body.member_id, body.email, body.member)
                if not row: raise HTTPException(status_code=404,detail="Member not found")
                primary=row["active_track"] or "rtcdp"
                if body.track==primary:
                    raise HTTPException(status_code=400,detail="Your primary enablement track can't be removed.")
                tracks=[t for t in json.loads(row["enrolled_tracks"] or "[]") if t!=body.track]
                cur.execute("UPDATE onboarding_requests SET enrolled_tracks=%s WHERE id=%s",(json.dumps(tracks),row["id"]))
        return {"ok":True,"enrolled_tracks":tracks}
    except HTTPException: raise
    except Exception as e: raise HTTPException(status_code=500,detail=str(e))

@app.get("/api/tracks/enrolled")
def get_enrolled_tracks(member: str, member_id: Optional[int] = None, email: Optional[str] = None):
    """Get all tracks a member is enrolled in."""
    import json
    try:
        with get_db() as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                row = _resolve_member_where(cur, member_id, email, member)
                if not row: return {"primary_track":"rtcdp","enrolled_tracks":[]}
                tracks=json.loads(row["enrolled_tracks"] or "[]")
                primary=row["active_track"] or "rtcdp"
                if primary not in tracks: tracks=[primary]+tracks
        return{"primary_track":primary,"enrolled_tracks":tracks}
    except Exception as e: raise HTTPException(status_code=500,detail=str(e))

@app.get("/api/tracks/progress")
def get_all_tracks_progress(member: str, manager: str):
    """Get completion count per track for a member."""
    try:
        with get_db() as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute("""
                    SELECT track, COUNT(*) as modules_done
                    FROM user_module_progress
                    WHERE member_name=%s AND manager=%s
                    GROUP BY track ORDER BY track""", (member, manager))
                rows=cur.fetchall()
        return{"progress":{r["track"]:int(r["modules_done"]) for r in rows}}
    except Exception as e: raise HTTPException(status_code=500,detail=str(e))

class TrackCapstoneBody(BaseModel):
    member: str
    track: str
    score: Optional[float] = None

@app.post("/api/tracks/capstone/complete")
def complete_track_capstone(body: TrackCapstoneBody):
    """Record that a member finished a specific track's capstone. Idempotent
    per (member, track). This is how capstone progress is tracked when a learner
    has multiple parallel cross-skill tracks — each track's capstone is its own
    row here, separate from the primary onboarding capstone flag."""
    track = (body.track or "").strip().lower()
    if not body.member.strip() or not track:
        raise HTTPException(status_code=422, detail="member and track are required.")
    try:
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("""INSERT INTO track_capstones (member_name, track, status, score, completed_at)
                    VALUES (%s,%s,'completed',%s,NOW())
                    ON CONFLICT (member_name, track)
                    DO UPDATE SET status='completed', score=EXCLUDED.score, completed_at=NOW()""",
                    (body.member.strip(), track, body.score))
        return {"ok": True, "member": body.member, "track": track}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/tracks/capstone")
def get_track_capstones(member: str):
    """List which tracks' capstones a member has completed (per-track progress)."""
    try:
        with get_db() as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute("""SELECT track, status, score, completed_at
                    FROM track_capstones WHERE LOWER(member_name)=LOWER(%s)
                    ORDER BY completed_at DESC""", (member,))
                rows = cur.fetchall()
        return {"completed_tracks": [r["track"] for r in rows],
                "detail": [{"track": r["track"], "status": r["status"],
                            "score": float(r["score"]) if r["score"] is not None else None,
                            "completed_at": str(r["completed_at"]) if r["completed_at"] else None}
                           for r in rows]}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/billing/summary")
def billing_summary(manager: str, member: str, year: Optional[int]=None, quarter: Optional[int]=None):
    """Total billable hours logged on project updates for a member in a quarter."""
    from datetime import date
    yr  = year    or date.today().year
    qtr = quarter or _current_quarter()
    q_start, q_end = _quarter_bounds(yr, qtr)
    try:
        with get_db() as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute(
                    """SELECT COALESCE(SUM(au.billable_hours),0) AS total_billed,
                              COUNT(DISTINCT au.week_of) AS weeks_logged
                       FROM allocation_updates au
                       JOIN project_allocations pa ON pa.id=au.allocation_id
                       WHERE au.member_name=%s AND pa.manager=%s
                         AND au.week_of BETWEEN %s AND %s
                         AND au.billable_hours > 0""",
                    (member, manager, q_start, q_end)
                )
                row = cur.fetchone()
                # Also get per-project breakdown
                cur.execute(
                    """SELECT pa.project_name,
                              COALESCE(SUM(au.billable_hours),0) AS project_billed
                       FROM allocation_updates au
                       JOIN project_allocations pa ON pa.id=au.allocation_id
                       WHERE au.member_name=%s AND pa.manager=%s
                         AND au.week_of BETWEEN %s AND %s
                       GROUP BY pa.project_name ORDER BY project_billed DESC""",
                    (member, manager, q_start, q_end)
                )
                breakdown = cur.fetchall()
        return {
            "quarter": f"Q{qtr} {yr}",
            "total_billed": float(row["total_billed"]) if row else 0,
            "weeks_logged": int(row["weeks_logged"]) if row else 0,
            "breakdown": [{"project": r["project_name"], "hours": float(r["project_billed"])} for r in breakdown],
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# ── Customer-Facing Utilization — weekly log + quarterly aggregate ──────────

def _quarter_bounds(year:int, quarter:int):
    starts = {1:(1,1),2:(4,1),3:(7,1),4:(10,1)}
    ends   = {1:(3,31),2:(6,30),3:(9,30),4:(12,31)}
    from datetime import date
    s=starts[quarter]; e=ends[quarter]
    return date(year,s[0],s[1]), date(year,e[0],e[1])

def _current_quarter():
    from datetime import date
    m=date.today().month
    return (m-1)//3+1

@app.post("/api/utilization")
def save_utilization(body: UtilizationEntry):
    """Upsert a member's weekly utilization entry."""
    try:
        with get_db() as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute(
                    """INSERT INTO user_utilization
                       (member_name,manager,week_of,billable_hours,non_billable_cf_hours,
                        ramp_credit,working_hours,holiday_hours,loa_hours,cf_target)
                       VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                       ON CONFLICT (member_name,week_of) DO UPDATE SET
                         manager=%s, billable_hours=%s, non_billable_cf_hours=%s,
                         ramp_credit=%s, working_hours=%s, holiday_hours=%s,
                         loa_hours=%s, cf_target=%s, updated_at=NOW()
                       RETURNING id""",
                    (body.member_name,body.manager,body.week_of,body.billable_hours,
                     body.non_billable_cf_hours,body.ramp_credit,body.working_hours,
                     body.holiday_hours,body.loa_hours,body.cf_target,
                     body.manager,body.billable_hours,body.non_billable_cf_hours,
                     body.ramp_credit,body.working_hours,body.holiday_hours,
                     body.loa_hours,body.cf_target)
                )
                row=cur.fetchone()
        return {"ok":True,"id":row["id"]}
    except Exception as e:
        raise HTTPException(status_code=500,detail=str(e))

@app.get("/api/utilization")
def get_utilization(manager:str, member:str, year:Optional[int]=None, quarter:Optional[int]=None):
    """Get aggregated quarterly utilization for one member.
    Defaults to current quarter. Returns summed hours + computed metrics."""
    from datetime import date
    yr  = year    or date.today().year
    qtr = quarter or _current_quarter()
    q_start, q_end = _quarter_bounds(yr, qtr)
    try:
        with get_db() as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute(
                    """SELECT
                         COALESCE(SUM(billable_hours),0)         AS billable_hours,
                         COALESCE(SUM(non_billable_cf_hours),0)  AS non_billable_cf_hours,
                         COALESCE(SUM(ramp_credit),0)            AS ramp_credit,
                         COALESCE(SUM(working_hours),0)          AS working_hours,
                         COALESCE(SUM(holiday_hours),0)          AS holiday_hours,
                         COALESCE(SUM(loa_hours),0)              AS loa_hours,
                         MAX(cf_target)                          AS cf_target,
                         COUNT(*)                                AS weeks_logged
                       FROM user_utilization
                       WHERE member_name=%s AND manager=%s
                         AND week_of BETWEEN %s AND %s""",
                    (member, manager, q_start, q_end)
                )
                row = cur.fetchone()
        def compute(r):
            if not r or r["weeks_logged"]==0: return None
            avail=max(float(r["working_hours"])-float(r["holiday_hours"])-float(r["loa_hours"]),0.01)
            cf_total=float(r["billable_hours"])+float(r["non_billable_cf_hours"])+float(r["ramp_credit"])
            cf_target=float(r["cf_target"] or 75)
            cf_util=round((cf_total/avail)*100,1)
            cf_achieved=round((cf_util/cf_target)*100,1) if cf_target>0 else 0
            return{"billable_hours":float(r["billable_hours"]),
                   "non_billable_cf_hours":float(r["non_billable_cf_hours"]),
                   "ramp_credit":float(r["ramp_credit"]),
                   "working_hours":float(r["working_hours"]),
                   "holiday_hours":float(r["holiday_hours"]),
                   "loa_hours":float(r["loa_hours"]),
                   "cf_target":cf_target,
                   "available_hours":round(avail,1),
                   "cf_hours_total":round(cf_total,1),
                   "cf_utilization":cf_util,
                   "cf_target_achieved":cf_achieved,
                   "weeks_logged":int(r["weeks_logged"])}
        return{"quarter":f"Q{qtr} {yr}","quarter_start":str(q_start),
               "quarter_end":str(q_end),"entry":compute(row)}
    except Exception as e:
        raise HTTPException(status_code=500,detail=str(e))

@app.get("/api/utilization/team")
def get_team_utilization(manager:str, year:Optional[int]=None, quarter:Optional[int]=None):
    """Quarterly utilization for all members in a manager's team."""
    from datetime import date
    yr  = year    or date.today().year
    qtr = quarter or _current_quarter()
    q_start, q_end = _quarter_bounds(yr, qtr)
    try:
        with get_db() as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute(
                    """SELECT member_name,
                         COALESCE(SUM(billable_hours),0)        AS billable_hours,
                         COALESCE(SUM(non_billable_cf_hours),0) AS non_billable_cf_hours,
                         COALESCE(SUM(ramp_credit),0)           AS ramp_credit,
                         COALESCE(SUM(working_hours),0)         AS working_hours,
                         COALESCE(SUM(holiday_hours),0)         AS holiday_hours,
                         COALESCE(SUM(loa_hours),0)             AS loa_hours,
                         MAX(cf_target)                         AS cf_target,
                         COUNT(*)                               AS weeks_logged
                       FROM user_utilization
                       WHERE manager=%s AND week_of BETWEEN %s AND %s
                       GROUP BY member_name ORDER BY member_name""",
                    (manager, q_start, q_end)
                )
                rows=cur.fetchall()
        def compute(r):
            avail=max(float(r["working_hours"])-float(r["holiday_hours"])-float(r["loa_hours"]),0.01)
            cf_total=float(r["billable_hours"])+float(r["non_billable_cf_hours"])+float(r["ramp_credit"])
            cf_target=float(r["cf_target"] or 75)
            cf_util=round((cf_total/avail)*100,1)
            return{**dict(r),"member_name":r["member_name"],
                   "available_hours":round(avail,1),"cf_hours_total":round(cf_total,1),
                   "cf_utilization":cf_util,"cf_target":cf_target,
                   "cf_target_achieved":round((cf_util/cf_target)*100,1) if cf_target>0 else 0,
                   "weeks_logged":int(r["weeks_logged"])}
        return{"quarter":f"Q{qtr} {yr}","quarter_start":str(q_start),
               "quarter_end":str(q_end),"members":[compute(r) for r in rows]}
    except Exception as e:
        raise HTTPException(status_code=500,detail=str(e))

@app.get("/api/allocations/team-feed")
def get_team_feed(manager: str, limit: int = 30):
    """All weekly updates across the whole team, newest first — for the manager's
    'All Milestones'-style feed. Joins through project_allocations to get project names."""
    _ensure_tracker_tables()
    try:
        with get_db() as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute(
                    """SELECT u.id, u.member_name, u.comment, u.health_status, u.created_at,
                              a.project_id, a.project_name
                       FROM allocation_updates u
                       JOIN project_allocations a ON a.id = u.allocation_id
                       WHERE a.manager=%s
                       ORDER BY u.created_at DESC
                       LIMIT %s""",
                    (manager, limit)
                )
                rows = cur.fetchall()
        for r in rows:
            r["created_at"] = str(r["created_at"])
        return {"feed": [dict(r) for r in rows]}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# ── Email reminders ────────────────────────────────────────────────────────────

@app.get("/api/email/status")
def email_status():
    """Check whether SMTP is configured, without exposing credentials."""
    return {"enabled": EMAIL_ENABLED, "host": SMTP_HOST or None, "from": SMTP_FROM or None}

@app.post("/api/notify/weekly-reminder")
def send_weekly_reminders(manager: str):
    """Notifies every approved team member to update their Weekly Tracker.
    Always creates an in-app notification; also emails if SMTP is configured."""
    try:
        with get_db() as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute(
                    "SELECT name, email FROM onboarding_requests WHERE status='approved' AND manager=%s",
                    (manager,)
                )
                rows = cur.fetchall()
        sent = 0
        failures = []
        for r in rows:
            create_notification(r["name"], manager, "weekly_reminder",
                "Time for your weekly update",
                "It's time to log this week's progress in your Weekly Tracker — add any new projects and post an update on what changed.")
            sent += 1
            if EMAIL_ENABLED and r.get("email"):
                html = email_template(
                    "Time for your weekly update",
                    f"Hi {r['name']}, it's time to log this week's progress in your Nexus Weekly Tracker — "
                    f"add any new projects and post an update on what changed.",
                    "Open Weekly Tracker", "http://localhost:5173"
                )
                result = send_email(r["email"], "Nexus — weekly tracker reminder", html)
                if not result.get("ok"):
                    failures.append({"email": r["email"], "error": result.get("error")})
        return {"ok": True, "sent": sent, "total": len(rows), "failures": failures, "email_enabled": EMAIL_ENABLED}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/notify/capstone-check")
def check_capstone_deadlines(manager: str, days_threshold: int = 7):
    """Notifies learners whose capstone has been open longer than the threshold (default 7 days).
    Always creates an in-app notification; also emails if SMTP is configured."""
    try:
        with get_db() as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute("ALTER TABLE onboarding_requests ADD COLUMN IF NOT EXISTS capstone_started_at TIMESTAMP")
                cur.execute(
                    """SELECT name, email, capstone_started_at FROM onboarding_requests
                       WHERE status='approved' AND manager=%s
                       AND capstone_started_at IS NOT NULL
                       AND capstone_completed=FALSE
                       AND capstone_started_at < NOW() - INTERVAL '%s days'""",
                    (manager, days_threshold)
                )
                rows = cur.fetchall()
        sent = 0
        for r in rows:
            create_notification(r["name"], manager, "capstone_overdue",
                "Your capstone is overdue",
                f"Your capstone has been open for more than {days_threshold} days. Reach out to your manager if you're blocked, or head back in to finish it up.")
            sent += 1
            if EMAIL_ENABLED and r.get("email"):
                html = email_template(
                    "Your capstone is overdue",
                    f"Hi {r['name']}, your capstone has been open for more than {days_threshold} days. "
                    f"Reach out to your manager if you're blocked, or head back in to finish it up.",
                    "Open Capstone", "http://localhost:5173"
                )
                send_email(r["email"], "Nexus — capstone deadline reminder", html)
        return {"ok": True, "sent": sent, "overdue": len(rows)}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ── Capstone duration tracking ─────────────────────────────────────────────────

@app.put("/api/onboarding/{request_id}/capstone-start")
def start_capstone_clock(request_id: int):
    """Called once, the first time a learner's confidence crosses the gate and the
    Capstone tab unlocks. Only sets the timestamp if it hasn't been set already."""
    try:
        with get_db() as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute("ALTER TABLE onboarding_requests ADD COLUMN IF NOT EXISTS capstone_started_at TIMESTAMP")
                cur.execute(
                    """UPDATE onboarding_requests SET capstone_started_at=NOW()
                       WHERE id=%s AND capstone_started_at IS NULL
                       RETURNING capstone_started_at""",
                    (request_id,)
                )
                row = cur.fetchone()
        return {"ok": True, "capstone_started_at": str(row["capstone_started_at"]) if row else None}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ── Module test-out (skip a module by passing a quiz at >=90%) ────────────────

class TestOutResult(BaseModel):
    member_name: str
    manager: str
    track: str
    module_id: int
    module_title: str
    score: float          # 0-100
    total_questions: int
    correct_answers: int

@app.post("/api/test-out")
def record_test_out(body: TestOutResult):
    """Record a module test-out attempt. Passing requires score >= 90."""
    passed = body.score >= 90.0
    try:
        conn = psycopg2.connect(DATABASE_URL)
        conn.autocommit = True
        with conn.cursor() as c:
            c.execute("""
                CREATE TABLE IF NOT EXISTS module_test_outs (
                    id SERIAL PRIMARY KEY,
                    member_name VARCHAR(120) NOT NULL,
                    manager VARCHAR(120),
                    track VARCHAR(50),
                    module_id INTEGER,
                    module_title VARCHAR(255),
                    score NUMERIC,
                    total_questions INTEGER,
                    correct_answers INTEGER,
                    passed BOOLEAN,
                    created_at TIMESTAMP DEFAULT NOW()
                )
            """)
        conn.close()
        with get_db() as conn2:
            with conn2.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute(
                    """INSERT INTO module_test_outs
                       (member_name, manager, track, module_id, module_title, score, total_questions, correct_answers, passed)
                       VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id""",
                    (body.member_name, body.manager, body.track, body.module_id, body.module_title,
                     body.score, body.total_questions, body.correct_answers, passed)
                )
                row = cur.fetchone()
        if passed:
            # Passing a test-out counts as completing the module — feeds the same
            # progress table that "mark complete" uses, and awards points once.
            _ensure_progress_tables()
            try:
                with get_db() as conn3:
                    with conn3.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur3:
                        cur3.execute(
                            """INSERT INTO user_module_progress (member_name, manager, track, module_id, module_title, via)
                               VALUES (%s,%s,%s,%s,%s,'test-out')
                               ON CONFLICT (member_name, track, module_id) DO NOTHING
                               RETURNING id""",
                            (body.member_name, body.manager, body.track, body.module_id, body.module_title)
                        )
                        newly = cur3.fetchone()
                if newly:
                    award_points(body.member_name, body.manager, 50, f"Tested out: {body.module_title}")
            except Exception:
                pass
        return {"ok": True, "id": row["id"], "passed": passed, "score": body.score, "threshold": 90.0}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/test-out")
def get_test_outs(member_name: str, track: str):
    """Modules this member has already tested out of (passing attempts only)."""
    try:
        with get_db() as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute(
                    """SELECT module_id, module_title, score, created_at FROM module_test_outs
                       WHERE member_name=%s AND track=%s AND passed=TRUE
                       ORDER BY created_at DESC""",
                    (member_name, track)
                )
                rows = cur.fetchall()
        for r in rows:
            r["created_at"] = str(r["created_at"])
        return {"passed_modules": [dict(r) for r in rows]}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ── Generalised guardrails — every agent, not just Socratic ───────────────────

class GenericGuardrail(BaseModel):
    agent_name: str
    score: float          # 1-10
    issue: Optional[str] = None
    response_preview: Optional[str] = None

@app.post("/api/guardrail/generic")
def save_generic_guardrail(body: GenericGuardrail):
    """Log a guardrail score for any agent (Curriculum, Evaluation, Capstone, Study, Practice Scenarios).
    Reuses guardrail_logs, adding an agent_name column so all agents share one table."""
    try:
        conn = psycopg2.connect(DATABASE_URL)
        conn.autocommit = True
        with conn.cursor() as c:
            c.execute("ALTER TABLE guardrail_logs ADD COLUMN IF NOT EXISTS agent_name VARCHAR(50) DEFAULT 'socratic'")
        conn.close()
        with get_db() as conn2:
            with conn2.cursor() as cur:
                cur.execute(
                    """INSERT INTO guardrail_logs (agent_name, score, issue, response_preview, word_count, has_one_question, avoids_direct_answer)
                       VALUES (%s,%s,%s,%s,NULL,NULL,NULL)""",
                    (body.agent_name, body.score, body.issue, body.response_preview)
                )
        return {"ok": True}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/guardrail/by-agent")
def guardrail_by_agent():
    """Average guardrail score per agent — for the Admin AI Safety view."""
    try:
        conn = psycopg2.connect(DATABASE_URL)
        conn.autocommit = True
        with conn.cursor() as c:
            c.execute("ALTER TABLE guardrail_logs ADD COLUMN IF NOT EXISTS agent_name VARCHAR(50) DEFAULT 'socratic'")
        conn.close()
        with get_db() as conn2:
            with conn2.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute(
                    """SELECT COALESCE(agent_name,'socratic') as agent_name, COUNT(*) as total, AVG(score) as avg_score
                       FROM guardrail_logs GROUP BY agent_name ORDER BY agent_name"""
                )
                rows = cur.fetchall()
        return {"agents": [dict(r) for r in rows]}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/admin/ai-safety")
def admin_ai_safety():
    """Everything the AI Safety admin view needs, in one call:
      - injection_blocks_by_agent: real counts from llm_logs (model='guardrail'
        rows — every call through llm_call()/call_with_tools() logs a row here
        when check_input_guardrail() refuses a request, regardless of agent)
      - ragas: per-agent RAGAS quality summary + recent evaluation rows
      - genericGuardrail: the existing /api/guardrail/by-agent heuristic scores
        (Capstone/CapstoneHint/CrossSkilling/Practice), folded in here so the
        UI has one endpoint instead of three round trips.
    """
    result = {"injection_blocks_by_agent": [], "ragas_summary": {}, "ragas_recent": [],
              "ragas_thresholds": {"good": 0.7, "warn": 0.4},
              "generic_guardrail_by_agent": []}
    try:
        with get_db() as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute("""
                    SELECT agent_name, COUNT(*) as blocked_count, MAX(created_at) as last_blocked
                    FROM llm_logs WHERE model='guardrail' AND success=false
                    GROUP BY agent_name ORDER BY blocked_count DESC
                """)
                result["injection_blocks_by_agent"] = [dict(r) for r in cur.fetchall()]
    except Exception as e:
        result["injection_blocks_error"] = str(e)

    try:
        from evaluation import get_evaluation_summary, get_recent_evaluations, get_ragas_thresholds
        result["ragas_summary"] = get_evaluation_summary()
        result["ragas_recent"] = get_recent_evaluations(limit=20)
        result["ragas_thresholds"] = get_ragas_thresholds()
    except Exception as e:
        result["ragas_error"] = str(e)

    try:
        conn = psycopg2.connect(DATABASE_URL)
        conn.autocommit = True
        with conn.cursor() as c:
            c.execute("ALTER TABLE guardrail_logs ADD COLUMN IF NOT EXISTS agent_name VARCHAR(50) DEFAULT 'socratic'")
        conn.close()
        with get_db() as conn2:
            with conn2.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute("""SELECT COALESCE(agent_name,'socratic') as agent_name, COUNT(*) as total, AVG(score) as avg_score
                               FROM guardrail_logs GROUP BY agent_name ORDER BY agent_name""")
                result["generic_guardrail_by_agent"] = [dict(r) for r in cur.fetchall()]
    except Exception as e:
        result["generic_guardrail_error"] = str(e)

    return result


@app.get("/api/admin/agent-prompts")
def admin_agent_prompts():
    """The REAL current system prompt for every agent, read straight from the
    running backend code — not a hardcoded client-side mock. Static prompts
    (Curriculum/CrossSkilling/Capstone/Practice/RAG/Socratic) are returned
    verbatim, {placeholder} tokens and all, so the admin sees exactly what's
    templated at runtime. Reasoning and Study Aid build their prompt
    dynamically per learner context (confidence, track, retrieved topics) —
    those are rendered with representative example values and labeled as such,
    since there's no single static string to show."""
    prompts = {}
    errors = {}

    def _get(agent_key, fn):
        try:
            prompts[agent_key] = fn()
        except Exception as e:
            errors[agent_key] = str(e)

    _get("curriculum", lambda: __import__("agents.curriculum", fromlist=["GUIDANCE_SYSTEM"]).GUIDANCE_SYSTEM)
    _get("crossskill_recommend", lambda: __import__("agents.crossskill", fromlist=["GUIDANCE_SYSTEM"]).GUIDANCE_SYSTEM)
    _get("crossskill_chat", lambda: __import__("agents.crossskill", fromlist=["CHAT_SYSTEM"]).CHAT_SYSTEM)
    _get("capstone", lambda: __import__("agents.capstone", fromlist=["CAPSTONE_SYSTEM"]).CAPSTONE_SYSTEM)
    _get("practice_scenario", lambda: __import__("agents.practice", fromlist=["SCENARIO_SYSTEM"]).SCENARIO_SYSTEM)
    _get("practice_validate", lambda: __import__("agents.practice", fromlist=["VALIDATION_SYSTEM"]).VALIDATION_SYSTEM)
    _get("rag_rewrite", lambda: __import__("agents.rag", fromlist=["REWRITE_SYSTEM"]).REWRITE_SYSTEM)
    _get("rag_rerank", lambda: __import__("agents.rag", fromlist=["RERANK_SYSTEM"]).RERANK_SYSTEM)
    _get("rag_answer", lambda: __import__("agents.rag", fromlist=["ANSWER_SYSTEM"]).ANSWER_SYSTEM)
    _get("rag_guard", lambda: __import__("agents.rag", fromlist=["GUARD_SYSTEM"]).GUARD_SYSTEM)
    _get("socratic", lambda: __import__("agents.socratic_agent", fromlist=["_GUIDED_SYSTEM"])._GUIDED_SYSTEM)

    def _study_aid_example():
        from agents.study_aid import _build_system
        return _build_system("Real-Time CDP", "Segmentation Basics",
                              "1. Segmentation Fundamentals — understand modes\n2. Audience Activation — send to destinations",
                              "\nLEARNER CONFIDENCE: 0.55 (MODERATE). Mix straightforward application cards with a couple of 'what happens if' scenarios.\n")
    _get("study_aid", _study_aid_example)

    def _reasoning_example():
        from agents.reasoning import _build_system_prompt
        return _build_system_prompt({
            "profile": {"name": "Example Learner", "role": "Analytics Engineer"},
            "learner_context": "", "track": "rtcdp",
            "intent": "new_question", "confidence": 0.6,
            "extra": {"learner_context": {}},
        })
    _get("reasoning", _reasoning_example)

    dynamic = {"study_aid", "reasoning"}
    return {
        "prompts": {k: {"text": v, "dynamic": k in dynamic} for k, v in prompts.items()},
        "errors": errors,
    }


# ── Real per-user module progress (replaces static per-track module status) ───

def _ensure_progress_tables():
    try:
        conn = psycopg2.connect(DATABASE_URL)
        conn.autocommit = True
        with conn.cursor() as c:
            c.execute("""
                CREATE TABLE IF NOT EXISTS user_module_progress (
                    id SERIAL PRIMARY KEY,
                    member_name VARCHAR(120) NOT NULL,
                    manager VARCHAR(120),
                    track VARCHAR(50) NOT NULL,
                    module_id INTEGER NOT NULL,
                    module_title VARCHAR(255),
                    via VARCHAR(20) DEFAULT 'completed',
                    completed_at TIMESTAMP DEFAULT NOW(),
                    UNIQUE(member_name, track, module_id)
                )
            """)
            c.execute("""
                CREATE TABLE IF NOT EXISTS points_ledger (
                    id SERIAL PRIMARY KEY,
                    member_name VARCHAR(120) NOT NULL,
                    manager VARCHAR(120),
                    points INTEGER NOT NULL,
                    reason VARCHAR(255),
                    created_at TIMESTAMP DEFAULT NOW()
                )
            """)
        conn.close()
    except Exception:
        pass

def award_points(member_name: str, manager: str, points: int, reason: str):
    """Best-effort points award — never blocks or fails the calling endpoint."""
    try:
        _ensure_progress_tables()
        conn = psycopg2.connect(DATABASE_URL)
        conn.autocommit = True
        with conn.cursor() as c:
            c.execute(
                "INSERT INTO points_ledger (member_name, manager, points, reason) VALUES (%s,%s,%s,%s)",
                (member_name, manager, points, reason)
            )
        conn.close()
    except Exception:
        pass

class ModuleComplete(BaseModel):
    member_name: str
    manager: str
    track: str
    module_id: int
    module_title: Optional[str] = ""
    via: Optional[str] = "completed"  # "completed" | "test-out"

@app.post("/api/progress/complete")
def mark_module_complete(body: ModuleComplete):
    """Mark a module done for this member+track. Idempotent. Awards points once."""
    _ensure_progress_tables()
    try:
        with get_db() as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute(
                    """INSERT INTO user_module_progress (member_name, manager, track, module_id, module_title, via)
                       VALUES (%s,%s,%s,%s,%s,%s)
                       ON CONFLICT (member_name, track, module_id) DO NOTHING
                       RETURNING id""",
                    (body.member_name, body.manager, body.track, body.module_id, body.module_title, body.via)
                )
                row = cur.fetchone()
        if row:  # only award points the first time, not on repeat calls
            award_points(body.member_name, body.manager, 50, f"Completed: {body.module_title or ('Module '+str(body.module_id))}")
        return {"ok": True, "newly_completed": bool(row)}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/progress")
def get_progress(member_name: str, track: str):
    """Which module ids this member has completed on this track."""
    _ensure_progress_tables()
    try:
        with get_db() as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute(
                    "SELECT module_id, via, completed_at FROM user_module_progress WHERE member_name=%s AND track=%s ORDER BY module_id",
                    (member_name, track)
                )
                rows = cur.fetchall()
        for r in rows:
            r["completed_at"] = str(r["completed_at"])
        return {"completed": [r["module_id"] for r in rows], "detail": [dict(r) for r in rows]}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ── Points & team leaderboard ───────────────────────────────────────────────────

@app.get("/api/points/me")
def get_my_points(member_name: str):
    _ensure_progress_tables()
    try:
        with get_db() as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute("SELECT COALESCE(SUM(points),0) as total FROM points_ledger WHERE member_name=%s", (member_name,))
                total = cur.fetchone()["total"]
                cur.execute("SELECT points, reason, created_at FROM points_ledger WHERE member_name=%s ORDER BY created_at DESC LIMIT 15", (member_name,))
                recent = cur.fetchall()
        for r in recent:
            r["created_at"] = str(r["created_at"])
        return {"total": int(total), "recent": [dict(r) for r in recent]}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/points/team")
def get_team_leaderboard(manager: str):
    """Ranks every member who has any points, under this manager."""
    _ensure_progress_tables()
    try:
        with get_db() as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute(
                    """SELECT member_name, SUM(points) as total
                       FROM points_ledger WHERE manager=%s
                       GROUP BY member_name ORDER BY total DESC""",
                    (manager,)
                )
                rows = cur.fetchall()
        return {"leaderboard": [{"member_name": r["member_name"], "total": int(r["total"])} for r in rows]}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ── Live team summary for Manager dashboards (Team Overview / ROI) ────────────

@app.get("/api/team/live-summary")
def team_live_summary(manager: str):
    """Aggregated real numbers for a manager's registered, approved team —
    used to replace static demo figures on Team Overview and ROI Velocity."""
    _ensure_progress_tables()
    try:
        with get_db() as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute("ALTER TABLE onboarding_requests ADD COLUMN IF NOT EXISTS capstone_started_at TIMESTAMP")
                cur.execute(
                    """SELECT id, name, preferred_name, email, team, joining_date,
                              COALESCE(capstone_completed, FALSE) as capstone_completed,
                              capstone_completed_at, capstone_started_at,
                              COALESCE(active_track, 'rtcdp') as active_track
                       FROM onboarding_requests WHERE status='approved' AND manager=%s""",
                    (manager,)
                )
                members = cur.fetchall()

                cur.execute(
                    """SELECT member_name, track, COUNT(*) as modules_done
                       FROM user_module_progress WHERE manager=%s GROUP BY member_name, track""",
                    (manager,)
                )
                progress_by_member = {}
                for r in cur.fetchall():
                    progress_by_member.setdefault(r["member_name"], {})[r["track"]] = r["modules_done"]

                cur.execute(
                    "SELECT member_name, SUM(points) as total FROM points_ledger WHERE manager=%s GROUP BY member_name",
                    (manager,)
                )
                points_by_member = {r["member_name"]: int(r["total"]) for r in cur.fetchall()}

        from datetime import date
        out = []
        days_to_capstone = []
        for m in members:
            track = m["active_track"] or "rtcdp"
            modules_done = progress_by_member.get(m["name"], {}).get(track, 0)
            try:
                days_since_joining = (date.today() - m["joining_date"]).days if m["joining_date"] else None
            except Exception:
                days_since_joining = None
            if m["capstone_completed"] and m["capstone_completed_at"] and m["joining_date"]:
                try:
                    days_to_capstone.append((m["capstone_completed_at"].date() - m["joining_date"]).days)
                except Exception:
                    pass
            out.append({
                "name": m["name"], "email": m["email"], "team": m["team"], "track": track,
                "modules_done": modules_done,
                "capstone_completed": bool(m["capstone_completed"]),
                "days_since_joining": days_since_joining,
                "points": points_by_member.get(m["name"], 0),
            })

        avg_days_to_capstone = round(sum(days_to_capstone) / len(days_to_capstone), 1) if days_to_capstone else None
        at_risk = [m for m in out if m["days_since_joining"] is not None and m["days_since_joining"] > 56 and not m["capstone_completed"]]

        return {
            "members": out,
            "total_members": len(out),
            "avg_days_to_capstone": avg_days_to_capstone,
            "capstones_completed": sum(1 for m in out if m["capstone_completed"]),
            "at_risk_count": len(at_risk),
            "at_risk_names": [m["name"] for m in at_risk],
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/cohort/ranking")
def cohort_ranking(email: str):
    """Real peer-progress ranking for a new joiner's home page: other
    approved learners on the same active_track, ranked by modules actually
    completed (user_module_progress) then points (points_ledger). No
    fictional names — only people who are actually registered and learning."""
    conn = psycopg2.connect(DATABASE_URL)
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as c:
            c.execute("""SELECT name, COALESCE(active_track,'rtcdp') AS active_track
                         FROM onboarding_requests WHERE LOWER(email)=LOWER(%s)""", (email.strip(),))
            me = c.fetchone()
            if not me:
                return {"track": None, "cohort": []}
            track = me["active_track"]

            c.execute("""SELECT name, email FROM onboarding_requests
                         WHERE COALESCE(active_track,'rtcdp')=%s AND status='approved'
                         AND COALESCE(role,'') NOT IN ('manager','admin')""", (track,))
            peers = c.fetchall()
            names = [p["name"] for p in peers]

            progress_by_name, points_by_name = {}, {}
            if names:
                c.execute("""SELECT member_name, COUNT(*) AS modules_done FROM user_module_progress
                             WHERE track=%s AND member_name=ANY(%s) GROUP BY member_name""", (track, names))
                progress_by_name = {r["member_name"]: r["modules_done"] for r in c.fetchall()}
                c.execute("""SELECT member_name, SUM(points) AS total FROM points_ledger
                             WHERE member_name=ANY(%s) GROUP BY member_name""", (names,))
                points_by_name = {r["member_name"]: int(r["total"] or 0) for r in c.fetchall()}

            board = [{
                "name": p["name"],
                "modules_done": progress_by_name.get(p["name"], 0),
                "points": points_by_name.get(p["name"], 0),
                "is_you": p["email"].strip().lower() == email.strip().lower(),
            } for p in peers]
            board.sort(key=lambda x: (-x["modules_done"], -x["points"], x["name"]))
            for i, b in enumerate(board):
                b["rank"] = i + 1
    finally:
        conn.close()
    return {"track": track, "cohort": board}


def _tenure_band(years):
    """Buckets years-of-tenure into a small set of bands used to group
    experienced staff into a cohort. Kept as one function so the bands are
    defined in exactly one place."""
    if years is None: return None
    if years < 1: return "0-1 yr"
    if years < 3: return "1-3 yrs"
    if years < 5: return "3-5 yrs"
    return "5+ yrs"


@app.get("/api/cohort/exp-ranking")
def cohort_exp_ranking(email: str):
    """Real peer-progress ranking for an experienced employee's home page:
    other approved employees on the SAME team AND in the SAME tenure band
    (years since joining_date), ranked by modules completed then points.
    No fictional names — only people actually registered in this team/band."""
    from datetime import date
    conn = psycopg2.connect(DATABASE_URL)
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as c:
            c.execute("""SELECT name, team, joining_date FROM onboarding_requests
                         WHERE LOWER(email)=LOWER(%s)""", (email.strip(),))
            me = c.fetchone()
            if not me or not me["team"]:
                return {"team": None, "tenure_band": None, "cohort": []}

            my_years = (date.today() - me["joining_date"]).days / 365.25 if me["joining_date"] else None
            my_band = _tenure_band(my_years)

            c.execute("""SELECT name, email, joining_date FROM onboarding_requests
                         WHERE team=%s AND status='approved'
                         AND COALESCE(role,'') NOT IN ('manager','admin')""", (me["team"],))
            peers = [p for p in c.fetchall()
                     if _tenure_band((date.today() - p["joining_date"]).days / 365.25 if p["joining_date"] else None) == my_band]
            names = [p["name"] for p in peers]

            progress_by_name, points_by_name = {}, {}
            if names:
                c.execute("""SELECT member_name, COUNT(*) AS modules_done FROM user_module_progress
                             WHERE member_name=ANY(%s) GROUP BY member_name""", (names,))
                progress_by_name = {r["member_name"]: r["modules_done"] for r in c.fetchall()}
                c.execute("""SELECT member_name, SUM(points) AS total FROM points_ledger
                             WHERE member_name=ANY(%s) GROUP BY member_name""", (names,))
                points_by_name = {r["member_name"]: int(r["total"] or 0) for r in c.fetchall()}

            board = [{
                "name": p["name"],
                "modules_done": progress_by_name.get(p["name"], 0),
                "points": points_by_name.get(p["name"], 0),
                "is_you": p["email"].strip().lower() == email.strip().lower(),
            } for p in peers]
            board.sort(key=lambda x: (-x["modules_done"], -x["points"], x["name"]))
            for i, b in enumerate(board):
                b["rank"] = i + 1
    finally:
        conn.close()
    return {"team": me["team"], "tenure_band": my_band, "cohort": board}


# ── In-app notifications (replaces email for now — no SMTP dependency/cost) ───

def _ensure_notification_tables():
    try:
        conn = psycopg2.connect(DATABASE_URL)
        conn.autocommit = True
        with conn.cursor() as c:
            c.execute("""
                CREATE TABLE IF NOT EXISTS notifications (
                    id SERIAL PRIMARY KEY,
                    member_name VARCHAR(120) NOT NULL,
                    manager VARCHAR(120),
                    type VARCHAR(50),
                    title VARCHAR(255),
                    message TEXT,
                    is_read BOOLEAN DEFAULT FALSE,
                    created_at TIMESTAMP DEFAULT NOW()
                )
            """)
        conn.close()
    except Exception:
        pass

def create_notification(member_name: str, manager: str, ntype: str, title: str, message: str):
    """Best-effort — never blocks or fails the calling endpoint."""
    try:
        _ensure_notification_tables()
        conn = psycopg2.connect(DATABASE_URL)
        conn.autocommit = True
        with conn.cursor() as c:
            c.execute(
                "INSERT INTO notifications (member_name, manager, type, title, message) VALUES (%s,%s,%s,%s,%s)",
                (member_name, manager, ntype, title, message)
            )
        conn.close()
    except Exception:
        pass

@app.get("/api/notifications")
def get_notifications(member_name: str, limit: int = 30):
    _ensure_notification_tables()
    try:
        with get_db() as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute(
                    "SELECT id, type, title, message, is_read, created_at FROM notifications WHERE member_name=%s ORDER BY created_at DESC LIMIT %s",
                    (member_name, limit)
                )
                rows = cur.fetchall()
                cur.execute("SELECT COUNT(*) as unread FROM notifications WHERE member_name=%s AND is_read=FALSE", (member_name,))
                unread = cur.fetchone()["unread"]
        for r in rows:
            r["created_at"] = str(r["created_at"])
        return {"notifications": [dict(r) for r in rows], "unread": unread}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.put("/api/notifications/{notif_id}/read")
def mark_notification_read(notif_id: int):
    try:
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("UPDATE notifications SET is_read=TRUE WHERE id=%s", (notif_id,))
        return {"ok": True}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.put("/api/notifications/read-all")
def mark_all_notifications_read(member_name: str):
    _ensure_notification_tables()
    try:
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("UPDATE notifications SET is_read=TRUE WHERE member_name=%s AND is_read=FALSE", (member_name,))
        return {"ok": True}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ── Persisted skill assessments (CAT quiz results, team-wide) ─────────────────

class SkillAssessment(BaseModel):
    member_name: str
    manager: str
    skill: str
    level: str          # none | developing | proficient | expert
    theta: Optional[float] = None

def _ensure_skill_tables():
    try:
        conn = psycopg2.connect(DATABASE_URL)
        conn.autocommit = True
        with conn.cursor() as c:
            c.execute("""
                CREATE TABLE IF NOT EXISTS skill_assessments (
                    id SERIAL PRIMARY KEY,
                    member_name VARCHAR(120) NOT NULL,
                    manager VARCHAR(120),
                    skill VARCHAR(100) NOT NULL,
                    level VARCHAR(20),
                    theta NUMERIC,
                    assessed_at TIMESTAMP DEFAULT NOW(),
                    UNIQUE(member_name, skill)
                )
            """)
        conn.close()
    except Exception:
        pass

@app.post("/api/skills/assess")
def save_skill_assessment(body: SkillAssessment):
    """Upsert — retaking a skill's CAT quiz updates the existing row rather than duplicating."""
    _ensure_skill_tables()
    try:
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    """INSERT INTO skill_assessments (member_name, manager, skill, level, theta, assessed_at)
                       VALUES (%s,%s,%s,%s,%s,NOW())
                       ON CONFLICT (member_name, skill) DO UPDATE
                       SET level=EXCLUDED.level, theta=EXCLUDED.theta, manager=EXCLUDED.manager, assessed_at=NOW()""",
                    (body.member_name, body.manager, body.skill, body.level, body.theta)
                )
        return {"ok": True}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/skills/me")
def get_my_skills(member_name: str):
    """Restores previous CAT quiz results after a refresh/relogin."""
    _ensure_skill_tables()
    try:
        with get_db() as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute("SELECT skill, level, theta, assessed_at FROM skill_assessments WHERE member_name=%s", (member_name,))
                rows = cur.fetchall()
        for r in rows:
            r["assessed_at"] = str(r["assessed_at"])
        return {"skills": {r["skill"]: r["level"] for r in rows}, "detail": [dict(r) for r in rows]}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/skills/team")
def get_team_skills(manager: str):
    """Real, persisted skill levels across every registered member who has taken at least one assessment."""
    _ensure_skill_tables()
    try:
        with get_db() as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute(
                    "SELECT member_name, skill, level, theta, assessed_at FROM skill_assessments WHERE manager=%s ORDER BY member_name, skill",
                    (manager,)
                )
                rows = cur.fetchall()
        for r in rows:
            r["assessed_at"] = str(r["assessed_at"])
        return {"assessments": [dict(r) for r in rows]}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# ── Manager accounts — real registration & login, mirrors the learner flow ────

class ManagerRegister(BaseModel):
    name: str
    email: str
    password: str
    team: Optional[str] = ""

class ManagerLoginRequest(BaseModel):
    email: str
    password: str

def _ensure_manager_table():
    try:
        conn = psycopg2.connect(DATABASE_URL)
        conn.autocommit = True
        with conn.cursor() as c:
            c.execute("""
                CREATE TABLE IF NOT EXISTS manager_accounts (
                    id SERIAL PRIMARY KEY,
                    name VARCHAR(120) NOT NULL,
                    email VARCHAR(160) UNIQUE NOT NULL,
                    password_hash VARCHAR(64) NOT NULL,
                    team VARCHAR(120),
                    status VARCHAR(20) DEFAULT 'pending',
                    actioned_by VARCHAR(120),
                    actioned_at TIMESTAMP,
                    created_at TIMESTAMP DEFAULT NOW()
                )
            """)
            c.execute("ALTER TABLE manager_accounts ADD COLUMN IF NOT EXISTS status VARCHAR(20) DEFAULT 'pending'")
            c.execute("ALTER TABLE manager_accounts ADD COLUMN IF NOT EXISTS actioned_by VARCHAR(120)")
            c.execute("ALTER TABLE manager_accounts ADD COLUMN IF NOT EXISTS actioned_at TIMESTAMP")
        conn.close()
    except Exception:
        pass

@app.post("/api/manager/register")
def register_manager(body: ManagerRegister):
    """Self-registration for managers. Does NOT grant access on its own — an admin
    must approve the account (mirrors the learner onboarding flow, just with the
    Admin persona as approver instead of a manager). This exists because nothing
    in a self-registration flow can otherwise confirm someone actually holds a
    people-manager role."""
    import hashlib
    _ensure_manager_table()
    email_clean = body.email.strip().lower()
    if not email_clean.endswith("@adobe.com"):
        raise HTTPException(status_code=400, detail="Please use your Adobe email address (@adobe.com).")
    pwd_hash = hashlib.sha256(body.password.encode()).hexdigest()
    try:
        with get_db() as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute("SELECT id FROM manager_accounts WHERE LOWER(email)=LOWER(%s)", (email_clean,))
                if cur.fetchone():
                    raise HTTPException(status_code=409, detail="An account with this email already exists. Try signing in instead.")
                cur.execute(
                    "INSERT INTO manager_accounts (name, email, password_hash, team, status) VALUES (%s,%s,%s,%s,'pending') RETURNING id",
                    (body.name, email_clean, pwd_hash, body.team)
                )
                row = cur.fetchone()
        return {"ok": True, "id": row["id"], "status": "pending"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/manager/login")
def login_manager(req: ManagerLoginRequest):
    """Manager sign-in. Blocks access until an admin has approved the account."""
    import hashlib
    _ensure_manager_table()
    _ensure_profile_columns()
    pwd_hash = hashlib.sha256(req.password.encode()).hexdigest()
    try:
        with get_db() as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute(
                    "SELECT id, name, email, password_hash, team, status, username, avatar_emoji, avatar_color FROM manager_accounts WHERE LOWER(email)=LOWER(%s)",
                    (req.email.strip(),)
                )
                row = cur.fetchone()
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

    if not row:
        raise HTTPException(status_code=404, detail="No manager account found for this email. Register first.")
    if row["password_hash"] != pwd_hash:
        raise HTTPException(status_code=401, detail="Incorrect password.")
    if row["status"] == "pending":
        return {"ok": False, "status": "pending", "message": "Your manager account is awaiting admin approval. Check back soon.", "name": row["name"]}
    if row["status"] == "declined":
        return {"ok": False, "status": "declined", "message": "Your manager registration was not approved. Contact your platform administrator."}

    return {
        "ok": True,
        "profile": {
            "id": row["id"],
            "name": row["name"],
            "email": row["email"],
            "team": row["team"],
            "username": row.get("username"),
            "avatar_emoji": row.get("avatar_emoji"),
            "avatar_color": row.get("avatar_color"),
        }
    }

@app.get("/api/manager/pending")
def get_pending_managers():
    """Admin reviews manager registrations awaiting approval."""
    _ensure_manager_table()
    try:
        with get_db() as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute(
                    "SELECT id, name, email, team, created_at FROM manager_accounts WHERE status='pending' ORDER BY created_at"
                )
                rows = cur.fetchall()
        for r in rows:
            r["created_at"] = str(r["created_at"])
        return {"pending": [dict(r) for r in rows]}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

class ManagerApprovalAction(BaseModel):
    action: str  # "approve" | "decline"
    actioned_by: Optional[str] = None

@app.put("/api/manager/{manager_id}/action")
def action_manager_registration(manager_id: int, body: ManagerApprovalAction):
    """Admin approves or declines a manager registration."""
    if body.action not in ("approve", "decline"):
        raise HTTPException(status_code=400, detail="action must be approve or decline")
    new_status = body.action + "d"
    _ensure_manager_table()
    try:
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "UPDATE manager_accounts SET status=%s, actioned_by=%s, actioned_at=NOW() WHERE id=%s",
                    (new_status, body.actioned_by, manager_id)
                )
        return {"ok": True, "status": new_status}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/manager/list")
def list_managers():
    """Names of all registered managers — populates the manager dropdown on the
    learner onboarding form, so new registrations stay consistent with real accounts."""
    _ensure_manager_table()
    try:
        with get_db() as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute("SELECT name FROM manager_accounts WHERE status='approved' ORDER BY name")
                rows = cur.fetchall()
        return {"managers": [r["name"] for r in rows]}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# ── Profile: avatar, username, password change ─────────────────────────────────
# Applies to both learners (onboarding_requests) and managers (manager_accounts).

AVATAR_EMOJIS = ["🦊","🐼","🚀","🎯","⚡","🌟","🔥","💡","🎨","🧠","🦉","🐙","🌈","🍀","🎮","🛰️","🧩","🦋","🐢","🦄"]
AVATAR_COLORS = ["#1473E6","#E34850","#12805C","#B86B00","#6030D0","#0891B2","#097348","#9B1C2E","#2357E8","#D6409F"]

def _ensure_profile_columns():
    try:
        conn = psycopg2.connect(DATABASE_URL)
        conn.autocommit = True
        with conn.cursor() as c:
            for table in ("onboarding_requests", "manager_accounts"):
                c.execute(f"ALTER TABLE {table} ADD COLUMN IF NOT EXISTS username VARCHAR(60)")
                c.execute(f"ALTER TABLE {table} ADD COLUMN IF NOT EXISTS avatar_emoji VARCHAR(10)")
                c.execute(f"ALTER TABLE {table} ADD COLUMN IF NOT EXISTS avatar_color VARCHAR(10)")
        conn.close()
    except Exception:
        pass

@app.get("/api/profile/avatar-options")
def get_avatar_options():
    return {"emojis": AVATAR_EMOJIS, "colors": AVATAR_COLORS}

class ProfileUpdate(BaseModel):
    email: str
    persona: str  # "learner" | "manager"
    username: Optional[str] = None
    avatar_emoji: Optional[str] = None
    avatar_color: Optional[str] = None
    preferred_name: Optional[str] = None

@app.put("/api/profile/update")
def update_profile(body: ProfileUpdate):
    _ensure_profile_columns()
    table = "manager_accounts" if body.persona == "manager" else "onboarding_requests"
    try:
        with get_db() as conn:
            with conn.cursor() as cur:
                if body.persona == "manager":
                    # manager_accounts has no preferred_name column.
                    cur.execute(
                        f"UPDATE {table} SET username=COALESCE(%s,username), avatar_emoji=COALESCE(%s,avatar_emoji), avatar_color=COALESCE(%s,avatar_color) WHERE LOWER(email)=LOWER(%s)",
                        (body.username, body.avatar_emoji, body.avatar_color, body.email.strip())
                    )
                else:
                    cur.execute(
                        f"UPDATE {table} SET username=COALESCE(%s,username), avatar_emoji=COALESCE(%s,avatar_emoji), avatar_color=COALESCE(%s,avatar_color), preferred_name=COALESCE(%s,preferred_name), profile_confirmed=TRUE WHERE LOWER(email)=LOWER(%s) AND status='approved'",
                        (body.username, body.avatar_emoji, body.avatar_color, body.preferred_name, body.email.strip())
                    )
        return {"ok": True}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

class PasswordChange(BaseModel):
    email: str
    persona: str  # "learner" | "manager"
    current_password: str
    new_password: str

@app.put("/api/profile/change-password")
def change_password(body: PasswordChange):
    import hashlib
    if len(body.new_password) < 6:
        raise HTTPException(status_code=400, detail="New password must be at least 6 characters.")
    current_hash = hashlib.sha256(body.current_password.encode()).hexdigest()
    new_hash = hashlib.sha256(body.new_password.encode()).hexdigest()
    table = "manager_accounts" if body.persona == "manager" else "onboarding_requests"
    try:
        with get_db() as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute(f"SELECT password_hash FROM {table} WHERE LOWER(email)=LOWER(%s)", (body.email.strip(),))
                row = cur.fetchone()
                if not row:
                    raise HTTPException(status_code=404, detail="Account not found.")
                if row["password_hash"] != current_hash:
                    raise HTTPException(status_code=401, detail="Current password is incorrect.")
                cur.execute(f"UPDATE {table} SET password_hash=%s WHERE LOWER(email)=LOWER(%s)", (new_hash, body.email.strip()))
        return {"ok": True}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# ── Generated content cache — flashcards, scenarios, capstone tasks ───────────
# Avoids a fresh LLM call every time someone reopens the same module's content.
# Test-out quizzes are deliberately NOT cached — regenerating each attempt
# avoids learners sharing/memorising answers across the team.

def _ensure_cache_table():
    try:
        conn = psycopg2.connect(DATABASE_URL)
        conn.autocommit = True
        with conn.cursor() as c:
            c.execute("""
                CREATE TABLE IF NOT EXISTS generated_content_cache (
                    cache_key VARCHAR(255) PRIMARY KEY,
                    agent_name VARCHAR(50),
                    content TEXT NOT NULL,
                    created_at TIMESTAMP DEFAULT NOW()
                )
            """)
        conn.close()
    except Exception:
        pass

@app.get("/api/cache/{cache_key}")
def get_cached_content(cache_key: str):
    _ensure_cache_table()
    try:
        with get_db() as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute("SELECT content, created_at FROM generated_content_cache WHERE cache_key=%s", (cache_key,))
                row = cur.fetchone()
        if not row:
            return {"hit": False}
        return {"hit": True, "content": row["content"], "created_at": str(row["created_at"])}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

class CacheWrite(BaseModel):
    agent_name: str
    content: str  # JSON-stringified by the caller

@app.post("/api/cache/{cache_key}")
def set_cached_content(cache_key: str, body: CacheWrite):
    _ensure_cache_table()
    try:
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    """INSERT INTO generated_content_cache (cache_key, agent_name, content, created_at)
                       VALUES (%s,%s,%s,NOW())
                       ON CONFLICT (cache_key) DO UPDATE SET content=EXCLUDED.content, agent_name=EXCLUDED.agent_name, created_at=NOW()""",
                    (cache_key, body.agent_name, body.content)
                )
        return {"ok": True}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/api/cache/{cache_key}")
def clear_cached_content(cache_key: str):
    """Used by 'Regenerate' buttons to force a fresh result."""
    try:
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("DELETE FROM generated_content_cache WHERE cache_key=%s", (cache_key,))
        return {"ok": True}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# ── Conversation history — full Socratic/cross-skilling transcripts ───────────
# Previously only an end-of-session summary was saved; the actual turn-by-turn
# messages were lost on refresh. This persists every message so a learner
# returning to the AI Tutor for the same module picks up where they left off.

def _ensure_conversation_table():
    try:
        conn = psycopg2.connect(DATABASE_URL)
        conn.autocommit = True
        with conn.cursor() as c:
            c.execute("""
                CREATE TABLE IF NOT EXISTS conversation_messages (
                    id SERIAL PRIMARY KEY,
                    member_name VARCHAR(120) NOT NULL,
                    manager VARCHAR(120),
                    module VARCHAR(255),
                    mode VARCHAR(30),
                    role VARCHAR(20),
                    content TEXT,
                    created_at TIMESTAMP DEFAULT NOW()
                )
            """)
            c.execute("CREATE INDEX IF NOT EXISTS idx_conv_lookup ON conversation_messages (member_name, module, mode)")
        conn.close()
    except Exception:
        pass

class ConversationMessage(BaseModel):
    member_name: str
    manager: Optional[str] = None
    module: str
    mode: str
    role: str
    content: str

@app.post("/api/conversations/message")
def save_conversation_message(body: ConversationMessage):
    _ensure_conversation_table()
    try:
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "INSERT INTO conversation_messages (member_name, manager, module, mode, role, content) VALUES (%s,%s,%s,%s,%s,%s)",
                    (body.member_name, body.manager, body.module, body.mode, body.role, body.content)
                )
        return {"ok": True}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/conversations")
def get_conversation_history(member_name: str, module: str, mode: str, limit: int = 100):
    _ensure_conversation_table()
    try:
        with get_db() as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute(
                    """SELECT role, content, created_at FROM conversation_messages
                       WHERE member_name=%s AND module=%s AND mode=%s
                       ORDER BY created_at ASC LIMIT %s""",
                    (member_name, module, mode, limit)
                )
                rows = cur.fetchall()
        for r in rows:
            r["created_at"] = str(r["created_at"])
        return {"messages": [dict(r) for r in rows]}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/api/conversations")
def clear_conversation_history(member_name: str, module: str, mode: str):
    """Used by a 'Start fresh' control if a learner wants to clear their thread."""
    try:
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "DELETE FROM conversation_messages WHERE member_name=%s AND module=%s AND mode=%s",
                    (member_name, module, mode)
                )
        return {"ok": True}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# ── Real embeddings-based retrieval ─────────────────────────────────────────────
# Upgrades /api/rag from keyword search (GitHub code search, lexical match only)
# to genuine semantic search: questions and docs are embedded into vectors, and
# retrieval is by meaning, not shared vocabulary. Uses fastembed — a local, free,
# CPU-only embedding model. No API key, no per-call cost, no pgvector dependency
# (the corpus is small enough that cosine similarity in plain Python is plenty fast).

_embedding_model = None  # lazy-loaded on first use, so server startup isn't slowed if unused

def _get_embedding_model():
    global _embedding_model
    if _embedding_model is None:
        from fastembed import TextEmbedding
        _embedding_model = TextEmbedding(model_name="BAAI/bge-small-en-v1.5")
    return _embedding_model

def _embed_texts(texts):
    model = _get_embedding_model()
    return [list(v) for v in model.embed(texts)]

def _cosine_sim(a, b):
    import math
    dot = sum(x*y for x, y in zip(a, b))
    na = math.sqrt(sum(x*x for x in a))
    nb = math.sqrt(sum(y*y for y in b))
    if na == 0 or nb == 0:
        return 0.0
    return dot / (na * nb)

def _ensure_embeddings_table():
    try:
        conn = psycopg2.connect(DATABASE_URL)
        conn.autocommit = True
        with conn.cursor() as c:
            c.execute("""
                CREATE TABLE IF NOT EXISTS doc_embeddings (
                    id SERIAL PRIMARY KEY,
                    repo VARCHAR(120),
                    file_path VARCHAR(500),
                    el_url VARCHAR(500),
                    title VARCHAR(255),
                    track VARCHAR(50),
                    chunk_index INTEGER,
                    chunk_text TEXT,
                    embedding TEXT,
                    created_at TIMESTAMP DEFAULT NOW()
                )
            """)
        conn.close()
    except Exception:
        pass

def vector_search(query: str, track: str = None, top_k: int = 3):
    """Semantic search over doc_embeddings. Returns [] if the index hasn't been
    built yet (graceful — caller falls back to keyword search).

    Prefers the persistent pgvector index (top-k ANN runs in Postgres — no
    whole-corpus load per query) when it's available; otherwise falls back to the
    legacy full-scan cosine below. Return shape is preserved either way:
    {title, repo, content, url, score}."""
    import json
    # ── Persistent pgvector path (no whole-corpus load) ───────────────────────
    try:
        from agents import vector_store as _vstore
        if _vstore.is_available():
            docs = _vstore.vector_retrieve(query, track=track, top_k=top_k, min_score=0.25)
            if docs is not None:   # None ⇒ store degraded → fall through to legacy
                return [
                    {"title":   d.get("title", ""),
                     "repo":    d.get("repo", ""),
                     "content": (d.get("content") or "")[:900],
                     "url":     d.get("url", ""),
                     "score":   d.get("score", 0.0)}
                    for d in docs
                ]
    except Exception as e:
        print(f"vector_search pgvector path skipped: {e}")

    # ── Legacy full-scan cosine fallback ──────────────────────────────────────
    _ensure_embeddings_table()
    try:
        with get_db() as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                if track:
                    cur.execute("SELECT title, repo, el_url, chunk_text, embedding FROM doc_embeddings WHERE track=%s", (track,))
                else:
                    cur.execute("SELECT title, repo, el_url, chunk_text, embedding FROM doc_embeddings")
                rows = cur.fetchall()
        if not rows:
            return []  # index not built — caller falls back to keyword search

        query_vec = _embed_texts([query])[0]
        scored = []
        for r in rows:
            try:
                doc_vec = json.loads(r["embedding"])
                score = _cosine_sim(query_vec, doc_vec)
                scored.append((score, r))
            except Exception:
                continue
        scored.sort(key=lambda x: x[0], reverse=True)
        return [
            {"title": r["title"], "repo": r["repo"], "content": r["chunk_text"][:900], "url": r["el_url"], "score": round(s, 3)}
            for s, r in scored[:top_k] if s > 0.25  # discard weak matches rather than return noise
        ]
    except Exception as e:
        print(f"vector_search error: {e}")
        return []

@app.get("/api/embeddings/status")
def embeddings_status():
    """How much of the corpus is actually indexed — run build_embeddings_index.py to populate."""
    _ensure_embeddings_table()
    try:
        with get_db() as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute("SELECT track, COUNT(*) as chunks, COUNT(DISTINCT title) as topics FROM doc_embeddings GROUP BY track")
                rows = cur.fetchall()
        return {"indexed": [dict(r) for r in rows], "built": len(rows) > 0}
    except Exception as e:
        return {"indexed": [], "built": False, "error": str(e)}

# ── Certifications ─────────────────────────────────────────────────────────────

@app.post("/api/admin/certs/import")
async def admin_certs_import(file: UploadFile = File(...), user: dict = Depends(require_persona("admin"))):
    """Admin uploads raw .xlsx file (ALM export format); backend parses it with openpyxl.
    Handles ALM exports where rows 1-N are metadata and the real header row contains
    'Learner Email' somewhere in the sheet."""
    import openpyxl, io
    contents = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
        ws = wb.active

        # ── Find the real header row ──────────────────────────────────────────
        # ALM exports have metadata in the first few rows.
        # Scan until we find a row containing an email-like header.
        EMAIL_SIGNALS = {"learner email", "email", "email address", "user email"}
        header_row_idx = None
        headers = []
        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            row_vals = [str(v or "").strip().lower() for v in row]
            if any(v in EMAIL_SIGNALS for v in row_vals):
                header_row_idx = i
                headers = row_vals
                break

        if header_row_idx is None:
            raise ValueError("Could not find a header row with an email column. "
                             "Expected a column named 'Learner Email' or 'Email'.")

        def col(row, *keys):
            for k in keys:
                if k in headers:
                    v = row[headers.index(k)]
                    if v is not None:
                        s = str(v).strip()
                        if s and s.lower() not in ("none", "n/a", "-", "—"):
                            return s
            return ""

        rows = []
        for r in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
            if not any(r):
                continue
            first_name = col(r, "learner first name", "first name", "firstname")
            last_name  = col(r, "learner last name",  "last name",  "lastname")
            full_name  = col(r, "full name", "name", "learner name", "employee name")
            if not full_name and (first_name or last_name):
                full_name = f"{first_name} {last_name}".strip()

            rows.append({
                "email":       col(r, "learner email", "email", "email address", "user email"),
                "full_name":   full_name,
                "cert_name":   col(r, "certification name", "cert name", "certification", "credential name"),
                "cert_type":   col(r, "credential capability", "cert type", "certification role combo",
                                    "credential type", "level", "type"),
                "status":      col(r, "cert status", "status", "certification status"),
                "issued_date": col(r, "certification date", "issued date", "issue date",
                                    "date issued", "issued", "cert date"),
                "expiry_date": col(r, "date_renewal", "date renewal", "renewal date", "expiry date",
                                    "expiry", "expiration date", "expiration", "expires"),
            })
    except Exception as ex:
        raise HTTPException(400, f"Could not parse Excel file: {ex}")

    rows = [r for r in rows if r["email"] and r["cert_name"]]
    if not rows:
        raise HTTPException(400, "No data rows found — check that the file has 'Learner Email' "
                                 "and 'Certification Name' columns.")
    from datetime import date, datetime as _dt

    def parse_date(val):
        if not val:
            return None
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y", "%d-%m-%Y"):
            try:
                return _dt.strptime(str(val).strip(), fmt).date()
            except ValueError:
                continue
        return None

    def compute_status(expiry, supplied):
        if expiry:
            delta = (expiry - date.today()).days
            if delta < 0:
                return "Expired", delta
            if delta <= 90:
                return "Renew Soon", delta
            return "Active", delta
        return (supplied or "Active"), None

    inserted = updated = skipped = 0
    conn = psycopg2.connect(DATABASE_URL)
    try:
        with conn:
            c = conn.cursor()
            for row in rows:
                email = (row.get("email") or "").strip().lower()
                cert_name = (row.get("cert_name") or "").strip()
                if not email or not cert_name:
                    skipped += 1
                    continue
                expiry = parse_date(row.get("expiry_date"))
                issued = parse_date(row.get("issued_date"))
                status, days = compute_status(expiry, row.get("status"))
                c.execute("SELECT id, name FROM onboarding_requests WHERE LOWER(email)=%s LIMIT 1", (email,))
                user_row = c.fetchone()
                user_id = user_row[0] if user_row else None
                full_name = (row.get("full_name") or (user_row[1] if user_row else "")).strip()
                c.execute("""
                    INSERT INTO user_certifications
                        (user_id, email, full_name, cert_name, cert_type, status,
                         issued_date, expiry_date, days_remaining, updated_at)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,NOW())
                    ON CONFLICT (email, cert_name) DO UPDATE SET
                        cert_type=EXCLUDED.cert_type, status=EXCLUDED.status,
                        issued_date=EXCLUDED.issued_date, expiry_date=EXCLUDED.expiry_date,
                        days_remaining=EXCLUDED.days_remaining, full_name=EXCLUDED.full_name,
                        updated_at=NOW()
                    RETURNING (xmax = 0) AS is_insert
                """, (user_id, email, full_name, cert_name,
                      row.get("cert_type", ""), status, issued, expiry, days))
                res = c.fetchone()
                if res and res[0]:
                    inserted += 1
                else:
                    updated += 1
    finally:
        conn.close()
    return {"ok": True, "inserted": inserted, "updated": updated, "skipped": skipped, "total": len(rows)}


@app.get("/api/admin/certs/summary")
def certs_summary(user: dict = Depends(require_persona("admin"))):
    """Persistent snapshot of certifications currently on file — shown every
    time the Certification Import page opens, not just right after an upload."""
    conn = psycopg2.connect(DATABASE_URL)
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as c:
            c.execute("SELECT COUNT(*) AS n, COUNT(DISTINCT LOWER(email)) AS members FROM user_certifications")
            totals = c.fetchone()
            c.execute("""SELECT status, COUNT(*) AS n FROM user_certifications
                         GROUP BY status ORDER BY n DESC""")
            by_status = [dict(r) for r in c.fetchall()]
    finally:
        conn.close()
    return {"total_certs": totals["n"], "total_members": totals["members"], "by_status": by_status}


@app.delete("/api/admin/certs/wipe")
def wipe_certs(confirm: str = "", user: dict = Depends(require_persona("admin"))):
    """Hard reset — permanently deletes every certification record. Irreversible;
    requires ?confirm=WIPE."""
    if confirm != "WIPE":
        raise HTTPException(400, 'Pass ?confirm=WIPE to actually delete everything.')
    conn = psycopg2.connect(DATABASE_URL)
    try:
        with conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as c:
                c.execute("SELECT COUNT(*) AS n FROM user_certifications")
                n = c.fetchone()["n"]
                c.execute("DELETE FROM user_certifications")
    finally:
        conn.close()
    return {"ok": True, "deleted_certs": n}


@app.get("/api/certs/my")
def my_certs(user_id: int = None, email: str = None):
    """Return all certifications for a single user."""
    if not user_id and not email:
        raise HTTPException(400, "Provide user_id or email")
    conn = psycopg2.connect(DATABASE_URL)
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as c:
            if user_id:
                c.execute("""SELECT cert_name, cert_type, status,
                                    TO_CHAR(issued_date,'YYYY-MM-DD') AS issued_date,
                                    TO_CHAR(expiry_date,'YYYY-MM-DD') AS expiry_date,
                                    days_remaining
                             FROM user_certifications WHERE user_id=%s
                             ORDER BY expiry_date NULLS LAST""", (user_id,))
            else:
                c.execute("""SELECT cert_name, cert_type, status,
                                    TO_CHAR(issued_date,'YYYY-MM-DD') AS issued_date,
                                    TO_CHAR(expiry_date,'YYYY-MM-DD') AS expiry_date,
                                    days_remaining
                             FROM user_certifications WHERE LOWER(email)=%s
                             ORDER BY expiry_date NULLS LAST""", (email.strip().lower(),))
            rows = c.fetchall()
    finally:
        conn.close()
    return {"certs": [dict(r) for r in rows]}


@app.get("/api/certs/team")
def team_certs(manager_email: str = None, manager_name: str = None):
    """Return certifications for every team member, grouped by person.
    If manager_email/name provided, pulls team from employee_directory.
    Otherwise falls back to all approved learners in onboarding_requests."""
    from collections import defaultdict
    conn = psycopg2.connect(DATABASE_URL)
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as c:
            if manager_email or manager_name:
                # Pull team from directory
                clauses, params = [], []
                if manager_email:
                    clauses.append("LOWER(manager_email)=LOWER(%s)")
                    params.append(manager_email.strip())
                if manager_name:
                    clauses.append("LOWER(manager_name)=LOWER(%s)")
                    params.append(manager_name.strip())
                where = "(" + " OR ".join(clauses) + ") AND is_active=TRUE"
                c.execute(f"""SELECT email,
                                     CONCAT(COALESCE(first_name,''),' ',COALESCE(last_name,'')) AS full_name
                              FROM employee_directory WHERE {where}
                              ORDER BY last_name, first_name""", params)
            else:
                c.execute("""SELECT email, name AS full_name FROM onboarding_requests
                             WHERE role NOT IN ('manager','admin') AND status='approved'
                             ORDER BY name""")
            members = c.fetchall()
            if not members:
                return {"team": []}
            emails = [m["email"].lower() for m in members]
            c.execute("""SELECT LOWER(email) AS email, cert_name, cert_type, status,
                                TO_CHAR(issued_date,'YYYY-MM-DD') AS issued_date,
                                TO_CHAR(expiry_date,'YYYY-MM-DD') AS expiry_date,
                                days_remaining
                         FROM user_certifications WHERE LOWER(email) = ANY(%s)
                         ORDER BY email, expiry_date NULLS LAST""", (emails,))
            cert_rows = c.fetchall()
    finally:
        conn.close()

    cert_map = defaultdict(list)
    for r in cert_rows:
        cert_map[r["email"]].append({
            "cert_name":      r["cert_name"],
            "cert_type":      r["cert_type"],
            "status":         r["status"],
            "issued_date":    r["issued_date"],
            "expiry_date":    r["expiry_date"],
            "days_remaining": r["days_remaining"],
        })

    return {"team": [
        {"user_id":   m["id"],
         "email":     m["email"],
         "full_name": m["full_name"] or m["email"],
         "certs":     cert_map.get(m["email"].lower(), [])}
        for m in members
    ]}


# ── Project Management ─────────────────────────────────────────────────────────

@app.get("/api/projects")
def list_projects(manager_email: str = None, manager_name: str = None):
    if not manager_email and not manager_name:
        raise HTTPException(400, "Provide manager_email or manager_name")
    conn = psycopg2.connect(DATABASE_URL)
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as c:
            if manager_email:
                c.execute("SELECT * FROM projects WHERE LOWER(manager_email)=LOWER(%s) ORDER BY created_at DESC", (manager_email.strip(),))
            else:
                c.execute("SELECT * FROM projects WHERE LOWER(manager_email)=LOWER(%s) ORDER BY created_at DESC", (manager_name.strip(),))
            projects = c.fetchall()
            result = []
            for p in projects:
                d = dict(p)
                for k in ("created_at","updated_at"):
                    if d.get(k): d[k] = str(d[k])
                # members
                c.execute("SELECT member_email, member_name FROM project_members WHERE project_id=%s ORDER BY added_at", (d["id"],))
                d["members"] = [dict(r) for r in c.fetchall()]
                # issues
                c.execute("SELECT id, title, priority, status, visibility, TO_CHAR(created_at,'YYYY-MM-DD') AS created_at FROM project_issues WHERE project_id=%s ORDER BY created_at DESC", (d["id"],))
                d["issues"] = [dict(r) for r in c.fetchall()]
                result.append(d)
    finally:
        conn.close()
    return {"projects": result}


@app.post("/api/projects")
def create_project(body: dict = Body(...)):
    manager_email = (body.get("manager_email") or "").strip().lower()
    title = (body.get("title") or "").strip()
    if not manager_email or not title:
        raise HTTPException(400, "manager_email and title required")
    conn = psycopg2.connect(DATABASE_URL)
    try:
        with conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as c:
                c.execute("""INSERT INTO projects (manager_email,title,sector,tag,sprint,status,description,color)
                             VALUES (%s,%s,%s,%s,%s,%s,%s,%s) RETURNING *""",
                    (manager_email, title,
                     body.get("sector",""), body.get("tag",""),
                     body.get("sprint",""), body.get("status","Planning"),
                     body.get("description",""), body.get("color","#1473E6")))
                row = dict(c.fetchone())
                for k in ("created_at","updated_at"):
                    if row.get(k): row[k] = str(row[k])
                row["members"] = []
                row["issues"] = []
    finally:
        conn.close()
    return {"ok": True, "project": row}


@app.put("/api/projects/{project_id}")
def update_project(project_id: int, body: dict = Body(...),
                    user: dict = Depends(require_persona("manager", "admin"))):
    fields, params = [], []
    for col in ("title","sector","tag","sprint","status","description","color"):
        if col in body:
            fields.append(f"{col}=%s")
            params.append(body[col])
    if not fields:
        raise HTTPException(400, "No fields to update")
    params.append(project_id)
    conn = psycopg2.connect(DATABASE_URL)
    try:
        with conn:
            with conn.cursor() as c:
                c.execute(f"UPDATE projects SET {','.join(fields)}, updated_at=NOW() WHERE id=%s", params)
    finally:
        conn.close()
    return {"ok": True}


@app.delete("/api/projects/{project_id}")
def delete_project(project_id: int, user: dict = Depends(require_persona("manager", "admin"))):
    conn = psycopg2.connect(DATABASE_URL)
    try:
        with conn:
            with conn.cursor() as c:
                c.execute("DELETE FROM projects WHERE id=%s", (project_id,))
    finally:
        conn.close()
    return {"ok": True}


@app.post("/api/projects/{project_id}/members")
def add_project_member(project_id: int, body: dict = Body(...)):
    email = (body.get("member_email") or "").strip().lower()
    name  = (body.get("member_name")  or "").strip()
    if not email:
        raise HTTPException(400, "member_email required")
    conn = psycopg2.connect(DATABASE_URL)
    try:
        with conn:
            with conn.cursor() as c:
                c.execute("""INSERT INTO project_members (project_id, member_email, member_name)
                             VALUES (%s,%s,%s) ON CONFLICT (project_id, member_email) DO NOTHING""",
                          (project_id, email, name))
    finally:
        conn.close()
    return {"ok": True}


@app.delete("/api/projects/{project_id}/members/{member_email}")
def remove_project_member(project_id: int, member_email: str):
    conn = psycopg2.connect(DATABASE_URL)
    try:
        with conn:
            with conn.cursor() as c:
                c.execute("DELETE FROM project_members WHERE project_id=%s AND LOWER(member_email)=LOWER(%s)",
                          (project_id, member_email))
    finally:
        conn.close()
    return {"ok": True}


@app.post("/api/projects/{project_id}/issues")
def add_issue(project_id: int, body: dict = Body(...)):
    title = (body.get("title") or "").strip()
    if not title:
        raise HTTPException(400, "title required")
    conn = psycopg2.connect(DATABASE_URL)
    try:
        with conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as c:
                c.execute("""INSERT INTO project_issues (project_id,title,priority,status,visibility)
                             VALUES (%s,%s,%s,%s,%s) RETURNING id,title,priority,status,visibility,
                             TO_CHAR(created_at,'YYYY-MM-DD') AS created_at""",
                          (project_id, title,
                           body.get("priority","Medium"),
                           body.get("status","Open"),
                           body.get("visibility","Everyone")))
                row = dict(c.fetchone())
    finally:
        conn.close()
    return {"ok": True, "issue": row}


@app.put("/api/projects/{project_id}/issues/{issue_id}")
def update_issue(project_id: int, issue_id: int, body: dict = Body(...)):
    fields, params = [], []
    for col in ("title","priority","status","visibility"):
        if col in body:
            fields.append(f"{col}=%s")
            params.append(body[col])
    if not fields:
        raise HTTPException(400, "No fields")
    params += [issue_id, project_id]
    conn = psycopg2.connect(DATABASE_URL)
    try:
        with conn:
            with conn.cursor() as c:
                c.execute(f"UPDATE project_issues SET {','.join(fields)},updated_at=NOW() WHERE id=%s AND project_id=%s", params)
    finally:
        conn.close()
    return {"ok": True}


@app.delete("/api/projects/{project_id}/issues/{issue_id}")
def delete_issue(project_id: int, issue_id: int):
    conn = psycopg2.connect(DATABASE_URL)
    try:
        with conn:
            with conn.cursor() as c:
                c.execute("DELETE FROM project_issues WHERE id=%s AND project_id=%s", (issue_id, project_id))
    finally:
        conn.close()
    return {"ok": True}


# ── Weekly Tracker Excel Import ────────────────────────────────────────────────

# Sheet names that are system/report tabs — skip these when parsing
TRACKER_SYSTEM_SHEETS = {
    "how to use", "manager dashboard", "all projects", "all members",
    "all milestones", "dashboard", "summary", "instructions", "readme",
    "sheet1", "cover", "index", "lookup", "template", "example",
    "sample", "guide", "help", "ref", "reference", "notes",
}

def _parse_tracker_excel(raw: bytes, manager_email: str):
    """Parse team project tracker Excel. Supports two layouts:
    1. Flat layout — a single sheet with a 'Member'/'Team Member'/'Resource' column;
       each row belongs to whichever member that row's cell names.
    2. Per-member layout (legacy) — each non-system sheet tab = one team member,
       with the member's name taken from the tab name or a 'Name:' meta cell.
    Finds the project data table by locating the header row containing 'Project ID' or 'Project Name'.
    Returns list of {member_name, member_role, projects:[...]} dicts. Initiatives and
    milestones are managed separately (see member_initiatives/member_milestones) and
    are not parsed from the Excel.
    """
    import openpyxl, io, re as _re
    from datetime import datetime, date

    def to_str(v):
        if v is None: return ""
        return str(v).strip()

    def to_date(v):
        if v is None: return None
        if isinstance(v, (datetime, date)):
            return v.date() if isinstance(v, datetime) else v
        try:
            s = str(v).strip()
            for fmt in ("%Y-%m-%d","%d-%m-%Y","%m/%d/%Y","%d/%m/%Y","%Y/%m/%d"):
                try: return datetime.strptime(s, fmt).date()
                except: continue
        except: pass
        return None

    def to_float(v):
        try: return float(str(v).strip())
        except: return 0.0

    def to_int(v):
        try: return int(float(str(v).strip()))
        except: return 0

    def norm_header(v):
        """Lowercase + collapse whitespace + tighten spacing around '/' so header
        variants like 'Hrs / Week' / 'Hrs/Week' and 'Industry / Vertical' /
        'Industry/Vertical' land on the same COL_MAP key."""
        h = to_str(v).lower()
        h = _re.sub(r"\s*/\s*", "/", h)
        h = _re.sub(r"\s+", " ", h).strip()
        return h

    try:
        wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True, keep_vba=False)
    except Exception:
        # Fallback for some xlsm files that need keep_vba=True
        wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True, keep_vba=True)
    print(f"[tracker parser] sheets: {wb.sheetnames}")
    results = []

    # Header keywords that identify the project table row / the per-row member column
    PROJ_HEADER_SIGNALS = {"project id", "project name", "project_id", "project_name"}
    MEMBER_HEADER_SIGNALS = {"member", "member name", "team member", "resource",
                              "resource name", "assigned to", "owner"}
    # Column name (normalised) → DB field mapping
    COL_MAP = {
        "project id": "project_code", "project_name": "title", "project name": "title",
        "project type": "project_type", "industry": "industry", "industry/vertical": "industry",
        "phase": "phase", "stage": "stage",
        "start date": "start_date", "end date": "end_date",
        "days remaining": "days_remaining",
        "hrs/weekallocated": "hrs_per_week",
        "hrs/week": "hrs_per_week", "hrs per week": "hrs_per_week",
        "allocated hours": "hrs_per_week",
        "solutions used": "solutions_used", "use cases": "use_cases",
        "product features": "product_features",
        "data sources": "data_sources",
        "destinations": "destinations",
        "audiences": "num_audiences",
        "product issues (ticket id)": "ticket_ids", "product issues": "ticket_ids",
        "ticket id": "ticket_ids", "ticket ids": "ticket_ids",
        "health/status": "health_status", "health status": "health_status",
        "health": "health_status", "status": "health_status",
        "renewal?": "renewal", "renewal": "renewal",
        "weekly comments": "weekly_comments", "comments": "weekly_comments",
        "high level project notes": "high_level_notes",
        "region": "region",
        "sector": "sector",
        "tag": "tag",
    }

    for sheet_name in wb.sheetnames:
        sn_lower = sheet_name.strip().lower()
        # Normalise: remove special chars for comparison
        sn_norm = _re.sub(r'[^a-z0-9 ]', '', sn_lower).strip()

        ws = wb[sheet_name]
        member_name_meta = sheet_name.strip()
        member_role_meta = ""

        # Scan for meta rows (Name / Role) in first 8 rows — used only for the
        # legacy per-sheet-is-one-member layout.
        for row in ws.iter_rows(min_row=1, max_row=8, values_only=True):
            row_strs = [to_str(c).lower() for c in row]
            for i, cell in enumerate(row_strs):
                if cell in ("name", "your name") and i+1 < len(row) and row[i+1]:
                    member_name_meta = to_str(row[i+1]) or member_name_meta
                if cell in ("role", "your role") and i+1 < len(row) and row[i+1]:
                    member_role_meta = to_str(row[i+1])

        # Find project table header row, and — critically — check whether it has
        # an explicit per-row Member/Resource column (the "flat" tracker layout).
        header_row_idx = None
        headers = []
        member_col = None
        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            row_norm = [norm_header(c) for c in row]
            if any(sig in row_norm for sig in PROJ_HEADER_SIGNALS):
                header_row_idx = i
                headers = row_norm
                for ci, h in enumerate(headers):
                    if h in MEMBER_HEADER_SIGNALS:
                        member_col = ci
                        break
                break

        # A sheet with a genuine "Project ID"/"Project Name" header is real data
        # regardless of what the tab is named — only apply the system-sheet-name
        # skip when we found no such header at all (or found one with no data,
        # in which case there's nothing to lose by skipping it).
        if header_row_idx is None:
            continue
        is_system_name = (sn_lower in TRACKER_SYSTEM_SHEETS or sn_norm in TRACKER_SYSTEM_SHEETS or
            any(sn_norm.startswith(s) for s in ["instruction", "how to", "dashboard",
                "all member", "all milestone", "template", "readme",
                "cover", "index", "guide", "column guide", "rules"]))
        if is_system_name and member_col is None:
            continue

        def _parse_project_row(row, headers=headers):
            rec = {}
            for ci, h in enumerate(headers):
                if h in COL_MAP and ci < len(row):
                    field = COL_MAP[h]
                    val = row[ci]
                    if field in ("start_date", "end_date"):
                        rec[field] = to_date(val)
                    elif field == "hrs_per_week":
                        rec[field] = to_float(val)
                    elif field in ("num_audiences", "days_remaining"):
                        rec[field] = to_int(val)
                    else:
                        rec[field] = to_str(val)
            return rec

        def _is_valid_project_row(rec):
            if not rec.get("title","").strip() and not rec.get("project_code","").strip():
                return False
            _t = (rec.get("title","") or rec.get("project_code","")).strip().lower()
            if any(sig in _t for sig in [
                "initiative", "notes & key", "notes and key", "key milestone",
                "weekly update", "team-level track", "log updates",
                "note / milestone", "blockers", "milestone / update",
                "project name", "project type", "project id", "industry / vertical",
                "hrs / week", "hrs/week", "solutions used",
                "health /status", "health/status", "renewal?",
                "log updates, blockers", "important dates",
                "dj nexus", "enablement dashboard",  # known initiatives — skip from projects
            ]):
                return False
            import unicodedata as _ud
            _t_no_emoji = "".join(c for c in _t if _ud.category(c) not in ("So","Sm","Sk","Mn")).strip()
            if len(_t_no_emoji.replace(" ","")) < 3:
                return False
            if _t and ord(_t[0]) > 9000:
                return False
            return True

        if member_col is not None:
            # ── Flat layout: group rows by their own Member cell ────────────
            by_member = {}
            for row in ws.iter_rows(min_row=header_row_idx+1, values_only=True):
                if not any(c for c in row if c is not None and str(c).strip()):
                    continue
                row_member = to_str(row[member_col]) if member_col < len(row) else ""
                if not row_member:
                    continue
                rec = _parse_project_row(row)
                if not _is_valid_project_row(rec):
                    continue
                if not rec.get("title",""):
                    rec["title"] = rec.get("project_code","Unnamed project")
                by_member.setdefault(row_member, []).append(rec)

            for row_member, projs in by_member.items():
                if projs:
                    results.append({
                        "member_name": row_member,
                        "member_role": "",
                        "email": "",
                        "projects": projs,
                    })
            continue

        # ── Legacy layout: whole sheet = one member ─────────────────────────
        member_name = member_name_meta
        member_role = member_role_meta
        projects = []
        for row in ws.iter_rows(min_row=header_row_idx+1, values_only=True):
            if not any(c for c in row if c is not None and str(c).strip()):
                continue
            rec = _parse_project_row(row)
            if not _is_valid_project_row(rec):
                continue
            if not rec.get("title",""):
                rec["title"] = rec.get("project_code","Unnamed project")
            projects.append(rec)

        if projects:
            results.append({
                "member_name": member_name,
                "member_role": member_role,
                "email": "",
                "projects": projects,
            })

    return results


@app.post("/api/admin/tracker/import")
async def import_tracker(file: UploadFile = File(...), manager_email: str = "", mgr_email: str = Form(default=""),
                          user: dict = Depends(require_persona("admin"))):
    """Admin uploads the team Excel tracker. Each member's projects are
    attributed to THEIR OWN manager automatically, resolved by full name
    against the HR roster (employee_directory.manager_email/manager_name) —
    no manual manager-email entry needed. `manager_email` is only used as a
    fallback for members not found in the HR roster (e.g. no roster uploaded
    yet); those land under "unassigned" if it's also blank."""
    manager_email = (manager_email or mgr_email or "").strip()
    print(f"[tracker/import] fallback manager_email={repr(manager_email)} file={file.filename}")
    import io
    contents = await file.read()
    import traceback
    try:
        parsed = _parse_tracker_excel(contents, manager_email)
    except Exception as ex:
        tb = traceback.format_exc()
        print(f"[tracker/import] ERROR: {tb}")
        raise HTTPException(400, f"Could not parse tracker file: {ex}")

    if not parsed:
        raise HTTPException(400, "No member sheets found. Check that member tabs exist and contain a 'Project Name' or 'Project ID' column.")

    from datetime import date as _date

    stats = {"members_processed": 0, "projects_inserted": 0, "projects_updated": 0,
             "members_linked": 0, "skipped": 0, "errors": []}

    conn = psycopg2.connect(DATABASE_URL)
    conn.autocommit = True  # each statement is its own tx — one failure won't abort others
    try:
        c = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

        for member_data in parsed:
                member_name = member_data["member_name"]
                member_role = member_data["member_role"]

                # Resolve a real email by name — try the HR roster first, then
                # any registered app profile (covers demo/manual signups when
                # no HR roster has been uploaded yet). The HR roster row also
                # gives us who this person's manager actually is, so each
                # member's projects get attributed to their real manager
                # automatically — no manual "manager email" entry needed.
                c.execute("""SELECT email, manager_email, manager_name FROM employee_directory
                             WHERE LOWER(CONCAT(first_name,' ',last_name))=LOWER(%s)
                             OR LOWER(first_name)=LOWER(%s) LIMIT 1""",
                          (member_name, member_name.split()[0] if member_name else ""))
                dir_row = c.fetchone()
                member_email = dir_row["email"] if dir_row else ""
                if not member_email:
                    c.execute("""SELECT email FROM onboarding_requests
                                 WHERE LOWER(name)=LOWER(%s) OR LOWER(preferred_name)=LOWER(%s)
                                 LIMIT 1""",
                              (member_name, member_name))
                    app_row = c.fetchone()
                    member_email = app_row["email"] if app_row else ""

                # This member's own manager — from the HR roster if we found
                # them there, else fall back to whatever the caller passed
                # (kept for backward compatibility), else "unassigned".
                row_manager_email = (
                    (dir_row and (dir_row.get("manager_email") or dir_row.get("manager_name")))
                    or manager_email or "unassigned"
                )

                stats["members_processed"] += 1

                # Clear existing allocations for this member to prevent duplicates on re-import
                c.execute("DELETE FROM project_allocations WHERE LOWER(member_name)=LOWER(%s) AND LOWER(manager)=LOWER(%s)",
                          (member_name, row_manager_email))

                for proj in member_data["projects"]:
                  def _t(v, n=250): return str(v or "")[:n]
                  try:
                    title    = proj.get("title","").strip()
                    p_code   = proj.get("project_code","").strip()
                    sector   = proj.get("sector","").strip()
                    tag      = proj.get("tag","").strip() or proj.get("solutions_used","").split(",")[0].strip()
                    stage    = proj.get("stage","").strip()
                    phase    = proj.get("phase","").strip()
                    status_raw = proj.get("health_status","").strip()
                    # Map health status to our status vocabulary
                    status = "In Progress"
                    if status_raw.lower() in ("blocked","on hold"): status = "Blocked"
                    elif status_raw.lower() in ("completed","done","near completion"): status = "Completed"
                    elif status_raw.lower() in ("planning","discovery"): status = "Planning"
                    hrs = proj.get("hrs_per_week", 0) or 0
                    days_remaining = proj.get("days_remaining")
                    if not days_remaining and proj.get("end_date"):
                        days_remaining = (proj["end_date"] - _date.today()).days
                    num_audiences = proj.get("num_audiences", 0) or 0

                    # Upsert project (match on project_code + manager_email, or title)
                    lookup_col = "project_code" if p_code else "title"
                    lookup_val = p_code if p_code else title
                    c.execute(f"SELECT id FROM projects WHERE LOWER({lookup_col})=LOWER(%s) AND LOWER(manager_email)=LOWER(%s)",
                              (lookup_val, row_manager_email))
                    existing = c.fetchone()

                    if existing:
                        proj_id = existing["id"]
                        c.execute("""UPDATE projects SET title=%s,project_code=%s,sector=%s,tag=%s,
                                     stage=%s,phase=%s,status=%s,
                                     project_type=%s,industry=%s,solutions_used=%s,
                                     health_status=%s,weekly_comments=%s,high_level_notes=%s,
                                     renewal=%s,region=%s,use_cases=%s,
                                     product_features=%s,data_sources=%s,destinations=%s,
                                     num_audiences=%s,ticket_ids=%s,days_remaining=%s,
                                     start_date=%s,end_date=%s,imported_from_tracker=TRUE,updated_at=NOW()
                                     WHERE id=%s""",
                            (_t(title),_t(p_code),_t(sector),_t(tag),_t(stage),_t(phase),_t(status),
                             _t(proj.get("project_type","")),_t(proj.get("industry","")),
                             _t(proj.get("solutions_used","")),_t(status_raw,50),
                             _t(proj.get("weekly_comments",""),2000),_t(proj.get("high_level_notes",""),2000),
                             _t(proj.get("renewal",""),20),_t(proj.get("region",""),50),_t(proj.get("use_cases",""),1000),
                             _t(proj.get("product_features",""),255),_t(proj.get("data_sources",""),255),
                             _t(proj.get("destinations",""),255),int(num_audiences),_t(proj.get("ticket_ids",""),255),
                             days_remaining,
                             proj.get("start_date"),proj.get("end_date"),proj_id))
                        stats["projects_updated"] += 1
                    else:
                        c.execute("""INSERT INTO projects (manager_email,title,project_code,sector,tag,
                                     stage,phase,status,project_type,industry,solutions_used,
                                     health_status,weekly_comments,high_level_notes,renewal,region,use_cases,
                                     product_features,data_sources,destinations,num_audiences,ticket_ids,days_remaining,
                                     start_date,end_date,color,imported_from_tracker)
                                     VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,TRUE)
                                     RETURNING id""",
                            (_t(row_manager_email),_t(title),_t(p_code),_t(sector),_t(tag),_t(stage),_t(phase),_t(status),
                             _t(proj.get("project_type","")),_t(proj.get("industry","")),
                             _t(proj.get("solutions_used","")),_t(status_raw,50),
                             _t(proj.get("weekly_comments",""),2000),_t(proj.get("high_level_notes",""),2000),
                             _t(proj.get("renewal",""),20),_t(proj.get("region",""),50),_t(proj.get("use_cases",""),1000),
                             _t(proj.get("product_features",""),255),_t(proj.get("data_sources",""),255),
                             _t(proj.get("destinations",""),255),int(num_audiences),_t(proj.get("ticket_ids",""),255),
                             days_remaining,
                             proj.get("start_date"),proj.get("end_date"),"#1473E6"))
                        proj_id = c.fetchone()["id"]
                        stats["projects_inserted"] += 1

                    # Link member to project
                    if member_email or member_name:
                        c.execute("""INSERT INTO project_members (project_id,member_email,member_name,hrs_per_week,role_on_project)
                                     VALUES (%s,%s,%s,%s,%s)
                                     ON CONFLICT (project_id,member_email) DO UPDATE SET
                                     hrs_per_week=EXCLUDED.hrs_per_week,
                                     role_on_project=EXCLUDED.role_on_project""",
                            (proj_id, member_email or member_name.lower().replace(" ",".")+"@adobe.com",
                             member_name, float(hrs), member_role))
                        stats["members_linked"] += 1

                    # Also write to project_allocations (Weekly Tracker source)
                    # so imported history appears in the tracker immediately
                    c.execute("""
                        INSERT INTO project_allocations
                            (member_name, manager, project_id, project_name, project_type,
                             industry, phase, stage, start_date, end_date, hrs_per_week,
                             use_cases, solutions_used, health_status, renewal,
                             comments, project_notes, region)
                        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
                        (_t(member_name), _t(row_manager_email),
                         _t(p_code or str(proj_id)), _t(title),
                         _t(proj.get("project_type","")),
                         _t(proj.get("industry","")),
                         _t(proj.get("phase","")),
                         _t(proj.get("stage","")),
                         proj.get("start_date"),
                         proj.get("end_date"),
                         float(hrs),
                         _t(proj.get("use_cases",""), 1000),
                         _t(proj.get("solutions_used","")),
                         _t(proj.get("health_status","On track") or "On track", 50),
                         _t(proj.get("renewal","TBD") or "TBD", 20),
                         _t(proj.get("weekly_comments",""), 2000),
                         _t(proj.get("high_level_notes",""), 2000),
                         _t(proj.get("region",""), 50)
                    ))
                  except Exception as proj_err:
                    stats["skipped"] += 1
                    stats["errors"].append(f"{member_name}/{proj.get('title','?')}: {proj_err}")
                    print(f"[tracker/import] row error: {proj_err}")
    finally:
        conn.close()

    return {"ok": True, **stats, "message": f"Processed {stats['members_processed']} member sheets, {stats['projects_inserted']} projects created, {stats['projects_updated']} updated.", "errors": stats["errors"][:5]}


@app.post("/api/admin/tracker/preview")
async def preview_tracker(file: UploadFile = File(...)):
    """Preview what would be imported before committing."""
    import traceback
    contents = await file.read()
    if not contents:
        raise HTTPException(400, "File is empty.")
    print(f"[tracker/preview] file={file.filename} size={len(contents)} bytes")
    try:
        parsed = _parse_tracker_excel(contents, "")
    except Exception as ex:
        tb = traceback.format_exc()
        print(f"[tracker/preview] ERROR:\n{tb}")
        raise HTTPException(400, f"Could not parse file: {ex}")
    print(f"[tracker/preview] parsed {len(parsed)} member sheets")
    if not parsed:
        raise HTTPException(400, "No member sheets found. Check that member tabs contain a 'Project Name' or 'Project ID' column.")
    summary = [{"member": m["member_name"], "role": m["member_role"],
                "project_count": len(m["projects"]),
                "projects": [p.get("title","") or p.get("project_code","") for p in m["projects"]]}
               for m in parsed]
    return {"ok": True, "members": len(parsed), "summary": summary}


# ── Learning tracks (admin-configurable — Reasoning agent reads these) ─────────

@app.get("/api/tracks")
def list_tracks_public():
    """Public list of active learning tracks (code + label) for track pickers
    in Study Cards / Capstone / Practice Scenarios. Deliberately unauthenticated
    — unlike /api/admin/tracks, demo personas (no IMS session) use these
    features too and must still be able to see track options."""
    try:
        from agents.reasoning import _load_track_config
        cfg = _load_track_config()
        return {"tracks": [{"track_code": c, "label": cfg["labels"][c]} for c in cfg["track_codes"]]}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


class TrackUpsert(BaseModel):
    track_code: str
    label: str
    keywords: List[str] = []
    grounding_terms: List[str] = []
    active: bool = True
    sort_order: int = 100


@app.get("/api/admin/tracks")
def list_learning_tracks(user: dict = Depends(get_current_user)):
    """List all configured learning tracks (active and inactive)."""
    try:
        with get_db() as conn:
            from agents.reasoning import _ensure_tracks_table
            _ensure_tracks_table(conn)
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute(
                    """SELECT track_code, label, keywords, grounding_terms, active, sort_order
                       FROM learning_tracks ORDER BY sort_order, track_code""")
                rows = [dict(r) for r in cur.fetchall()]
        return {"tracks": rows}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/admin/tracks")
def upsert_learning_track(body: TrackUpsert,
                          user: dict = Depends(require_persona("admin"))):
    """Create or update a learning track. Admin-only. This is the whole point of
    Tier-2 'dynamic': a new Adobe product becomes a data row, not a deploy. The
    reasoning agent picks it up within its cache TTL (no restart needed)."""
    code = (body.track_code or "").strip().lower()
    if not code or not body.label.strip():
        raise HTTPException(status_code=422, detail="track_code and label are required.")
    try:
        with get_db() as conn:
            with conn.cursor() as cur:
                # Table is created/seeded lazily by the agent; ensure it exists here too.
                from agents.reasoning import _ensure_tracks_table
                _ensure_tracks_table(conn)
                cur.execute(
                    """INSERT INTO learning_tracks
                         (track_code, label, keywords, grounding_terms, active, sort_order, updated_at)
                       VALUES (%s,%s,%s,%s,%s,%s,NOW())
                       ON CONFLICT (track_code) DO UPDATE SET
                         label=EXCLUDED.label, keywords=EXCLUDED.keywords,
                         grounding_terms=EXCLUDED.grounding_terms, active=EXCLUDED.active,
                         sort_order=EXCLUDED.sort_order, updated_at=NOW()""",
                    (code, body.label.strip(), json.dumps(body.keywords or []),
                     json.dumps(body.grounding_terms or []), body.active, body.sort_order))
        # Force the agent's cache to refresh so the change is visible immediately.
        try:
            from agents.reasoning import _load_track_config
            _load_track_config(force=True)
        except Exception:
            pass
        return {"ok": True, "track_code": code}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/admin/learner-feedback/{email}")
def get_learner_feedback(email: str, user: dict = Depends(require_persona("admin"))):
    """Mentor/admin view: what the reasoning agent's feedback-adaptation layer has
    learned from a learner's 👍/👎 history — the sample size, the up/down split,
    and the plain-English directive currently steering their answers (or why it's
    not adapting yet). Admin-only; read-only. Makes the previously-invisible
    adaptation logic auditable without DB access."""
    try:
        from agents.reasoning import get_learner_feedback_insight
        return get_learner_feedback_insight((email or "").lower())
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))



# ════════════════════════════════════════════════════════════════════════════════
# Thread-based chat history (ChatGPT / Claude style)
# ════════════════════════════════════════════════════════════════════════════════
# The flat conversation_messages table above keeps ONE thread per (member,module,
# mode) and trusts a client-supplied member_name — a learner can neither keep
# several separate chats nor is their history protected from another caller.
#
# This adds proper, per-user conversation THREADS:
#   - chat_conversations : one row per chat thread (title, mode, module, track)
#   - chat_messages      : messages belonging to a thread (role, content, metadata)
# Every endpoint is authenticated (Depends(get_current_user)) and OWNERSHIP-checked
# against the SERVER-SIDE session email — a caller can only ever see or mutate their
# own threads, closing the IDOR present in the flat endpoints above.

def _ensure_chat_tables():
    try:
        conn = psycopg2.connect(DATABASE_URL)
        conn.autocommit = True
        with conn.cursor() as c:
            c.execute("""
                CREATE TABLE IF NOT EXISTS chat_conversations (
                    id          SERIAL PRIMARY KEY,
                    user_email  VARCHAR(255) NOT NULL,
                    title       VARCHAR(200) NOT NULL DEFAULT 'New chat',
                    track       VARCHAR(60),
                    module      VARCHAR(255),
                    mode        VARCHAR(30) NOT NULL DEFAULT 'reasoning',
                    dashboard   VARCHAR(30),
                    archived    BOOLEAN DEFAULT FALSE,
                    created_at  TIMESTAMP DEFAULT NOW(),
                    updated_at  TIMESTAMP DEFAULT NOW()
                )
            """)
            # Backfill for databases created before the dashboard column existed.
            c.execute("ALTER TABLE chat_conversations ADD COLUMN IF NOT EXISTS dashboard VARCHAR(30)")
            c.execute("""
                CREATE TABLE IF NOT EXISTS chat_messages (
                    id               SERIAL PRIMARY KEY,
                    conversation_id  INTEGER NOT NULL REFERENCES chat_conversations(id) ON DELETE CASCADE,
                    role             VARCHAR(20) NOT NULL,
                    content          TEXT NOT NULL,
                    metadata         JSONB,
                    created_at       TIMESTAMP DEFAULT NOW()
                )
            """)
            c.execute("CREATE INDEX IF NOT EXISTS idx_chat_conv_user ON chat_conversations (user_email, updated_at DESC)")
            c.execute("CREATE INDEX IF NOT EXISTS idx_chat_msg_conv ON chat_messages (conversation_id, created_at)")
        conn.close()
    except Exception:
        pass


def _session_email(user: dict) -> str:
    """The authoritative learner identity for chat ownership: the session email."""
    email = (user.get("email") or "").lower()
    if not email:
        raise HTTPException(status_code=401, detail="Session has no email.")
    return email


def _assert_owns_conversation(cur, cid: int, email: str):
    """Raise 404 unless conversation `cid` exists AND belongs to `email`.
    404 (not 403) so a caller cannot even probe which conversation ids exist.
    Callers use either a plain cursor (tuple rows) or a RealDictCursor (dict-like
    rows) — handle both rather than assuming positional indexing everywhere."""
    cur.execute("SELECT user_email FROM chat_conversations WHERE id=%s", (cid,))
    row = cur.fetchone()
    if row is None:
        raise HTTPException(status_code=404, detail="Conversation not found.")
    owner = row["user_email"] if isinstance(row, dict) else row[0]
    if owner is None or owner.lower() != email:
        raise HTTPException(status_code=404, detail="Conversation not found.")


def _derive_title(text: str) -> str:
    """Make a short thread title from the first user message (fallback = 'New chat')."""
    t = " ".join((text or "").split())
    if not t:
        return "New chat"
    return (t[:60] + "…") if len(t) > 60 else t


class NewConversation(BaseModel):
    title:     Optional[str] = None
    track:     Optional[str] = None
    module:    Optional[str] = None
    mode:      str = "reasoning"
    # Which dashboard this thread belongs to ("new_joiner" / "experience"). Keeps
    # a learner's New-Joiner threads from leaking into their Experience dashboard.
    dashboard: Optional[str] = None


class AppendMessage(BaseModel):
    role:     str
    content:  str
    metadata: Optional[dict] = None


class EditMessage(BaseModel):
    content: str


class PatchConversation(BaseModel):
    title:    Optional[str] = None
    archived: Optional[bool] = None


@app.get("/api/chat/conversations")
def list_chat_conversations(include_archived: bool = False, dashboard: str = "",
                            user: dict = Depends(get_current_user)):
    """List the signed-in learner's chat threads, most-recently-updated first.
    Optional `dashboard` filter scopes threads to one dashboard (new_joiner /
    experience); legacy rows with a NULL dashboard are always included so
    pre-existing threads don't vanish after the column was added."""
    _ensure_chat_tables()
    email = _session_email(user)
    try:
        with get_db() as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute(
                    """SELECT c.id, c.title, c.track, c.module, c.mode, c.dashboard, c.archived,
                              c.created_at, c.updated_at,
                              (SELECT COUNT(*) FROM chat_messages m WHERE m.conversation_id=c.id) AS message_count
                       FROM chat_conversations c
                       WHERE c.user_email=%s AND (%s OR c.archived=FALSE)
                         AND (%s='' OR c.dashboard=%s OR c.dashboard IS NULL)
                       ORDER BY c.updated_at DESC LIMIT 200""",
                    (email, include_archived, dashboard, dashboard),
                )
                rows = cur.fetchall()
        for r in rows:
            r["created_at"] = str(r["created_at"])
            r["updated_at"] = str(r["updated_at"])
        return {"conversations": [dict(r) for r in rows]}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/chat/conversations")
def create_chat_conversation(body: NewConversation,
                             user: dict = Depends(get_current_user)):
    """Create a new chat thread owned by the signed-in learner."""
    _ensure_chat_tables()
    email = _session_email(user)
    title = (body.title or "New chat").strip()[:200] or "New chat"
    try:
        with get_db() as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute(
                    """INSERT INTO chat_conversations (user_email, title, track, module, mode, dashboard)
                       VALUES (%s,%s,%s,%s,%s,%s)
                       RETURNING id, title, track, module, mode, dashboard, archived, created_at, updated_at""",
                    (email, title, body.track, body.module, body.mode or "reasoning", body.dashboard),
                )
                row = cur.fetchone()
        row["created_at"] = str(row["created_at"])
        row["updated_at"] = str(row["updated_at"])
        row["message_count"] = 0
        return dict(row)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/chat/conversations/{cid}")
def get_chat_conversation(cid: int, user: dict = Depends(get_current_user)):
    """Return one thread's full message list (ownership-checked)."""
    _ensure_chat_tables()
    email = _session_email(user)
    try:
        with get_db() as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                _assert_owns_conversation(cur, cid, email)
                cur.execute(
                    """SELECT id, role, content, metadata, created_at
                       FROM chat_messages WHERE conversation_id=%s
                       ORDER BY created_at ASC, id ASC""",
                    (cid,),
                )
                rows = cur.fetchall()
        for r in rows:
            r["created_at"] = str(r["created_at"])
        return {"id": cid, "messages": [dict(r) for r in rows]}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/chat/conversations/{cid}/messages")
def append_chat_message(cid: int, body: AppendMessage,
                        user: dict = Depends(get_current_user)):
    """Append a message to a thread (ownership-checked). Bumps updated_at and, if the
    thread is still the default 'New chat', auto-titles it from the first user turn."""
    _ensure_chat_tables()
    email = _session_email(user)
    if body.role not in ("user", "assistant", "system"):
        raise HTTPException(status_code=422, detail="Invalid role.")
    try:
        with get_db() as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                _assert_owns_conversation(cur, cid, email)
                cur.execute(
                    """INSERT INTO chat_messages (conversation_id, role, content, metadata)
                       VALUES (%s,%s,%s,%s) RETURNING id, created_at""",
                    (cid, body.role, body.content,
                     json.dumps(body.metadata) if body.metadata is not None else None),
                )
                new_row = cur.fetchone()
                # Auto-title from the first user message if still the default.
                if body.role == "user":
                    cur.execute(
                        """UPDATE chat_conversations
                           SET title=%s, updated_at=NOW()
                           WHERE id=%s AND (title='New chat' OR title IS NULL OR title='')""",
                        (_derive_title(body.content), cid),
                    )
                cur.execute("UPDATE chat_conversations SET updated_at=NOW() WHERE id=%s", (cid,))
        return {"ok": True, "id": new_row["id"], "created_at": str(new_row["created_at"])}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.patch("/api/chat/conversations/{cid}")
def patch_chat_conversation(cid: int, body: PatchConversation,
                            user: dict = Depends(get_current_user)):
    """Rename or archive a thread (ownership-checked)."""
    _ensure_chat_tables()
    email = _session_email(user)
    try:
        with get_db() as conn:
            with conn.cursor() as cur:
                _assert_owns_conversation(cur, cid, email)
                if body.title is not None:
                    cur.execute("UPDATE chat_conversations SET title=%s, updated_at=NOW() WHERE id=%s",
                                (body.title.strip()[:200] or "New chat", cid))
                if body.archived is not None:
                    cur.execute("UPDATE chat_conversations SET archived=%s, updated_at=NOW() WHERE id=%s",
                                (body.archived, cid))
        return {"ok": True}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.delete("/api/chat/conversations/{cid}")
def delete_chat_conversation(cid: int, user: dict = Depends(get_current_user)):
    """Delete a thread and its messages (ownership-checked; cascade drops messages)."""
    _ensure_chat_tables()
    email = _session_email(user)
    try:
        with get_db() as conn:
            with conn.cursor() as cur:
                _assert_owns_conversation(cur, cid, email)
                cur.execute("DELETE FROM chat_conversations WHERE id=%s", (cid,))
        return {"ok": True}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.delete("/api/chat/conversations/{cid}/messages/last")
def delete_last_chat_message(cid: int, user: dict = Depends(get_current_user)):
    """Delete the most-recent message in a thread (ownership-checked). Used by the
    frontend 'Regenerate' action to drop the previous assistant answer before
    streaming a fresh one, so the thread doesn't accumulate duplicate turns."""
    _ensure_chat_tables()
    email = _session_email(user)
    try:
        with get_db() as conn:
            with conn.cursor() as cur:
                _assert_owns_conversation(cur, cid, email)
                cur.execute(
                    """DELETE FROM chat_messages WHERE id = (
                           SELECT id FROM chat_messages WHERE conversation_id=%s
                           ORDER BY created_at DESC, id DESC LIMIT 1)""",
                    (cid,),
                )
                deleted = cur.rowcount
                cur.execute("UPDATE chat_conversations SET updated_at=NOW() WHERE id=%s", (cid,))
        return {"ok": True, "deleted": deleted}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.patch("/api/chat/messages/{mid}")
def edit_chat_message(mid: int, body: EditMessage,
                      user: dict = Depends(get_current_user)):
    """Edit a previously-sent USER message and truncate the conversation after it.

    Editing a message forks the conversation from that point, so every message
    that came after it (the old assistant answer and any later turns) is deleted
    — the frontend then re-streams a fresh answer for the edited text. Only the
    learner's own user messages can be edited (ownership-checked); editing an
    assistant message is rejected, since that would let a learner rewrite the
    tutor's words. Returns how many trailing messages were removed."""
    _ensure_chat_tables()
    email = _session_email(user)
    new_content = (body.content or "").strip()
    if not new_content:
        raise HTTPException(status_code=422, detail="content cannot be empty.")
    try:
        with get_db() as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute(
                    """SELECT c.id AS cid, c.user_email AS owner, m.role AS role,
                              m.created_at AS created_at
                       FROM chat_messages m
                       JOIN chat_conversations c ON c.id = m.conversation_id
                       WHERE m.id=%s""",
                    (mid,),
                )
                row = cur.fetchone()
                if not row or (row["owner"] or "").lower() != email:
                    raise HTTPException(status_code=404, detail="Message not found.")
                if row["role"] != "user":
                    raise HTTPException(status_code=400, detail="Only user messages can be edited.")
                cid = row["cid"]
                # Update the edited message's content.
                cur.execute("UPDATE chat_messages SET content=%s WHERE id=%s", (new_content, mid))
                # Delete everything after it (created strictly later, or same
                # timestamp but a higher id — matches the ASC ordering used on load).
                cur.execute(
                    """DELETE FROM chat_messages
                       WHERE conversation_id=%s
                         AND (created_at > %s OR (created_at = %s AND id > %s))""",
                    (cid, row["created_at"], row["created_at"], mid),
                )
                removed = cur.rowcount
                cur.execute("UPDATE chat_conversations SET updated_at=NOW() WHERE id=%s", (cid,))
        return {"ok": True, "removed_after": removed}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


class MessageFeedback(BaseModel):
    rating: Optional[str] = None   # "up" | "down" | null to clear


@app.post("/api/chat/messages/{mid}/feedback")
def set_chat_message_feedback(mid: int, body: MessageFeedback,
                              user: dict = Depends(get_current_user)):
    """Record a 👍/👎 rating on an assistant message by merging it into that
    message's JSONB metadata (key 'feedback'). Ownership is enforced by joining to
    the parent conversation — a caller can only rate messages in their own threads.
    Passing rating=null clears a previous rating (toggle off)."""
    _ensure_chat_tables()
    email = _session_email(user)
    if body.rating not in (None, "up", "down"):
        raise HTTPException(status_code=422, detail="rating must be 'up', 'down', or null.")
    try:
        with get_db() as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute(
                    """SELECT c.user_email AS owner, m.metadata AS metadata
                       FROM chat_messages m
                       JOIN chat_conversations c ON c.id = m.conversation_id
                       WHERE m.id=%s""",
                    (mid,),
                )
                row = cur.fetchone()
                if not row or (row["owner"] or "").lower() != email:
                    raise HTTPException(status_code=404, detail="Message not found.")
                meta = row["metadata"] if isinstance(row["metadata"], dict) else {}
                if body.rating is None:
                    meta.pop("feedback", None)
                else:
                    meta["feedback"] = body.rating
                cur.execute("UPDATE chat_messages SET metadata=%s WHERE id=%s",
                            (json.dumps(meta), mid))
        return {"ok": True, "rating": body.rating}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ── Chat history full-text search ────────────────────────────────────────────────
@app.get("/api/chat/search")
def search_chat_history(q: str, mode: str = "", limit: int = 30,
                        user: dict = Depends(get_current_user)):
    """Full-text-ish search across the signed-in learner's own chat messages.
    Returns matching messages with the parent thread + a short snippet, so the UI
    can jump straight to the thread. Ownership is enforced by filtering on the
    session email — a learner can only ever search their own threads (no IDOR)."""
    _ensure_chat_tables()
    email = _session_email(user)
    term = (q or "").strip()
    if len(term) < 2:
        return {"results": []}
    like = f"%{term}%"
    limit = max(1, min(int(limit or 30), 100))
    try:
        with get_db() as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                # mode filter is optional (Socratic vs Reasoning threads are distinct).
                cur.execute(
                    """SELECT m.id AS message_id, m.role, m.content, m.created_at,
                              c.id AS conversation_id, c.title, c.mode, c.module
                       FROM chat_messages m
                       JOIN chat_conversations c ON c.id = m.conversation_id
                       WHERE c.user_email=%s
                         AND (%s='' OR c.mode=%s)
                         AND m.content ILIKE %s
                       ORDER BY m.created_at DESC
                       LIMIT %s""",
                    (email, mode, mode, like, limit),
                )
                rows = cur.fetchall()
        results = []
        tlow = term.lower()
        for r in rows:
            content = r["content"] or ""
            idx = content.lower().find(tlow)
            start = max(0, idx - 40)
            end = min(len(content), idx + len(term) + 40)
            snippet = ("…" if start > 0 else "") + content[start:end].strip() + ("…" if end < len(content) else "")
            results.append({
                "conversation_id": r["conversation_id"],
                "title": r["title"],
                "mode": r["mode"],
                "module": r["module"],
                "role": r["role"],
                "snippet": snippet,
                "created_at": str(r["created_at"]),
            })
        return {"results": results}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ── Admin: bulk curriculum import from Excel template ────────────────────────────
def _ensure_curriculum_upsert_index():
    """A UNIQUE index on (track, module_id, topic_order) is what makes the ON
    CONFLICT upsert below work. Created lazily; harmless if it already exists."""
    try:
        conn = psycopg2.connect(DATABASE_URL)
        conn.autocommit = True
        with conn.cursor() as c:
            c.execute("""CREATE UNIQUE INDEX IF NOT EXISTS uq_curriculum_topic
                         ON curriculum_topics (track, module_id, topic_order)""")
        conn.close()
    except Exception:
        pass


@app.post("/api/admin/curriculum/import")
async def admin_curriculum_import(file: UploadFile = File(...),
                                  _user: dict = Depends(require_persona("admin"))):
    """Admin-only. Upload the curriculum Excel template (sheet 'Curriculum', or
    the first sheet whose header row contains track/module_id/topic_order/title).
    Each valid row is upserted into curriculum_topics on (track, module_id,
    topic_order). Returns per-row validation errors (naming the row number) plus
    inserted/updated counts, so a partly-bad file reports exactly what to fix
    rather than silently dropping rows."""
    import openpyxl, io
    contents = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
    except Exception as ex:
        raise HTTPException(400, f"Could not open Excel file: {ex}")

    # Prefer a sheet literally named 'Curriculum'; else the first sheet that has
    # the required headers (so the instructions sheet in the template is skipped).
    def _header_index(ws):
        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            vals = [str(v or "").strip().lower() for v in row]
            if {"track", "module_id", "topic_order", "title"}.issubset(set(vals)):
                return i, vals
        return None, None

    ws = None
    header_idx = headers = None
    if "Curriculum" in wb.sheetnames:
        ws = wb["Curriculum"]
        header_idx, headers = _header_index(ws)
    if header_idx is None:
        for cand in wb.worksheets:
            hi, hv = _header_index(cand)
            if hi is not None:
                ws, header_idx, headers = cand, hi, hv
                break
    if header_idx is None:
        raise HTTPException(400, "No sheet has a header row containing track, module_id, "
                                 "topic_order and title. Use the provided template.")

    def col(row, key):
        if key in headers:
            v = row[headers.index(key)]
            if v is not None:
                s = str(v).strip()
                if s and s.lower() not in ("none", "n/a"):
                    return s
        return ""

    parsed, errors = [], []
    for excel_row, r in enumerate(ws.iter_rows(min_row=header_idx + 1, values_only=True), start=header_idx + 1):
        if not any(r):
            continue
        track = col(r, "track").lower()
        title = col(r, "title")
        mid_raw, ord_raw = col(r, "module_id"), col(r, "topic_order")
        row_errs = []
        if not track:
            row_errs.append("missing track")
        if not title:
            row_errs.append("missing title")
        try:
            module_id = int(float(mid_raw))
        except (ValueError, TypeError):
            module_id = None
            row_errs.append(f"module_id must be a number (got {mid_raw!r})")
        try:
            topic_order = int(float(ord_raw))
        except (ValueError, TypeError):
            topic_order = None
            row_errs.append(f"topic_order must be a number (got {ord_raw!r})")
        if row_errs:
            errors.append(f"Row {excel_row}: {'; '.join(row_errs)}")
            continue
        parsed.append({
            "track": track, "module_id": module_id, "topic_order": topic_order,
            "title": title, "objective": col(r, "objective"), "activity": col(r, "activity"),
            "output": col(r, "output"), "checkpoint": col(r, "checkpoint"),
            "video_title": col(r, "video_title"), "video_duration": col(r, "video_duration"),
            "el_url": col(r, "el_url"),
        })

    if not parsed:
        raise HTTPException(400, "No valid rows found." + (" Errors: " + "; ".join(errors[:10]) if errors else ""))

    _ensure_curriculum_upsert_index()
    inserted = updated = 0
    tracks_seen = set()
    conn = psycopg2.connect(DATABASE_URL)
    try:
        with conn:
            c = conn.cursor()
            for row in parsed:
                tracks_seen.add(row["track"])
                c.execute("""
                    INSERT INTO curriculum_topics
                        (track, module_id, topic_order, title, objective, activity,
                         output, checkpoint, video_title, video_duration, el_url)
                    VALUES (%(track)s,%(module_id)s,%(topic_order)s,%(title)s,%(objective)s,
                            %(activity)s,%(output)s,%(checkpoint)s,%(video_title)s,
                            %(video_duration)s,%(el_url)s)
                    ON CONFLICT (track, module_id, topic_order) DO UPDATE SET
                        title=EXCLUDED.title, objective=EXCLUDED.objective,
                        activity=EXCLUDED.activity, output=EXCLUDED.output,
                        checkpoint=EXCLUDED.checkpoint, video_title=EXCLUDED.video_title,
                        video_duration=EXCLUDED.video_duration, el_url=EXCLUDED.el_url
                    RETURNING (xmax = 0) AS is_insert
                """, row)
                res = c.fetchone()
                if res and res[0]:
                    inserted += 1
                else:
                    updated += 1
    except Exception as ex:
        raise HTTPException(500, f"Import failed during insert: {ex}")
    finally:
        conn.close()

    # Soft warning: curriculum rows for a track with no learning_tracks entry will
    # still work for the lesson UI, but Study Aid/Reasoning won't have a label for
    # it — surface that rather than failing.
    unknown_tracks = []
    try:
        from agents.reasoning import get_track_codes
        known = set(get_track_codes())
        unknown_tracks = sorted(t for t in tracks_seen if t not in known)
    except Exception:
        pass

    return {"ok": True, "inserted": inserted, "updated": updated,
            "total_valid": len(parsed), "row_errors": errors,
            "tracks": sorted(tracks_seen),
            "warning": (f"These tracks have curriculum but no Learning Tracks entry: "
                        f"{', '.join(unknown_tracks)}. Add them in Reasoning Config → Learning Tracks "
                        f"so Study Aid/Reasoning label them correctly." if unknown_tracks else None)}

# ── Member project view (individual team member sees their own assignments) ────

@app.get("/api/projects/my")
def my_projects(email: str = None, member_name: str = None):
    """Return projects a team member is assigned to, with their allocation details."""
    if not email and not member_name:
        raise HTTPException(400, "Provide email or member_name")
    conn = psycopg2.connect(DATABASE_URL)
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as c:
            if email:
                c.execute("""SELECT p.*, pm.hrs_per_week, pm.role_on_project, pm.id AS member_link_id
                             FROM projects p JOIN project_members pm ON pm.project_id=p.id
                             WHERE LOWER(pm.member_email)=LOWER(%s)
                             ORDER BY p.updated_at DESC""", (email.strip(),))
            else:
                c.execute("""SELECT p.*, pm.hrs_per_week, pm.role_on_project, pm.id AS member_link_id
                             FROM projects p JOIN project_members pm ON pm.project_id=p.id
                             WHERE LOWER(pm.member_name)=LOWER(%s)
                             ORDER BY p.updated_at DESC""", (member_name.strip(),))
            rows = c.fetchall()
    finally:
        conn.close()
    result = []
    for r in rows:
        d = dict(r)
        for k in ("created_at","updated_at","start_date","end_date"):
            if d.get(k): d[k] = str(d[k])
        result.append(d)
    return {"projects": result}


@app.put("/api/projects/{project_id}/my-update")
def member_update_project(project_id: int, body: dict = Body(...)):
    """Team member updates their own allocation details and project notes."""
    email = (body.get("email") or "").strip().lower()
    member_name = (body.get("member_name") or "").strip()
    if not email and not member_name:
        raise HTTPException(400, "email or member_name required")
    conn = psycopg2.connect(DATABASE_URL)
    try:
        with conn:
            with conn.cursor() as c:
                # Update member's own allocation — match by email first (the
                # normal case), falling back to a name match for members
                # imported from a tracker with no resolvable email (the
                # frontend passes their display name via `email` in that case).
                if "hrs_per_week" in body or "role_on_project" in body:
                    updated = 0
                    if email:
                        c.execute("""UPDATE project_members
                                     SET hrs_per_week=COALESCE(%s, hrs_per_week),
                                         role_on_project=COALESCE(%s, role_on_project)
                                     WHERE project_id=%s AND LOWER(member_email)=LOWER(%s)""",
                                  (body.get("hrs_per_week"), body.get("role_on_project"),
                                   project_id, email))
                        updated = c.rowcount
                    fallback_name = member_name or email
                    if updated == 0 and fallback_name:
                        c.execute("""UPDATE project_members
                                     SET hrs_per_week=COALESCE(%s, hrs_per_week),
                                         role_on_project=COALESCE(%s, role_on_project)
                                     WHERE project_id=%s AND LOWER(member_name)=LOWER(%s)""",
                                  (body.get("hrs_per_week"), body.get("role_on_project"),
                                   project_id, fallback_name))
                # Update project-level fields member is allowed to change.
                # weekly_comments is deliberately NOT here — it's append-only
                # history now (see /api/projects/{id}/updates), not a field
                # any single call can silently overwrite.
                proj_fields, proj_params = [], []
                for col in ("high_level_notes","health_status","stage","status"):
                    if col in body:
                        proj_fields.append(f"{col}=%s")
                        proj_params.append(body[col])
                if proj_fields:
                    proj_params.append(project_id)
                    c.execute(f"UPDATE projects SET {','.join(proj_fields)},updated_at=NOW() WHERE id=%s",
                              proj_params)
    finally:
        conn.close()
    return {"ok": True}


@app.post("/api/projects/{project_id}/updates")
def add_project_update(project_id: int, body: dict = Body(...)):
    """Post a new timestamped weekly note for a project — append-only, never
    overwrites a previous entry. Also mirrors the text into
    projects.weekly_comments as the "latest" snapshot, for views (Project
    Board, Team Weekly Tracker table) that only show the current note."""
    text = (body.get("update_text") or "").strip()
    if not text:
        raise HTTPException(400, "update_text required")
    member_email = (body.get("member_email") or "").strip()
    member_name  = (body.get("member_name") or "").strip()
    conn = psycopg2.connect(DATABASE_URL)
    try:
        with conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as c:
                c.execute("""INSERT INTO project_weekly_updates (project_id,member_email,member_name,update_text)
                             VALUES (%s,%s,%s,%s) RETURNING id, created_at""",
                          (project_id, member_email, member_name, text))
                row = c.fetchone()
                c.execute("UPDATE projects SET weekly_comments=%s, updated_at=NOW() WHERE id=%s",
                          (text, project_id))
    finally:
        conn.close()
    return {"ok": True, "id": row["id"], "created_at": str(row["created_at"])}


@app.get("/api/projects/{project_id}/updates")
def list_project_updates(project_id: int):
    """Full timestamped weekly-note history for one project, newest first."""
    conn = psycopg2.connect(DATABASE_URL)
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as c:
            c.execute("""SELECT id, member_name, member_email, update_text, created_at
                         FROM project_weekly_updates
                         WHERE project_id=%s ORDER BY created_at DESC""", (project_id,))
            rows = c.fetchall()
    finally:
        conn.close()
    result = []
    for r in rows:
        d = dict(r)
        d["created_at"] = str(d["created_at"])
        result.append(d)
    return {"updates": result}


# ── Initiatives & Milestones ───────────────────────────────────────────────────

@app.get("/api/initiatives/my")
def my_initiatives(email: str = None, member_name: str = None):
    conn = psycopg2.connect(DATABASE_URL)
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as c:
            if email:
                c.execute("""SELECT * FROM member_initiatives WHERE LOWER(member_email)=LOWER(%s)
                             ORDER BY updated_at DESC""", (email.strip(),))
            else:
                c.execute("""SELECT * FROM member_initiatives WHERE LOWER(member_name)=LOWER(%s)
                             ORDER BY updated_at DESC""", (member_name.strip(),))
            rows = c.fetchall()
    finally:
        conn.close()
    result = [dict(r) for r in rows]
    for r in result:
        for k in ("date_logged","created_at","updated_at"):
            if r.get(k): r[k] = str(r[k])
    return {"initiatives": result}

@app.post("/api/initiatives/update")
def update_initiative(body: dict = Body(...)):
    init_id = body.get("id")
    update_text = (body.get("latest_update") or "").strip()
    if not init_id:
        raise HTTPException(400, "id required")
    conn = psycopg2.connect(DATABASE_URL)
    try:
        with conn:
            with conn.cursor() as c:
                c.execute("UPDATE member_initiatives SET latest_update=%s, updated_at=NOW() WHERE id=%s",
                          (update_text, init_id))
    finally:
        conn.close()
    return {"ok": True}

@app.post("/api/initiatives/add")
def add_initiative(body: dict = Body(...)):
    conn = psycopg2.connect(DATABASE_URL)
    try:
        with conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as c:
                c.execute("""INSERT INTO member_initiatives
                             (member_name, member_email, initiative, latest_update, manager_email)
                             VALUES (%s,%s,%s,%s,%s) RETURNING id""",
                    (body.get("member_name",""), body.get("member_email",""),
                     body.get("initiative",""), body.get("latest_update",""),
                     body.get("manager_email","")))
                new_id = c.fetchone()["id"]
    finally:
        conn.close()
    return {"ok": True, "id": new_id}

@app.get("/api/milestones/my")
def my_milestones(email: str = None, member_name: str = None):
    conn = psycopg2.connect(DATABASE_URL)
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as c:
            if email:
                c.execute("""SELECT * FROM member_milestones WHERE LOWER(member_email)=LOWER(%s)
                             ORDER BY milestone_date DESC NULLS LAST, created_at DESC""", (email.strip(),))
            else:
                c.execute("""SELECT * FROM member_milestones WHERE LOWER(member_name)=LOWER(%s)
                             ORDER BY milestone_date DESC NULLS LAST, created_at DESC""", (member_name.strip(),))
            rows = c.fetchall()
    finally:
        conn.close()
    result = [dict(r) for r in rows]
    for r in result:
        for k in ("milestone_date","created_at"):
            if r.get(k): r[k] = str(r[k])
    return {"milestones": result}

@app.get("/api/initiatives/team")
def team_initiatives(manager_email: str = None, manager_name: str = None):
    conn = psycopg2.connect(DATABASE_URL)
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as c:
            if manager_email:
                c.execute("""SELECT * FROM member_initiatives WHERE LOWER(manager_email)=LOWER(%s)
                             ORDER BY member_name, updated_at DESC""", (manager_email.strip(),))
            else:
                c.execute("SELECT * FROM member_initiatives ORDER BY member_name, updated_at DESC")
            rows = c.fetchall()
    finally:
        conn.close()
    result = [dict(r) for r in rows]
    for r in result:
        for k in ("date_logged","created_at","updated_at"):
            if r.get(k): r[k] = str(r[k])
    return {"initiatives": result}


# ── Initiative dated updates ────────────────────────────────────────────────────

@app.post("/api/initiatives/{initiative_id}/updates")
def add_initiative_update(initiative_id: int, body: dict = Body(...)):
    text = (body.get("update_text") or "").strip()
    if not text:
        raise HTTPException(400, "update_text required")
    conn = psycopg2.connect(DATABASE_URL)
    try:
        with conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as c:
                c.execute("INSERT INTO initiative_updates (initiative_id, update_text) VALUES (%s,%s) RETURNING id, TO_CHAR(created_at,'YYYY-MM-DD HH24:MI') AS created_at",
                          (initiative_id, text))
                row = dict(c.fetchone())
                # Also update the latest_update field on the initiative
                c.execute("UPDATE member_initiatives SET latest_update=%s, updated_at=NOW() WHERE id=%s",
                          (text, initiative_id))
    finally:
        conn.close()
    return {"ok": True, **row}

@app.get("/api/initiatives/{initiative_id}/updates")
def get_initiative_updates(initiative_id: int):
    conn = psycopg2.connect(DATABASE_URL)
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as c:
            c.execute("""SELECT id, update_text, TO_CHAR(created_at,'YYYY-MM-DD HH24:MI') AS created_at
                         FROM initiative_updates WHERE initiative_id=%s ORDER BY created_at DESC""",
                      (initiative_id,))
            rows = [dict(r) for r in c.fetchall()]
    finally:
        conn.close()
    return {"updates": rows}

@app.post("/api/milestones/add")
def add_milestone(body: dict = Body(...)):
    conn = psycopg2.connect(DATABASE_URL)
    try:
        with conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as c:
                c.execute("""INSERT INTO member_milestones (member_name, member_email, milestone_date, note, project_name, manager_email)
                             VALUES (%s,%s,%s,%s,%s,%s) RETURNING id""",
                    (body.get("member_name",""), body.get("member_email",""),
                     body.get("milestone_date") or None, body.get("note",""),
                     body.get("project_name",""), body.get("manager_email","")))
                new_id = c.fetchone()["id"]
    finally:
        conn.close()
    return {"ok": True, "id": new_id}

@app.get("/api/projects/my-client")
def my_client_projects(email: str = None, member_name: str = None):
    """Return only client (non-initiative) projects for a member."""
    if not email and not member_name:
        raise HTTPException(400, "Provide email or member_name")
    conn = psycopg2.connect(DATABASE_URL)
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as c:
            if email:
                c.execute("""SELECT p.*, pm.hrs_per_week, pm.role_on_project, pm.id AS member_link_id
                             FROM projects p JOIN project_members pm ON pm.project_id=p.id
                             WHERE LOWER(pm.member_email)=LOWER(%s)
                             AND COALESCE(p.is_initiative,FALSE)=FALSE
                             ORDER BY p.updated_at DESC""", (email.strip(),))
            else:
                c.execute("""SELECT p.*, pm.hrs_per_week, pm.role_on_project, pm.id AS member_link_id
                             FROM projects p JOIN project_members pm ON pm.project_id=p.id
                             WHERE LOWER(pm.member_name)=LOWER(%s)
                             AND COALESCE(p.is_initiative,FALSE)=FALSE
                             ORDER BY p.updated_at DESC""", (member_name.strip(),))
            rows = c.fetchall()
    finally:
        conn.close()
    result = []
    for r in rows:
        d = dict(r)
        for k in ("created_at","updated_at","start_date","end_date"):
            if d.get(k): d[k] = str(d[k])
        result.append(d)
    return {"projects": result}


@app.get("/api/projects/tracker-table")
def projects_tracker_table(manager_email: str = None, manager_name: str = None):
    """Every imported project row, one per (project, member) pair, for every
    member reporting to this manager — the single flat table backing Team
    Weekly Tracker. Member→project→manager mapping falls straight out of the
    tracker import (project_members.member_name/email + projects.manager_email),
    no separate HR-roster upload required."""
    if not manager_email and not manager_name:
        raise HTTPException(400, "Provide manager_email or manager_name")
    conn = psycopg2.connect(DATABASE_URL)
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as c:
            clauses, params = [], []
            if manager_email:
                clauses.append("LOWER(p.manager_email)=LOWER(%s)")
                params.append(manager_email.strip())
            if manager_name:
                clauses.append("LOWER(p.manager_email)=LOWER(%s)")
                params.append(manager_name.strip())
            where = " OR ".join(clauses)
            c.execute(f"""SELECT p.*, pm.member_name, pm.member_email, pm.hrs_per_week, pm.id AS member_link_id
                          FROM project_members pm JOIN projects p ON p.id=pm.project_id
                          WHERE ({where}) AND COALESCE(p.is_initiative,FALSE)=FALSE
                          ORDER BY pm.member_name, p.updated_at DESC""", params)
            rows = c.fetchall()
    finally:
        conn.close()

    result = []
    for r in rows:
        d = dict(r)
        for k in ("created_at","updated_at","start_date","end_date"):
            if d.get(k): d[k] = str(d[k])
        result.append(d)
    return {"rows": result, "total": len(result)}


@app.get("/api/admin/tracker/summary")
def tracker_summary():
    """Persistent snapshot of everything currently imported — shown on the
    Tracker Import page every time it's opened (not just right after an
    upload), the same way User Provisioning always shows the current roster
    regardless of when it was last uploaded."""
    conn = psycopg2.connect(DATABASE_URL)
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as c:
            c.execute("SELECT COUNT(*) AS n FROM projects WHERE COALESCE(is_initiative,FALSE)=FALSE")
            total_projects = c.fetchone()["n"]
            c.execute("SELECT COUNT(DISTINCT LOWER(member_email)) AS n FROM project_members")
            total_members = c.fetchone()["n"]
            c.execute("""SELECT p.manager_email, COUNT(DISTINCT p.id) AS project_count,
                                COUNT(DISTINCT LOWER(pm.member_email)) AS member_count,
                                MAX(p.updated_at) AS last_updated
                         FROM projects p LEFT JOIN project_members pm ON pm.project_id=p.id
                         WHERE COALESCE(p.is_initiative,FALSE)=FALSE
                         GROUP BY p.manager_email
                         ORDER BY project_count DESC""")
            by_manager = []
            for r in c.fetchall():
                d = dict(r)
                if d.get("last_updated"): d["last_updated"] = str(d["last_updated"])
                by_manager.append(d)
    finally:
        conn.close()
    return {"total_projects": total_projects, "total_members": total_members, "by_manager": by_manager}


@app.delete("/api/admin/projects/wipe")
def wipe_projects(confirm: str = "", user: dict = Depends(require_persona("admin"))):
    """Hard reset — permanently deletes every imported project, member link,
    issue, and the legacy allocations/initiatives/milestones tables tied to
    tracker imports. Irreversible; requires ?confirm=WIPE."""
    if confirm != "WIPE":
        raise HTTPException(400, 'Pass ?confirm=WIPE to actually delete everything.')
    conn = psycopg2.connect(DATABASE_URL)
    try:
        with conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as c:
                c.execute("SELECT COUNT(*) AS n FROM projects")
                n_projects = c.fetchone()["n"]
                c.execute("SELECT COUNT(*) AS n FROM project_members")
                n_links = c.fetchone()["n"]
                # project_members/project_issues cascade on projects delete,
                # but clear them explicitly first so counts above stay accurate
                # even if a future schema change drops the ON DELETE CASCADE.
                c.execute("DELETE FROM project_issues")
                c.execute("DELETE FROM project_members")
                c.execute("DELETE FROM projects")
                c.execute("DELETE FROM project_allocations")
                c.execute("DELETE FROM allocation_updates")
                c.execute("DELETE FROM member_initiatives")
                c.execute("DELETE FROM initiative_updates")
                c.execute("DELETE FROM member_milestones")
    finally:
        conn.close()
    return {"ok": True, "deleted_projects": n_projects, "deleted_links": n_links}


@app.post("/api/admin/projects/dedupe")
def dedupe_projects(dry_run: bool = True, user: dict = Depends(require_persona("admin"))):
    """Cleanup for duplicate projects/links from earlier tracker imports —
    e.g. before member-email resolution was fixed, the same person could get
    linked to a project twice (different guessed email each time), or the
    same client project could get created twice under slightly different
    Project IDs. Defaults to a dry run that only reports what it *would*
    merge; call with ?dry_run=false to actually apply.

    1. Duplicate PROJECT rows: same manager_email + same title → merge into
       the most recently updated one, re-pointing project_members/project_issues.
    2. Duplicate LINK rows: same project_id + same member_name (different
       member_email) → keep the most recent link, taking the max hrs_per_week.
    """
    conn = psycopg2.connect(DATABASE_URL)
    report = {"dry_run": dry_run, "duplicate_project_groups": [], "duplicate_link_groups": []}
    try:
        with conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as c:
                # 1) Duplicate project rows
                c.execute("""SELECT id, manager_email, title, updated_at FROM projects
                             WHERE COALESCE(is_initiative,FALSE)=FALSE
                             ORDER BY LOWER(manager_email), LOWER(title), updated_at DESC NULLS LAST, id DESC""")
                groups = {}
                for r in c.fetchall():
                    key = ((r["manager_email"] or "").lower(), (r["title"] or "").strip().lower())
                    groups.setdefault(key, []).append(r["id"])

                for (mgr, title), ids in groups.items():
                    if len(ids) < 2:
                        continue
                    keep_id, dup_ids = ids[0], ids[1:]
                    report["duplicate_project_groups"].append(
                        {"manager_email": mgr, "title": title, "keep_id": keep_id, "merged_ids": dup_ids})
                    if dry_run:
                        continue
                    for dup_id in dup_ids:
                        c.execute("SELECT member_email FROM project_members WHERE project_id=%s", (dup_id,))
                        for row in c.fetchall():
                            c.execute("""SELECT 1 FROM project_members
                                         WHERE project_id=%s AND LOWER(member_email)=LOWER(%s)""",
                                      (keep_id, row["member_email"]))
                            if c.fetchone():
                                c.execute("DELETE FROM project_members WHERE project_id=%s AND member_email=%s",
                                          (dup_id, row["member_email"]))
                            else:
                                c.execute("""UPDATE project_members SET project_id=%s
                                             WHERE project_id=%s AND member_email=%s""",
                                          (keep_id, dup_id, row["member_email"]))
                        c.execute("UPDATE project_issues SET project_id=%s WHERE project_id=%s", (keep_id, dup_id))
                        c.execute("DELETE FROM projects WHERE id=%s", (dup_id,))

                # 2) Duplicate link rows (same project, same person, different email)
                c.execute("""SELECT id, project_id, member_name, member_email, hrs_per_week
                             FROM project_members
                             ORDER BY project_id, LOWER(member_name), added_at DESC NULLS LAST, id DESC""")
                groups2 = {}
                for r in c.fetchall():
                    key = (r["project_id"], (r["member_name"] or "").strip().lower())
                    groups2.setdefault(key, []).append(dict(r))

                for (proj_id, mname), rows2 in groups2.items():
                    if len(rows2) < 2:
                        continue
                    keep, dups = rows2[0], rows2[1:]
                    report["duplicate_link_groups"].append({
                        "project_id": proj_id, "member_name": mname,
                        "keep_email": keep["member_email"],
                        "removed_emails": [d["member_email"] for d in dups],
                    })
                    if dry_run:
                        continue
                    best_hrs = max([float(keep["hrs_per_week"] or 0)] + [float(d["hrs_per_week"] or 0) for d in dups])
                    if best_hrs > float(keep["hrs_per_week"] or 0):
                        c.execute("UPDATE project_members SET hrs_per_week=%s WHERE id=%s", (best_hrs, keep["id"]))
                    for d in dups:
                        c.execute("DELETE FROM project_members WHERE id=%s", (d["id"],))
    finally:
        conn.close()

    report["projects_merged"]  = sum(len(g["merged_ids"]) for g in report["duplicate_project_groups"])
    report["links_removed"]    = sum(len(g["removed_emails"]) for g in report["duplicate_link_groups"])
    return report